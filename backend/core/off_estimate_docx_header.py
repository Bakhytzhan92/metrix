"""Разбор и вставка шапки заявки из Word (.docx) в Excel."""
from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

COL_COUNT = 5
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Ширины колонок для альбомного A4 (фиксированные, без растягивания).
TABLE_COL_WIDTHS = {1: 6, 2: 34, 3: 20, 4: 12, 5: 24}

LOGO_TARGET_WIDTH = 180
LOGO_TARGET_HEIGHT = 110
LOGO_ROW_HEIGHT = 36

DEFAULT_BKC_HEADER_DOCX = (
    Path(__file__).resolve().parent / "assets" / "off_estimate_header_bkc.docx"
)


@dataclass
class DocxLetterhead:
    left_lines: list[str]
    right_lines: list[str]
    address_left: str = ""
    address_right: str = ""
    image_bytes: bytes | None = None
    image_name: str | None = None


def company_uses_default_bkc_header(company) -> bool:
    prefix = (getattr(company, "inventory_prefix", "") or "").upper()
    name = (getattr(company, "name", "") or "").lower()
    return prefix == "BKC" or "baykaz" in name or "bkc" in name


def resolve_header_template_path(company) -> str | None:
    uploaded = getattr(company, "off_estimate_excel_header_template", None)
    if uploaded:
        try:
            return uploaded.path
        except (ValueError, AttributeError):
            return None
    if DEFAULT_BKC_HEADER_DOCX.is_file() and company_uses_default_bkc_header(company):
        return str(DEFAULT_BKC_HEADER_DOCX)
    return None


def _split_letterhead_address(address: str) -> tuple[str, str]:
    if not address.strip():
        return "", ""

    email = "too-bkc@mail.ru"
    match = re.search(r"[\w.+-]+@[\w.-]+\.(?:ru|com|kz|net|org)", address, re.IGNORECASE)
    if match:
        email = match.group(0)

    marker = "город Кызылорда"
    if marker in address:
        left_raw = address.split(marker, 1)[0].rstrip(", ").strip()
        left = f"{left_raw} Эл.Пошта:{email}" if left_raw else f"Эл.Пошта:{email}"
        right = f"город Кызылорда, ул Женис 88 Эл.почта:{email}"
        return left, right

    return address.strip(), ""


def header_template_is_docx(path: str) -> bool:
    return path.lower().endswith(".docx")


def _para_text_and_images(p, rels_map: dict[str, str]) -> tuple[str, list[str]]:
    texts: list[str] = []
    images: list[str] = []
    for t in p.iter(W + "t"):
        if t.text:
            texts.append(t.text)
        if t.tail:
            texts.append(t.tail)
    for blip in p.iter(A + "blip"):
        embed = blip.attrib.get(REL_NS + "embed")
        if embed and embed in rels_map:
            images.append(rels_map[embed])
    return "".join(texts).strip(), images


def parse_docx_letterhead(path: str) -> DocxLetterhead:
    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("word/document.xml"))
        rels_root = ET.fromstring(zf.read("word/_rels/document.xml.rels"))
        rels_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root
            if rel.attrib.get("Id") and rel.attrib.get("Target")
        }

        left_lines: list[str] = []
        right_lines: list[str] = []
        address_raw = ""
        image_target: str | None = None

        body = root.find(W + "body")
        if body is not None:
            for child in body:
                tag = child.tag.split("}")[-1]
                if tag == "tbl":
                    for tr in child.iter(W + "tr"):
                        cells = list(tr.iter(W + "tc"))
                        if len(cells) >= 3:
                            left_lines = _cell_lines(cells[0], rels_map)
                            mid_lines, mid_images = _cell_content(cells[1], rels_map)
                            right_lines = _cell_lines(cells[2], rels_map)
                            if mid_images:
                                image_target = mid_images[0]
                            elif not image_target:
                                image_target = _first_image_in_cell(cells[1], rels_map)
                elif tag == "p":
                    text, imgs = _para_text_and_images(child, rels_map)
                    if text:
                        address_raw = text
                    if not image_target and imgs:
                        image_target = imgs[0]

        image_bytes = None
        image_name = None
        if image_target:
            media_path = (
                image_target
                if image_target.startswith("word/")
                else f"word/{image_target.lstrip('/')}"
            )
            if media_path in zf.namelist():
                image_bytes = zf.read(media_path)
                image_name = Path(media_path).name

    address_left, address_right = _split_letterhead_address(address_raw)

    return DocxLetterhead(
        left_lines=left_lines,
        right_lines=right_lines,
        address_left=address_left,
        address_right=address_right,
        image_bytes=image_bytes,
        image_name=image_name,
    )


def _cell_lines(tc, rels_map: dict[str, str]) -> list[str]:
    lines, _ = _cell_content(tc, rels_map)
    return lines


def _cell_content(tc, rels_map: dict[str, str]) -> tuple[list[str], list[str]]:
    lines: list[str] = []
    images: list[str] = []
    for p in tc.iter(W + "p"):
        text, imgs = _para_text_and_images(p, rels_map)
        if text:
            lines.append(text)
        images.extend(imgs)
    return lines, images


def _first_image_in_cell(tc, rels_map: dict[str, str]) -> str | None:
    for blip in tc.iter(A + "blip"):
        embed = blip.attrib.get(REL_NS + "embed")
        if embed and embed in rels_map:
            return rels_map[embed]
    return None


def write_docx_letterhead(ws: Worksheet, path: str, start_row: int) -> int:
    data = parse_docx_letterhead(path)

    row = start_row
    line_count = max(len(data.left_lines), len(data.right_lines), 1)

    for i in range(line_count):
        r = row + i
        ws.row_dimensions[r].height = LOGO_ROW_HEIGHT
        left = data.left_lines[i] if i < len(data.left_lines) else ""
        right = data.right_lines[i] if i < len(data.right_lines) else ""

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        left_cell = ws.cell(row=r, column=1, value=left)
        left_cell.font = Font(size=10)
        left_cell.alignment = LEFT

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        right_cell = ws.cell(row=r, column=4, value=right)
        right_cell.font = Font(size=10)
        right_cell.alignment = RIGHT

    if line_count > 1:
        ws.merge_cells(
            start_row=row, start_column=3, end_row=row + line_count - 1, end_column=3
        )
    ws.cell(row=row, column=3).alignment = CENTER

    if data.image_bytes:
        img = XLImage(BytesIO(data.image_bytes))
        scale = min(
            LOGO_TARGET_WIDTH / max(img.width, 1),
            LOGO_TARGET_HEIGHT / max(img.height, 1),
        )
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, f"C{row}")

    row += line_count

    if data.address_left or data.address_right:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        left_addr = ws.cell(row=row, column=1, value=data.address_left)
        left_addr.font = Font(size=9)
        left_addr.alignment = LEFT

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        right_addr = ws.cell(row=row, column=4, value=data.address_right)
        right_addr.font = Font(size=9)
        right_addr.alignment = RIGHT
        row += 1

    return row + 1
