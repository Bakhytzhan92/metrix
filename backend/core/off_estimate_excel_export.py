"""Экспорт одной заявки вне сметы в Excel (печатная форма)."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .models import OffEstimateSupplyRequest
from .off_estimate_docx_header import (
    TABLE_COL_WIDTHS,
    header_template_is_docx,
    resolve_header_template_path,
    write_docx_letterhead,
)

COL_COUNT = 5
LAST_COL = "E"
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def export_filename(req: OffEstimateSupplyRequest) -> str:
    num = (req.number or str(req.pk)).replace("/", "-").replace("\\", "-")
    return f"Заявка_на_материалы_№_{num}.xlsx"


def _fmt_qty(value) -> str:
    if value is None:
        return "0"
    try:
        d = value
        if hasattr(d, "normalize"):
            s = format(d.normalize(), "f")
        else:
            s = str(d)
        s = s.rstrip("0").rstrip(".")
        return s or "0"
    except Exception:
        return str(value)



def _default_header_lines(req: OffEstimateSupplyRequest) -> list[str]:
    project_name = req.project.name if req.project_id else "—"
    req_date = req.created_at.strftime("%d.%m.%Y") if req.created_at else "—"
    return [
        "ЗАЯВКА НА ПРИОБРЕТЕНИЕ МАТЕРИАЛОВ",
        "",
        f"Проект: {project_name}",
        f"Дата заявки: {req_date}",
        f"Номер заявки: {req.number}",
    ]


def _render_header_text(company, req: OffEstimateSupplyRequest) -> list[str]:
    raw = (company.off_estimate_excel_header_text or "").strip()
    if not raw:
        return _default_header_lines(req)
    ctx = {
        "project": req.project.name if req.project_id else "—",
        "date": req.created_at.strftime("%d.%m.%Y") if req.created_at else "—",
        "number": req.number or "—",
        "company": company.name,
    }
    lines = []
    for line in raw.splitlines():
        try:
            lines.append(line.format(**ctx))
        except (KeyError, ValueError):
            lines.append(line)
    return lines or _default_header_lines(req)


def _merge_row(ws, row: int, value: str, *, bold: bool = False, center: bool = False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COL_COUNT)
    cell = ws.cell(row=row, column=1, value=value)
    cell.font = Font(bold=bold, size=12 if bold else 11)
    cell.alignment = CENTER if center else LEFT
    return row + 1


def _write_label_with_line_row(ws, row: int, label: str) -> int:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = Font(size=11)
    label_cell.alignment = LEFT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=COL_COUNT)
    line_cell = ws.cell(row=row, column=2, value="______________________")
    line_cell.font = Font(size=11)
    line_cell.alignment = Alignment(horizontal="left", vertical="bottom")
    return row + 1


def _write_signature_block(ws, row: int) -> int:
    row = _merge_row(ws, row, "Составил:")
    row = _write_label_with_line_row(ws, row, "ФИО:")
    row = _write_label_with_line_row(ws, row, "Подпись:")
    return row


def _copy_template_header(ws, template_path: str, start_row: int) -> int:
    tmpl = load_workbook(template_path, data_only=True, read_only=True)
    try:
        src = tmpl.active
        row = start_row
        empty_streak = 0
        for src_row in src.iter_rows(max_row=25, values_only=True):
            if all(v is None or str(v).strip() == "" for v in src_row):
                empty_streak += 1
                if empty_streak >= 2:
                    break
                row += 1
                continue
            empty_streak = 0
            for col_idx, val in enumerate(src_row[:COL_COUNT], start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.alignment = WRAP
            if len(src_row) > COL_COUNT:
                extra = " ".join(str(v) for v in src_row[COL_COUNT:] if v not in (None, ""))
                if extra:
                    ws.cell(row=row, column=COL_COUNT, value=extra)
            row += 1
        return row + 1
    finally:
        tmpl.close()


def _write_request_info_header(ws, req: OffEstimateSupplyRequest, start_row: int) -> int:
    project_name = req.project.name if req.project_id else "—"
    req_date = req.created_at.strftime("%d.%m.%Y") if req.created_at else "—"
    lines = [
        "ЗАЯВКА НА ПРИОБРЕТЕНИЕ МАТЕРИАЛОВ",
        "",
        f"Проект: {project_name}",
        f"Дата заявки: {req_date}",
        f"Номер заявки: {req.number}",
    ]
    row = start_row
    for i, line in enumerate(lines):
        if i == 0 and line.strip():
            row = _merge_row(ws, row, line.strip(), bold=True, center=True)
        elif line.strip():
            row = _merge_row(ws, row, line.strip())
        else:
            row += 1
    return row + 1


def _write_text_header(ws, company, req: OffEstimateSupplyRequest, start_row: int) -> int:
    lines = _render_header_text(company, req)
    row = start_row
    for i, line in enumerate(lines):
        if i == 0 and line.strip():
            row = _merge_row(ws, row, line.strip(), bold=True, center=True)
        elif line.strip():
            row = _merge_row(ws, row, line.strip())
        else:
            row += 1
    return row + 1


def _style_table_cell(cell, *, header: bool = False):
    cell.border = BORDER
    cell.alignment = WRAP if not header else CENTER
    if header:
        cell.font = Font(bold=True)


def _apply_table_column_widths(ws):
    for col, width in TABLE_COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def export_single_off_estimate_request_xlsx(
    req: OffEstimateSupplyRequest,
) -> BytesIO:
    company = req.company
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявка"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    row = 1
    template_path = resolve_header_template_path(company)
    if template_path and header_template_is_docx(template_path):
        row = write_docx_letterhead(ws, template_path, row)
        row = _write_request_info_header(ws, req, row)
    elif template_path:
        row = _copy_template_header(ws, template_path, row)
    else:
        row = _write_text_header(ws, company, req, row)

    headers = ["№", "Наименование материала", "Ед. изм.", "Количество", "Примечание"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        _style_table_cell(cell, header=True)
    table_start = row
    row += 1

    items = list(req.items.order_by("sort_order", "id"))
    for idx, line in enumerate(items, start=1):
        values = [
            idx,
            line.material_name,
            line.unit,
            _fmt_qty(line.quantity),
            line.note or "",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_table_cell(cell)
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
        row += 1

    if not items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COL_COUNT)
        cell = ws.cell(row=row, column=1, value="Нет позиций")
        cell.alignment = CENTER
        cell.border = BORDER
        row += 1

    row += 1
    row = _write_signature_block(ws, row)

    _apply_table_column_widths(ws)
    ws.print_area = f"A1:{LAST_COL}{row}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
