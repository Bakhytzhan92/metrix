"""Журнал расходов компании."""
from __future__ import annotations

import os
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from django.db import transaction
from django.db.models import QuerySet, Sum
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import (
    Company,
    Employee,
    ExpenseJournalEntry,
    FinanceOperation,
    Project,
    SupplyOrder,
)

PAYMENT_METHOD_META = [
    {"code": code, "label": label}
    for code, label in ExpenseJournalEntry.PAYMENT_METHOD_CHOICES
]

EMPLOYEE_PILL_COLORS = [
    ("#fecaca", "#7f1d1d"),
    ("#fde68a", "#78350f"),
    ("#bfdbfe", "#1e3a8a"),
    ("#bbf7d0", "#14532d"),
    ("#e9d5ff", "#581c87"),
    ("#fed7aa", "#9a3412"),
]


def employee_pill_style(employee_id: int | None) -> dict[str, str]:
    if not employee_id:
        return {"bg": "#f1f5f9", "text": "#475569"}
    idx = int(employee_id) % len(EMPLOYEE_PILL_COLORS)
    bg, text = EMPLOYEE_PILL_COLORS[idx]
    return {"bg": bg, "text": text}


def payment_pill_style(method: str) -> dict[str, str]:
    styles = {
        ExpenseJournalEntry.PAY_KASPI: ("#ef4444", "#ffffff"),
        ExpenseJournalEntry.PAY_CASH: ("#22c55e", "#ffffff"),
        ExpenseJournalEntry.PAY_HALYK: ("#16a34a", "#ffffff"),
        ExpenseJournalEntry.PAY_ACCOUNTABLE: ("#f59e0b", "#1f2937"),
    }
    bg, text = styles.get(method, ("#f1f5f9", "#475569"))
    return {"bg": bg, "text": text}


def _parse_decimal(raw) -> Decimal:
    try:
        val = Decimal(str(raw).replace(",", ".").replace(" ", ""))
    except (InvalidOperation, TypeError):
        raise ValueError("bad_amount")
    if val < 0:
        raise ValueError("bad_amount")
    return val


def list_employees(company: Company) -> list[dict[str, Any]]:
    return [
        {
            "id": e.pk,
            "full_name": e.full_name,
            "pill": employee_pill_style(e.pk),
        }
        for e in Employee.objects.filter(
            company=company, status=Employee.STATUS_ACTIVE
        ).order_by("full_name", "id")
    ]


def filter_entries(
    company: Company,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    project_id: int | None = None,
    responsible_id: int | None = None,
    payment_method: str = "",
) -> QuerySet[ExpenseJournalEntry]:
    qs = (
        ExpenseJournalEntry.objects.filter(company=company)
        .select_related("responsible", "project", "supply_order")
        .order_by("-date", "-id")
    )
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if project_id:
        qs = qs.filter(project_id=project_id)
    if responsible_id:
        qs = qs.filter(responsible_id=responsible_id)
    if payment_method:
        qs = qs.filter(payment_method=payment_method)
    return qs


def compute_kpis(
    company: Company,
    *,
    project_id: int | None = None,
    today: date | None = None,
) -> dict[str, Decimal]:
    today = today or timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    month_start = date(today.year, today.month, 1)
    base = ExpenseJournalEntry.objects.filter(company=company)

    def _sum(qs):
        val = qs.aggregate(s=Sum("amount")).get("s")
        return Decimal(str(val or 0))

    kpis = {
        "today": _sum(base.filter(date=today)),
        "week": _sum(base.filter(date__gte=week_start, date__lte=today)),
        "month": _sum(base.filter(date__gte=month_start, date__lte=today)),
        "project": Decimal("0"),
    }
    if project_id:
        kpis["project"] = _sum(base.filter(project_id=project_id))
    return kpis


def serialize_entry(row: ExpenseJournalEntry) -> dict[str, Any]:
    receipt_name = ""
    receipt_url = ""
    receipt_preview_url = ""
    if row.receipt_pdf:
        receipt_name = os.path.basename(row.receipt_pdf.name)
        try:
            receipt_url = row.receipt_pdf.url
        except ValueError:
            receipt_url = ""
        receipt_preview_url = reverse(
            "api_expense_journal_receipt_preview", args=[row.pk]
        )
    return {
        "id": row.pk,
        "date": row.date.isoformat(),
        "amount": str(row.amount),
        "purpose": row.purpose or "",
        "responsible_id": row.responsible_id,
        "responsible_name": row.responsible.full_name if row.responsible else "",
        "payment_method": row.payment_method or "",
        "project_id": row.project_id,
        "project_name": row.project.name if row.project else "",
        "receipt_url": receipt_url,
        "receipt_preview_url": receipt_preview_url,
        "receipt_name": receipt_name,
        "category": row.category,
        "from_supply": bool(row.supply_order_id),
        "supply_order_id": row.supply_order_id,
    }


@transaction.atomic
def create_entry(
    company: Company,
    *,
    user,
    entry_date: date | None = None,
    amount: Decimal | None = None,
    purpose: str = "",
    responsible_id: int | None = None,
    payment_method: str = "",
    project_id: int | None = None,
    comment: str = "",
    category: str = ExpenseJournalEntry.CAT_OTHER,
) -> ExpenseJournalEntry:
    row = ExpenseJournalEntry.objects.create(
        company=company,
        date=entry_date or timezone.localdate(),
        amount=amount or Decimal("0"),
        purpose=(purpose or "")[:500],
        responsible_id=responsible_id or None,
        payment_method=payment_method or "",
        project_id=project_id or None,
        comment=(comment or "")[:5000],
        category=category or ExpenseJournalEntry.CAT_OTHER,
        created_by=user,
    )
    return row


@transaction.atomic
def update_entry(
    row: ExpenseJournalEntry,
    *,
    data: dict[str, Any],
) -> ExpenseJournalEntry:
    if "date" in data and data["date"]:
        row.date = data["date"]
    if "amount" in data:
        row.amount = _parse_decimal(data["amount"])
    if "purpose" in data:
        row.purpose = (data["purpose"] or "")[:500]
    if "responsible_id" in data:
        rid = data["responsible_id"]
        row.responsible_id = int(rid) if rid else None
    if "payment_method" in data:
        pm = data["payment_method"] or ""
        if pm and pm not in dict(ExpenseJournalEntry.PAYMENT_METHOD_CHOICES):
            raise ValueError("bad_payment_method")
        row.payment_method = pm
    if "project_id" in data:
        pid = data["project_id"]
        row.project_id = int(pid) if pid else None
    if "comment" in data:
        row.comment = (data["comment"] or "")[:5000]
    if "category" in data:
        cat = data["category"] or ExpenseJournalEntry.CAT_OTHER
        if cat not in dict(ExpenseJournalEntry.CATEGORY_CHOICES):
            raise ValueError("bad_category")
        row.category = cat
    row.save()
    return row


@transaction.atomic
def delete_entry(row: ExpenseJournalEntry) -> None:
    if row.receipt_pdf:
        row.receipt_pdf.delete(save=False)
    row.delete()


@transaction.atomic
def render_receipt_preview_png(row: ExpenseJournalEntry) -> bytes:
    """Первая страница чека как PNG для предпросмотра в UI."""
    import fitz

    if not row.receipt_pdf:
        raise ValueError("no_file")
    with row.receipt_pdf.open("rb") as fh:
        pdf_bytes = fh.read()
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        if doc.page_count < 1:
            raise ValueError("empty_pdf")
        page = doc.load_page(0)
        target_width = 440
        scale = target_width / page.rect.width if page.rect.width else 1.0
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        return pix.tobytes("png")
    finally:
        doc.close()


def upload_receipt_pdf(row: ExpenseJournalEntry, uploaded_file) -> ExpenseJournalEntry:
    if not uploaded_file:
        raise ValueError("no_file")
    ext = os.path.splitext(uploaded_file.name or "")[1].lower()
    if ext != ".pdf":
        raise ValueError("bad_extension")
    if row.receipt_pdf:
        row.receipt_pdf.delete(save=False)
    row.receipt_pdf = uploaded_file
    row.save(update_fields=["receipt_pdf", "updated_at"])
    return row


@transaction.atomic
def create_from_supply_payment(
    *,
    order: SupplyOrder,
    finance_operation: FinanceOperation,
    amount: Decimal,
    pay_date: date,
    user,
) -> ExpenseJournalEntry:
    all_items = list(order.items.all())
    items = all_items[:3]
    names = ", ".join(it.display_name for it in items if it.display_name)
    if len(all_items) > 3:
        names += "…"
    purpose = f"Заказ #{order.pk}"
    if order.supplier:
        purpose += f" — {order.supplier}"
    if names:
        purpose += f" ({names})"

    responsible = Employee.objects.filter(
        company=order.company, status=Employee.STATUS_ACTIVE
    ).order_by("id").first()

    return ExpenseJournalEntry.objects.create(
        company=order.company,
        date=pay_date,
        amount=amount,
        purpose=purpose[:500],
        responsible=responsible,
        payment_method="",
        project=order.project,
        comment=(order.procurement_note or "")[:5000],
        category=ExpenseJournalEntry.CAT_MATERIALS,
        supply_order=order,
        finance_operation=finance_operation,
        created_by=user,
    )


def export_xlsx(company: Company, qs: QuerySet[ExpenseJournalEntry]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал расходов"
    headers = [
        "Дата",
        "Сумма расхода",
        "Назначение платежа",
        "Ответственное лицо",
        "Способ оплаты",
        "Проект",
        "Чек",
        "Категория",
    ]
    header_fill = PatternFill("solid", fgColor="5B4CF0")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
    pay_labels = dict(ExpenseJournalEntry.PAYMENT_METHOD_CHOICES)
    cat_labels = dict(ExpenseJournalEntry.CATEGORY_CHOICES)
    for row_idx, ent in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=ent.date.strftime("%d.%m.%Y"))
        ws.cell(row=row_idx, column=2, value=float(ent.amount or 0))
        ws.cell(row=row_idx, column=3, value=ent.purpose)
        ws.cell(
            row=row_idx,
            column=4,
            value=ent.responsible.full_name if ent.responsible else "",
        )
        ws.cell(
            row=row_idx,
            column=5,
            value=pay_labels.get(ent.payment_method, ""),
        )
        ws.cell(row=row_idx, column=6, value=ent.project.name if ent.project else "")
        ws.cell(
            row=row_idx,
            column=7,
            value=os.path.basename(ent.receipt_pdf.name) if ent.receipt_pdf else "",
        )
        ws.cell(row=row_idx, column=8, value=cat_labels.get(ent.category, ""))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_pdf(company: Company, qs: QuerySet[ExpenseJournalEntry]) -> BytesIO:
    import fitz

    doc = fitz.open()
    page = doc.new_page(width=842, height=595)
    y = 36
    page.insert_text(
        (36, y),
        f"Журнал расходов — {company.name}",
        fontsize=14,
        fontname="helv",
    )
    y += 22
    pay_labels = dict(ExpenseJournalEntry.PAYMENT_METHOD_CHOICES)
    for ent in qs:
        if y > 560:
            page = doc.new_page(width=842, height=595)
            y = 36
        line = (
            f"{ent.date:%d.%m.%Y} | {ent.amount} | {ent.purpose[:40]} | "
            f"{ent.responsible.full_name if ent.responsible else '—'} | "
            f"{pay_labels.get(ent.payment_method, '—')} | "
            f"{ent.project.name if ent.project else '—'}"
        )
        page.insert_text((36, y), line, fontsize=8, fontname="helv")
        y += 12
    buf = BytesIO(doc.tobytes())
    doc.close()
    buf.seek(0)
    return buf
