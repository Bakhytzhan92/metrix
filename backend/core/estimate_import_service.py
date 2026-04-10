"""
Импорт сметы из Excel.
Формат: Раздел | Наименование | Тип | Ед | Кол-во | Цена (себестоимость за ед.).
Тип: материалы / работы / механизмы / доставка (и синонимы в TYPE_MAP).
Наценка при импорте = 0; цена заказчика считается в системе как цена × (1 + наценка/100).
"""
from decimal import Decimal
from typing import Any

from django.db.models import Max
from openpyxl import load_workbook

from .models import EstimateSection, EstimateItem, Project


TYPE_MAP = {
    "material": "material",
    "материалы": "material",
    "труд": "labor",
    "labor": "labor",
    "люди": "labor",
    "люд": "labor",
    "работа": "labor",
    "работы": "labor",
    "механизмы": "equipment",
    "мех": "equipment",
    "equipment": "equipment",
    "доставка": "delivery",
    "delivery": "delivery",
}


def _normalize_type(raw: str) -> str:
    v = (raw or "").strip().lower()
    return TYPE_MAP.get(v, "material")


def import_estimate_from_excel(project: Project, file) -> dict[str, Any]:
    """
    Импортирует смету из Excel.
    Ожидаемые колонки (по порядку): Раздел, Наименование, Тип, Ед, Кол-во, Цена.
    Первая строка — заголовки.
    Возвращает: {"sections_created": int, "items_created": int, "errors": list}.
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return {"sections_created": 0, "items_created": 0, "errors": ["Нет листа в файле"]}

    sections_by_name: dict[str, EstimateSection] = {}
    agg = EstimateSection.objects.filter(project=project).aggregate(m=Max("order"))
    next_order = (agg.get("m") or 0) + 1
    sections_created = 0
    items_created = 0
    errors: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None for cell in row):
            continue
        parts = [str(c).strip() if c is not None else "" for c in (list(row)[:6] or [])]
        section_name = (parts[0] or "").strip()
        name = (parts[1] or "").strip()
        type_str = _normalize_type(parts[2] if len(parts) > 2 else "")
        unit = (parts[3] or "шт").strip() or "шт"
        try:
            quantity = Decimal(str(parts[4]).replace(",", ".") or "0")
        except Exception:
            quantity = Decimal("0")
        try:
            cost_price = Decimal(str(parts[5]).replace(",", ".").replace(" ", "") or "0")
        except Exception:
            cost_price = Decimal("0")

        if not section_name:
            errors.append(f"Строка {row_idx}: не указан раздел")
            continue
        if not name:
            errors.append(f"Строка {row_idx}: не указано наименование")
            continue

        if section_name not in sections_by_name:
            section, created = EstimateSection.objects.get_or_create(
                project=project,
                name=section_name,
                defaults={"order": next_order},
            )
            if created:
                next_order += 1
                sections_created += 1
            sections_by_name[section_name] = section

        section = sections_by_name[section_name]
        EstimateItem.objects.create(
            section=section,
            name=name,
            type=type_str,
            unit=unit,
            quantity=quantity,
            cost_price=cost_price,
            markup_percent=Decimal("0"),
            order=row_idx,
        )
        items_created += 1

    wb.close()
    return {
        "sections_created": sections_created,
        "items_created": items_created,
        "errors": errors,
    }
