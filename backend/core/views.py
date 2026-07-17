from __future__ import annotations

import calendar
import json
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Max, Min, Prefetch, Q, Sum
from django.http import HttpRequest, HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.core.exceptions import ValidationError
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import ensure_csrf_cookie

from . import construction_services
from . import finance_project_services
from . import services as report_services
from . import supply_services
from . import off_estimate_supply_services as oes_services
from . import supply_workflow_services as wf_services
from .off_estimate_excel_export import export_filename, export_single_off_estimate_request_xlsx
from .auth_utils import login_user
from .access_utils import can_manage_access, get_current_company, has_permission, is_company_owner
from .forms import (
    RegisterForm,
    CompanyForm,
    ProjectForm,
    TaskForm,
    TaskQuickForm,
    FinanceForm,
    FinanceIncomeForm,
    FinanceExpenseForm,
    FinanceTransferForm,
    ProjectFinanceIncomeForm,
    ProjectFinanceExpenseForm,
    ProjectFinanceOperationEditForm,
    ProjectPaySupplyOrderForm,
    ProjectPayWorkActForm,
    WorkActForm,
    ConstructionWorkReportForm,
    InventoryItemForm,
    SupplyRequestForm,
    SupplyOrderForm,
    SupplyOrderCreateForm,
    ProjectSupplyRequestForm,
    ProjectOffEstimateSupplyRequestForm,
    ProjectSupplyOrderCreateForm,
    WarehouseIncomingForm,
    WarehouseOutgoingForm,
    WarehouseTransferForm,
    AddCompanyUserForm,
    EditCompanyUserForm,
    EstimateSectionForm,
    EstimateItemForm,
    EstimateItemInlineForm,
    WarehouseCreateForm,
    StockIncomingForm,
    StockWriteoffForm,
    StockTransferForm,
    MaterialCreateForm,
    WarehouseInventoryCreateForm,
    WarehouseInventoryUpdateForm,
    InventoryTransferForm,
    EquipmentForm,
)
from .models import (
    Company,
    CompanyRole,
    CompanyUser,
    Project,
    ProjectAccess,
    Task,
    Finance,
    InventoryItem,
    Account,
    FinanceCategory,
    FinanceOperation,
    WorkAct,
    Resource,
    SupplyRequest,
    SupplyOrder,
    SupplyOrderItem,
    OffEstimateSupplyRequest,
    OFF_ESTIMATE_UNIT_PRESETS,
    Warehouse,
    StockItem,
    WarehouseOperation,
    EstimateSection,
    EstimateItem,
    ESTIMATE_ITEM_NAME_MAX_LENGTH,
    ConstructionWorkLog,
    ConstructionWorkPhoto,
    Material,
    Stock,
    StockMovement,
    WarehouseInventoryItem,
    InventoryLog,
    FuelStock,
    Equipment,
    EquipmentDocument,
    EquipmentAuditLog,
)
from .warehouse_services import apply_incoming, apply_writeoff, apply_transfer
from .inventory_services import (
    get_written_off_warehouse,
    log_inventory_action,
    transfer_inventory_item,
    set_inventory_status,
    material_names_casefold_for_company,
    is_equipment_inventory_row,
)
from .inventory_numbering import allocate_inventory_number
from .inventory_qr import ensure_qr_image_file
from .rbac import permission_required
from .estimate_format import format_sell_price

User = get_user_model()


def register(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Создаём компанию по умолчанию для нового пользователя
            Company.objects.create(name=f"Компания {user.username}", owner=user)
            login_user(request, user)
            return redirect("dashboard")
    else:
        form = RegisterForm()
    return render(request, "registration/register.html", {"form": form})


@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    from .company_roles import ensure_company_default_roles

    company = get_current_company(request.user)
    if not company:
        company = Company.objects.create(
            name=f"Компания {request.user.username}", owner=request.user
        )
        ensure_company_default_roles(company)
    projects = company.projects.all().order_by("-created_at")
    status = request.GET.get("status")
    order = request.GET.get("order", "recent")

    if status:
        projects = projects.filter(status=status)

    if order == "oldest":
        projects = projects.order_by("created_at")

    context = {
        "company": company,
        "projects": projects,
        "status_filter": status or "",
        "order": order,
    }
    return render(request, "core/dashboard.html", context)


@login_required
def project_create(request: HttpRequest) -> HttpResponse:
    from .company_roles import ensure_company_default_roles
    from .subscription_limits import (
        apply_trial_for_new_company,
        can_create_project,
    )

    company = get_current_company(request.user)
    if not company:
        company = Company.objects.create(
            name=f"Компания {request.user.username}", owner=request.user
        )
        ensure_company_default_roles(company)
        apply_trial_for_new_company(
            company,
        )
    if request.method == "POST":
        ok_proj, err_proj = can_create_project(
            company,
        )
        if not ok_proj:
            messages.error(
                request,
                err_proj or "Нельзя создать проект.",
            )
            return redirect(
                "dashboard",
            )
        form = ProjectForm(request.POST)
        if form.is_valid():
            project = form.save(commit=False)
            project.company = company
            project.save()
            for cu in CompanyUser.objects.filter(company=company, auto_add_to_new_projects=True, is_active=True):
                ProjectAccess.objects.get_or_create(
                    company_user=cu,
                    project=project,
                    defaults={"role_in_project": ProjectAccess.ROLE_VIEWER},
                )
            messages.success(request, "Проект создан.")
            return redirect("dashboard")
    else:
        form = ProjectForm()
    return render(request, "core/project_form.html", {"form": form, "company": company})


@login_required
def project_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """Редактирование проекта."""
    company = get_current_company(request.user)
    if not company:
        return redirect("dashboard")
    project = get_object_or_404(Project, pk=pk, company=company)
    if request.method == "POST":
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, "Проект обновлён.")
            return redirect("project_analytics", pk=project.pk)
    else:
        form = ProjectForm(instance=project)
    return render(request, "core/project_form.html", {
        "form": form,
        "company": company,
        "project": project,
        "is_edit": True,
    })


@login_required
@require_POST
def project_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Удаление проекта."""
    company = get_current_company(request.user)
    if not company:
        return redirect("dashboard")
    project = get_object_or_404(Project, pk=pk, company=company)
    project.delete()
    messages.success(request, "Проект удалён.")
    return redirect("dashboard")


def _get_project_or_403(request: HttpRequest, pk: int):
    """Возвращает (project, None) если проект принадлежит компании пользователя, иначе (None, HttpResponse 403/redirect)."""
    company = get_current_company(request.user)
    if not company:
        return None, redirect("dashboard")
    project = get_object_or_404(Project, pk=pk)
    if project.company_id != company.id:
        from django.http import HttpResponseForbidden
        return None, HttpResponseForbidden("<h1>403</h1><p>Проект не принадлежит вашей компании.</p>")
    return project, None


@login_required
def project_overview(request: HttpRequest, pk: int) -> HttpResponse:
    """Редирект с /projects/<id>/ на раздел Аналитика."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    return redirect("project_analytics", pk=project.pk)


@login_required
def project_analytics(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    from . import project_analytics_services as analytics_svc

    return render(
        request,
        "core/project/analytics.html",
        {
            "project": project,
            "active_tab": "analytics",
            "analytics": analytics_svc.compute_project_analytics(project),
        },
    )


ESTIMATE_VIRTUAL_ROW_THRESHOLD = 120
SCHEDULE_VIRTUAL_ROW_THRESHOLD = 1


def _estimate_sheet_row_count(sections) -> int:
    return sum(len(s.items.all()) for s in sections)


def _estimate_virtual_payload(request: HttpRequest, project, sections):
    """JSON для клиентской виртуализации строк сметы (React + react-window)."""
    from decimal import Decimal

    from django.middleware.csrf import get_token

    from .templatetags.estimate_extras import qty_plain, strip_norm_code
    from .services.excel_estimate_parser import normalize_estimate_name

    out_sections = []
    for sec in sections:
        rows = []
        ordinal = 0
        items_list = list(sec.items.all())
        sc_dec = Decimal("0")
        sp_dec = Decimal("0")
        item_count = 0
        for item in items_list:
            if item.is_subsection_header:
                rows.append(
                    {
                        "kind": "banner",
                        "key": f"h-{item.pk}",
                        "title": item.name or "",
                    }
                )
                continue
            item_count += 1
            pos_no = (item.pdf_pos_no or "").strip()
            if pos_no.isdigit():
                ordinal = int(pos_no)
            else:
                ordinal += 1
            tc = item.total_cost or Decimal("0")
            tp = item.total_price or Decimal("0")
            sc_dec += tc
            sp_dec += tp
            rows.append(
                {
                    "kind": "item",
                    "key": f"i-{item.pk}",
                    "id": item.pk,
                    "ordinal": ordinal,
                    "type": item.type,
                    "name": strip_norm_code(
                        normalize_estimate_name(item.name or "")
                    ),
                    "unit": item.unit or "",
                    "quantity": qty_plain(item.quantity),
                    "cost_price": str(item.cost_price),
                    "markup_percent": str(item.markup_percent),
                    "sell_price": format_sell_price(item.sell_price),
                    "total_cost": f"{item.total_cost:.2f}",
                    "total_price": f"{item.total_price:.2f}",
                    "urls": {
                        "save": reverse(
                            "estimate_item_inline",
                            args=[project.pk, sec.pk, item.pk],
                        ),
                        "delete": reverse(
                            "estimate_item_delete",
                            args=[project.pk, sec.pk, item.pk],
                        ),
                        "supply": (
                            f"{reverse('project_supply', args=[project.pk])}"
                            f"?tab=requests&estimate_item={item.pk}"
                            f"#supply-create-request"
                        ),
                    },
                }
            )
        out_sections.append(
            {
                "id": sec.pk,
                "order": sec.order,
                "name": normalize_estimate_name(sec.name or ""),
                "header_style": getattr(sec, "header_style", "") or "",
                "displayBadge": sec.order,
                "item_count": item_count,
                "section_total_cost": f"{sc_dec:.2f}",
                "section_total_price": f"{sp_dec:.2f}",
                "save_url": reverse(
                    "estimate_section_inline", args=[project.pk, sec.pk]
                ),
                "delete_url": reverse(
                    "estimate_section_delete", args=[project.pk, sec.pk]
                ),
                "quick_add_action": reverse(
                    "estimate_item_quick_add", args=[project.pk, sec.pk]
                ),
                "rows": rows,
            }
        )
    return {
        "project_id": project.pk,
        "csrf_token": get_token(request),
        "sections": out_sections,
    }


def _get_estimate_totals(project):
    """Итоги сметы по проекту: total_cost, total_price, markup, vat, client_total."""
    from decimal import Decimal, ROUND_HALF_UP

    from django.db.models import Sum

    agg = EstimateItem.objects.filter(
        section__project=project, is_subsection_header=False
    ).aggregate(
        cost=Sum("total_cost"),
        price=Sum("total_price"),
    )
    cost = agg.get("cost") or Decimal("0")
    price = agg.get("price") or Decimal("0")
    if not isinstance(cost, Decimal):
        cost = Decimal(str(cost))
    if not isinstance(price, Decimal):
        price = Decimal(str(price))
    markup = price - cost
    q = lambda d: d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    vat_amt = q(price * Decimal("0.16")) if project.estimate_vat_enabled else Decimal("0")
    client_total = q(price * Decimal("1.16")) if project.estimate_vat_enabled else price
    return {
        "total_cost": cost,
        "total_price": price,
        "markup": markup,
        "vat_amount": vat_amt,
        "client_total": client_total,
    }


@login_required
def project_estimate(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    sections = (
        EstimateSection.objects.filter(project=project)
        .prefetch_related("items")
        .order_by("order", "id")
    )
    totals = _get_estimate_totals(project)
    rows_total = _estimate_sheet_row_count(sections)
    use_virtual = rows_total >= ESTIMATE_VIRTUAL_ROW_THRESHOLD
    ctx = {
        "project": project,
        "active_tab": "estimate",
        "sections": sections,
        "totals": totals,
        "section_form": EstimateSectionForm(),
        "item_form": EstimateItemForm(),
        "use_estimate_virtual": use_virtual,
    }
    if use_virtual:
        ctx["estimate_virtual_payload"] = _estimate_virtual_payload(
            request, project, sections
        )
    return render(request, "core/project/estimate.html", ctx)


@login_required
@require_POST
def estimate_vat_toggle(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    project.estimate_vat_enabled = request.POST.get("vat") in (
        "1",
        "on",
        "true",
        "True",
    )
    project.save(update_fields=["estimate_vat_enabled"])
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if wants_json:
        return JsonResponse({"ok": True, "vat_enabled": project.estimate_vat_enabled})
    return redirect("project_estimate", pk=project.pk)


@login_required
def estimate_section_add(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    if request.method == "POST":
        form = EstimateSectionForm(request.POST)
        if form.is_valid():
            section = form.save(commit=False)
            section.project = project
            max_order = (
                EstimateSection.objects.filter(project=project).aggregate(
                    m=Max("order")
                ).get("m")
                or 0
            )
            section.order = max_order + 1
            section.save()
            messages.success(request, "Раздел добавлен.")
            return redirect("project_estimate", pk=project.pk)
    else:
        form = EstimateSectionForm()
    return render(
        request,
        "core/project/estimate_section_form.html",
        {"project": project, "form": form, "active_tab": "estimate", "is_edit": False},
    )


@login_required
def estimate_section_edit(request: HttpRequest, pk: int, section_id: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    if request.method == "POST":
        form = EstimateSectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, "Раздел обновлён.")
            return redirect("project_estimate", pk=project.pk)
    else:
        form = EstimateSectionForm(instance=section)
    return render(
        request,
        "core/project/estimate_section_form.html",
        {"project": project, "section": section, "form": form, "active_tab": "estimate", "is_edit": True},
    )


@login_required
@require_POST
def estimate_section_inline(request: HttpRequest, pk: int, section_id: int) -> HttpResponse:
    """Сохранение названия и порядка раздела сметы с страницы сметы (без отдельной формы)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    if request.headers.get("X-Requested-With") != "XMLHttpRequest":
        return HttpResponseBadRequest("Expected XMLHttpRequest")
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    form = EstimateSectionForm(request.POST, instance=section)
    if form.is_valid():
        form.save()
        return JsonResponse({"ok": True, "name": section.name, "order": section.order})
    err_payload = {k: [str(x) for x in v] for k, v in form.errors.items()}
    return JsonResponse({"ok": False, "errors": err_payload}, status=400)


@login_required
@require_POST
def estimate_section_delete(request: HttpRequest, pk: int, section_id: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    section.delete()
    messages.success(request, "Раздел удалён.")
    return redirect("project_estimate", pk=project.pk)


@login_required
@require_POST
def estimate_sections_delete_all(request: HttpRequest, pk: int) -> HttpResponse:
    """Удалить все разделы сметы проекта и все позиции (CASCADE). Связь этапов графика с разделом обнуляется (SET_NULL)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    EstimateSection.objects.filter(project=project).delete()
    messages.success(request, "Все разделы сметы и позиции удалены.")
    return redirect("project_estimate", pk=project.pk)


@login_required
def estimate_item_add(request: HttpRequest, pk: int, section_id: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    if request.method == "POST":
        form = EstimateItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.section = section
            item.save()
            messages.success(request, "Позиция добавлена.")
            return redirect("project_estimate", pk=project.pk)
    else:
        form = EstimateItemForm()
    return render(
        request,
        "core/project/estimate_item_form.html",
        {"project": project, "section": section, "form": form, "active_tab": "estimate", "is_edit": False},
    )


@login_required
@require_POST
def estimate_item_quick_add(request: HttpRequest, pk: int, section_id: int) -> HttpResponse:
    """Добавить пустую строку сметы сразу в таблицу, без отдельной страницы."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    max_order = section.items.aggregate(m=Max("order")).get("m") or 0
    EstimateItem.objects.create(
        section=section,
        name="",
        type=EstimateItem.TYPE_MATERIAL,
        unit="-",
        quantity=0,
        cost_price=0,
        markup_percent=0,
        order=max_order + 1,
    )
    messages.success(request, "Позиция добавлена.")
    return redirect(f"{reverse('project_estimate', args=[project.pk])}#section-{section_id}")


def _normalize_estimate_inline_post(post):
    """Копия POST для сметы: запятая → точка, пустые числа → 0 (как у полей по умолчанию в модели)."""
    data = post.copy()
    for key in ("quantity", "cost_price", "markup_percent", "sell_price"):
        if key not in data:
            continue
        raw = data.get(key)
        if raw is None:
            continue
        s = str(raw).strip().replace(" ", "").replace(",", ".")
        if s == "" or s in ("-", ".", "-."):
            data[key] = "0"
        else:
            data[key] = s
    return data


@login_required
@require_POST
def estimate_item_inline(request: HttpRequest, pk: int, section_id: int, item_id: int) -> HttpResponse:
    """Сохранение позиции сметы из таблицы без перехода на отдельную страницу."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    item = get_object_or_404(EstimateItem, pk=item_id, section=section)
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    post_data = _normalize_estimate_inline_post(request.POST)
    form = EstimateItemInlineForm(post_data, instance=item)
    if form.is_valid():
        saved = form.save()
        if wants_json:
            agg = EstimateItem.objects.filter(
                section=section, is_subsection_header=False
            ).aggregate(
                sc=Sum("total_cost"), sp=Sum("total_price")
            )
            sc = agg.get("sc") or 0
            sp = agg.get("sp") or 0
            return JsonResponse(
                {
                    "ok": True,
                    "total_cost": f"{saved.total_cost:.2f}",
                    "total_price": f"{saved.total_price:.2f}",
                    "sell_price": format_sell_price(saved.sell_price),
                    "markup_percent": f"{saved.markup_percent:.2f}",
                    "section_total_cost": f"{sc:.2f}",
                    "section_total_price": f"{sp:.2f}",
                }
            )
        messages.success(request, "Строка сметы сохранена.")
    else:
        if wants_json:
            err_payload = {k: list(v) for k, v in form.errors.items()}
            return JsonResponse({"ok": False, "errors": err_payload}, status=400)
        messages.error(request, "Не удалось сохранить строку: проверьте значения.")
    return redirect(f"{reverse('project_estimate', args=[project.pk])}#section-{section_id}")


@login_required
def estimate_item_edit(request: HttpRequest, pk: int, section_id: int, item_id: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    item = get_object_or_404(EstimateItem, pk=item_id, section=section)
    if request.method == "POST":
        form = EstimateItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Позиция обновлена.")
            return redirect("project_estimate", pk=project.pk)
    else:
        form = EstimateItemForm(instance=item)
    return render(
        request,
        "core/project/estimate_item_form.html",
        {"project": project, "section": section, "item": item, "form": form, "active_tab": "estimate", "is_edit": True},
    )


@login_required
@require_POST
def estimate_item_delete(request: HttpRequest, pk: int, section_id: int, item_id: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    section = get_object_or_404(EstimateSection, pk=section_id, project=project)
    item = get_object_or_404(EstimateItem, pk=item_id, section=section)
    item.delete()
    messages.success(request, "Позиция удалена.")
    return redirect("project_estimate", pk=project.pk)


@login_required
def estimate_import(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    if request.method != "POST" or not request.FILES.get("file"):
        messages.error(request, "Выберите файл Excel.")
        return redirect("project_estimate", pk=project.pk)
    from .estimate_import_service import import_estimate_from_excel
    result = import_estimate_from_excel(project, request.FILES["file"])
    if result["errors"]:
        for e in result["errors"]:
            messages.warning(request, e)
    err_n = result.get("error_count", len(result["errors"]))
    messages.success(
        request,
        "Импорт Excel: "
        f"разделов {result['sections_created']}, "
        f"позиций {result['items_created']}, "
        f"пропущено строк {result.get('skipped', 0)}, "
        f"ошибок {err_n}.",
    )
    return redirect("project_estimate", pk=project.pk)


@login_required
def estimate_export(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    from io import BytesIO
    from openpyxl import Workbook
    sections = (
        EstimateSection.objects.filter(project=project)
        .prefetch_related("items")
        .order_by("order", "id")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.append(
        [
            "Раздел",
            "Наименование",
            "Тип",
            "Ед",
            "Кол-во",
            "Цена",
            "Себестоимость",
            "Наценка %",
            "Цена для заказчика",
            "Итого",
        ]
    )
    for section in sections:
        for item in section.items.all():
            ws.append([
                section.name,
                item.name,
                item.get_type_display(),
                item.unit,
                float(item.quantity),
                float(item.cost_price),
                float(item.total_cost),
                float(item.markup_percent),
                float(item.sell_price),
                float(item.total_price),
            ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="estimate_{project.pk}.xlsx"'
    return response


def _item_schedule_duration_days(item: EstimateItem) -> int | None:
    if item.schedule_start and item.schedule_end:
        return (item.schedule_end - item.schedule_start).days + 1
    return None


def _parse_schedule_api_date(value) -> tuple[date | None, str | None]:
    """Дата графика из API: YYYY-MM-DD, год 1990–2100."""
    if value in (None, ""):
        return None, None
    if not isinstance(value, str):
        return None, "invalid_date"
    parsed = parse_date(value)
    if not parsed:
        return None, "invalid_date"
    if parsed.year < 1990 or parsed.year > 2100:
        return None, "year_out_of_range"
    return parsed, None


def _schedule_display_item_passes_filters(
    item: EstimateItem,
    *,
    show_materials: bool,
) -> bool:
    if not show_materials and item.type == EstimateItem.TYPE_MATERIAL:
        return False
    return True


def _schedule_task_passes_filters(
    task: EstimateItem,
    *,
    status: str,
    assignee_id: int | None,
    d_from: date | None,
    d_to: date | None,
) -> bool:
    if status and task.schedule_status != status:
        return False
    if assignee_id is not None and task.schedule_assignee_id != assignee_id:
        return False
    if d_from or d_to:
        if task.schedule_start and task.schedule_end:
            if d_from and task.schedule_end < d_from:
                return False
            if d_to and task.schedule_start > d_to:
                return False
        elif status or assignee_id is not None or d_from or d_to:
            return False
    return True


def _schedule_summarize_children(children: list[EstimateItem]) -> dict:
    from decimal import Decimal

    item_count = 0
    total_price = Decimal("0")
    by_unit: dict[str, Decimal] = {}
    for child in children:
        if child.is_subsection_header:
            continue
        item_count += 1
        total_price += child.total_price or Decimal("0")
        unit = (child.unit or "—").strip() or "—"
        qty = child.quantity or Decimal("0")
        by_unit[unit] = by_unit.get(unit, Decimal("0")) + qty

    vol_parts: list[str] = []
    for unit, qty in sorted(by_unit.items(), key=lambda x: x[0]):
        qf = qty
        if qf == qf.to_integral_value():
            qs = str(int(qf))
        else:
            qs = f"{qf:.4f}".rstrip("0").rstrip(".")
        vol_parts.append(f"{qs} {unit}")

    suggested_days: int | None = None
    if item_count > 0:
        suggested_days = max(1, min(90, (item_count + 4) // 5))

    return {
        "item_count": item_count,
        "total_volume": " • ".join(vol_parts[:4]),
        "total_cost": f"{total_price:.2f}",
        "suggested_days": suggested_days,
    }


def _schedule_group_visible_items(
    items: list[EstimateItem],
    *,
    show_materials: bool,
) -> list[EstimateItem]:
    return [
        it
        for it in items
        if it.is_subsection_header
        or _schedule_display_item_passes_filters(it, show_materials=show_materials)
    ]


def _schedule_skip_section_banner(
    section: EstimateSection, visible_items: list[EstimateItem]
) -> bool:
    """Не дублировать заголовок раздела, если в нём одна группа с тем же названием."""
    headers = [it for it in visible_items if it.is_subsection_header]
    if len(headers) != 1:
        return False
    sec_name = (section.name or "").strip().casefold()
    hdr_name = (headers[0].name or "").strip().casefold()
    return bool(sec_name and hdr_name and sec_name == hdr_name)


def _ensure_section_schedule_group(section: EstimateSection) -> EstimateItem:
    """Группа работ для раздела без подзаголовков (например, после импорта Excel)."""
    from decimal import Decimal

    existing = (
        EstimateItem.objects.filter(section=section, is_subsection_header=True)
        .order_by("order", "pk")
        .first()
    )
    if existing:
        return existing
    min_order = EstimateItem.objects.filter(section=section).aggregate(
        m=Min("order")
    )["m"]
    header_order = 0 if min_order is None else max(0, (min_order or 1) - 1)
    return EstimateItem.objects.create(
        section=section,
        name=(section.name or "Группа работ")[:ESTIMATE_ITEM_NAME_MAX_LENGTH],
        type=EstimateItem.TYPE_LABOR,
        unit="—",
        quantity=Decimal("0"),
        cost_price=Decimal("0"),
        markup_percent=Decimal("0"),
        order=header_order,
        is_subsection_header=True,
    )


def _build_schedule_task_rows(
    section_blocks: list[tuple[EstimateSection, list[EstimateItem]]],
    *,
    status: str,
    assignee_id: int | None,
    d_from: date | None,
    d_to: date | None,
    show_materials: bool,
) -> tuple[list[dict], list[date], list[date]]:
    """График: сроки на каждой позиции сметы (без строк групп)."""
    schedule_rows: list[dict] = []
    row_idx = 0
    project_starts: list[date] = []
    project_ends: list[date] = []

    schedulable_items: list[EstimateItem] = []
    for _sec, vis in section_blocks:
        for it in vis:
            if not it.is_subsection_header:
                schedulable_items.append(it)

    def _successor_id(item_pk: int) -> int | None:
        for oit in schedulable_items:
            if oit.schedule_predecessor_id == item_pk:
                return oit.pk
        return None

    def _item_row_dict(
        child: EstimateItem,
        section: EstimateSection,
        task_id: int | None,
    ) -> dict:
        dur = _item_schedule_duration_days(child)
        if child.schedule_start and child.schedule_end:
            project_starts.append(child.schedule_start)
            project_ends.append(child.schedule_end)
        return {
            "kind": "item",
            "row_index": row_idx,
            "id": f"i-{child.pk}",
            "section_id": section.pk,
            "task_id": task_id,
            "item_id": child.pk,
            "name": child.name or "—",
            "number": (child.pdf_pos_no or "").strip(),
            "quantity": str(child.quantity).rstrip("0").rstrip(".")
            if child.quantity
            else "",
            "unit": child.unit or "",
            "type": child.type,
            "total_price": f"{child.total_price:.2f}",
            "schedule_start": child.schedule_start.isoformat()
            if child.schedule_start
            else None,
            "schedule_end": child.schedule_end.isoformat()
            if child.schedule_end
            else None,
            "duration_days": dur,
            "status": child.schedule_status,
            "assignee_id": child.schedule_assignee_id,
            "predecessor_id": child.schedule_predecessor_id,
            "successor_id": _successor_id(child.pk),
        }

    for section, all_items in section_blocks:
        filtered_items = [
            it
            for it in all_items
            if not it.is_subsection_header
            and _schedule_item_passes_filters(
                it,
                status=status,
                assignee_id=assignee_id,
                d_from=d_from,
                d_to=d_to,
                show_materials=show_materials,
            )
        ]
        if not filtered_items:
            continue

        skip_section_banner = _schedule_skip_section_banner(section, all_items)

        if not skip_section_banner:
            schedule_rows.append(
                {
                    "kind": "estimate_section",
                    "row_index": row_idx,
                    "id": f"es-{section.pk}",
                    "section_id": section.pk,
                    "item_id": None,
                    "name": section.name,
                    "number": "",
                    "quantity": "",
                    "unit": "",
                    "schedule_start": None,
                    "schedule_end": None,
                    "duration_days": None,
                    "status": "",
                    "assignee_id": None,
                    "predecessor_id": None,
                    "successor_id": None,
                    "item_count": 0,
                    "total_volume": "",
                    "total_cost": "",
                    "suggested_days": None,
                }
            )
            row_idx += 1

        for child in filtered_items:
            row = _item_row_dict(child, section, None)
            row["row_index"] = row_idx
            schedule_rows.append(row)
            row_idx += 1

    return schedule_rows, project_starts, project_ends


def _schedule_item_passes_filters(
    item: EstimateItem,
    *,
    status: str,
    assignee_id: int | None,
    d_from: date | None,
    d_to: date | None,
    show_materials: bool,
) -> bool:
    if not show_materials and item.type == EstimateItem.TYPE_MATERIAL:
        return False
    if status and item.schedule_status != status:
        return False
    if assignee_id is not None and item.schedule_assignee_id != assignee_id:
        return False
    if d_from or d_to:
        if item.schedule_start and item.schedule_end:
            if d_from and item.schedule_end < d_from:
                return False
            if d_to and item.schedule_start > d_to:
                return False
        else:
            return False
    return True


def _schedule_item_api_payload(item: EstimateItem) -> dict:
    return {
        "id": item.pk,
        "schedule_start": item.schedule_start.isoformat()
        if item.schedule_start
        else None,
        "schedule_end": item.schedule_end.isoformat()
        if item.schedule_end
        else None,
        "duration_days": _item_schedule_duration_days(item),
        "schedule_status": item.schedule_status,
        "schedule_assignee_id": item.schedule_assignee_id,
        "schedule_predecessor_id": item.schedule_predecessor_id,
    }


def _schedule_item_api_response_dict(
    item: EstimateItem,
    *,
    link_updates: list[EstimateItem] | None = None,
) -> dict:
    out: dict = {"ok": True, "item": _schedule_item_api_payload(item)}
    if link_updates:
        out["schedule_links_updated"] = [
            _schedule_item_api_payload(x) for x in link_updates
        ]
    return out


def _schedule_item_api_response(item: EstimateItem) -> dict:
    return _schedule_item_api_response_dict(item)


def _slim_schedule_row_for_virtual(row: dict) -> dict:
    """Минимальный набор полей для клиентского графика (меньше JSON)."""
    kind = row.get("kind")
    if kind == "estimate_section":
        return {
            "kind": kind,
            "section_id": row["section_id"],
            "name": row["name"],
        }
    if kind == "item":
        return {
            "kind": kind,
            "section_id": row["section_id"],
            "task_id": row.get("task_id"),
            "item_id": row["item_id"],
            "name": row["name"],
            "number": row.get("number") or "",
            "quantity": row.get("quantity") or "",
            "unit": row.get("unit") or "",
            "schedule_start": row.get("schedule_start"),
            "schedule_end": row.get("schedule_end"),
            "duration_days": row.get("duration_days"),
            "status": row.get("status") or "planned",
            "assignee_id": row.get("assignee_id"),
            "predecessor_id": row.get("predecessor_id"),
            "successor_id": row.get("successor_id"),
        }
    if kind == "task":
        return {
            "kind": kind,
            "section_id": row["section_id"],
            "item_id": row["item_id"],
            "name": row["name"],
            "item_count": row.get("item_count"),
            "total_volume": row.get("total_volume"),
        }
    return row


def _load_project_schedule(
    request: HttpRequest,
    project,
) -> dict:
    """Общая загрузка строк графика и фильтров (страница + JSON API)."""
    show_materials = request.GET.get("materials", "1") != "0"
    st = request.GET.get("status", "").strip()
    if st not in dict(EstimateItem.SCHEDULE_STATUS_CHOICES):
        st = ""
    aid = request.GET.get("assignee", "").strip()
    filter_assignee_id = int(aid) if aid.isdigit() else None
    df = request.GET.get("date_from", "").strip()
    dt_to = request.GET.get("date_to", "").strip()
    d_from = parse_date(df) if df else None
    d_to = parse_date(dt_to) if dt_to else None

    item_qs = EstimateItem.objects.select_related(
        "schedule_assignee", "schedule_predecessor"
    ).order_by("order", "id")

    sections = (
        EstimateSection.objects.filter(project=project)
        .prefetch_related(Prefetch("items", queryset=item_qs))
        .order_by("order", "id")
    )

    section_blocks: list[tuple[EstimateSection, list[EstimateItem]]] = []
    for section in sections:
        all_items = list(section.items.all())
        if any(
            not it.is_subsection_header
            and _schedule_display_item_passes_filters(
                it, show_materials=show_materials
            )
            for it in all_items
        ):
            section_blocks.append((section, all_items))

    assignee_users = list(
        get_user_model()
        .objects.filter(company_users__company=project.company)
        .distinct()
        .order_by("username")
    )
    assignee_choices_json = [
        {"id": u.pk, "username": u.get_username()} for u in assignee_users
    ]

    schedule_rows, project_starts, project_ends = _build_schedule_task_rows(
        section_blocks,
        status=st,
        assignee_id=filter_assignee_id,
        d_from=d_from,
        d_to=d_to,
        show_materials=show_materials,
    )

    return {
        "section_blocks": section_blocks,
        "schedule_rows": schedule_rows,
        "project_starts": project_starts,
        "project_ends": project_ends,
        "assignee_users": assignee_users,
        "assignee_choices_json": assignee_choices_json,
        "show_materials": show_materials,
        "filter_status": st,
        "filter_assignee_id": filter_assignee_id,
        "filter_date_from": df,
        "filter_date_to": dt_to,
    }


def _schedule_virtual_payload(
    request: HttpRequest,
    project,
    schedule_rows: list[dict],
    status_choices,
    assignee_choices_json: list[dict],
    today: date,
) -> dict:
    """JSON для клиентской виртуализации графика работ (React + react-window)."""
    from django.middleware.csrf import get_token

    section_names = {
        r["section_id"]: r["name"]
        for r in schedule_rows
        if r.get("kind") == "estimate_section"
    }
    succ_catalog = [
        {
            "id": r["item_id"],
            "label": (r.get("name") or "—")[:140],
            "section_id": r["section_id"],
            "section_name": section_names.get(r["section_id"], "—"),
        }
        for r in schedule_rows
        if r.get("kind") == "item" and r.get("item_id")
    ]
    rows_client = [
        _slim_schedule_row_for_virtual({k: v for k, v in r.items() if k != "succ_choices"})
        for r in schedule_rows
    ]

    return {
        "project_id": project.pk,
        "csrf_token": get_token(request),
        "today": today.isoformat(),
        "rows": rows_client,
        "succ_catalog": succ_catalog,
        "assignees": assignee_choices_json,
        "status_choices": list(status_choices),
        "api_url_template": reverse(
            "schedule_item_api",
            args=[project.pk, 999999],
        ).replace("999999", "__ID__"),
    }


@login_required
def project_schedule(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err

    sched = _load_project_schedule(request, project)
    schedule_rows = sched["schedule_rows"]
    project_starts = sched["project_starts"]
    project_ends = sched["project_ends"]
    assignee_users = sched["assignee_users"]
    assignee_choices_json = sched["assignee_choices_json"]

    totals = _get_estimate_totals(project)

    if project_starts and project_ends:
        sched_start = min(project_starts)
        sched_end = max(project_ends)
        sched_days = (sched_end - sched_start).days + 1
    else:
        sched_start = sched_end = None
        sched_days = None

    schedule_rows_json = [
        {k: v for k, v in row.items() if k != "succ_choices"}
        for row in schedule_rows
    ]

    use_schedule_virtual = len(schedule_rows) >= SCHEDULE_VIRTUAL_ROW_THRESHOLD
    ctx = {
        "project": project,
        "active_tab": "schedule",
        "schedule_rows": schedule_rows,
        "schedule_rows_json": schedule_rows_json,
        "status_choices": EstimateItem.SCHEDULE_STATUS_CHOICES,
        "assignee_choices": assignee_users,
        "assignee_choices_json": assignee_choices_json,
        "filter_status": sched["filter_status"],
        "filter_assignee_id": sched["filter_assignee_id"],
        "filter_date_from": sched["filter_date_from"],
        "filter_date_to": sched["filter_date_to"],
        "show_materials": sched["show_materials"],
        "totals": totals,
        "schedule_summary_start": sched_start,
        "schedule_summary_end": sched_end,
        "schedule_summary_days": sched_days,
        "today": date.today(),
        "use_schedule_virtual": use_schedule_virtual,
        "schedule_virtual_data_url": reverse(
            "project_schedule_virtual_data",
            args=[project.pk],
        ),
    }
    return render(request, "core/project/schedule.html", ctx)


@login_required
def project_schedule_virtual_data(request: HttpRequest, pk: int) -> HttpResponse:
    """JSON для виртуального графика (отдельно от HTML — не раздувает страницу)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err

    sched = _load_project_schedule(request, project)
    schedule_rows = sched["schedule_rows"]
    if not schedule_rows:
        return JsonResponse({"ok": False, "error": "empty", "rows": []}, status=404)

    payload = _schedule_virtual_payload(
        request,
        project,
        schedule_rows,
        EstimateItem.SCHEDULE_STATUS_CHOICES,
        sched["assignee_choices_json"],
        date.today(),
    )
    return JsonResponse(payload)


@login_required
@require_POST
def schedule_item_api(request: HttpRequest, pk: int, item_id: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    item = get_object_or_404(
        EstimateItem.objects.select_related("section"),
        pk=item_id,
        section__project=project,
    )
    if item.is_subsection_header:
        return JsonResponse(
            {
                "ok": False,
                "error": "Сроки задаются для позиций сметы, не для групп работ.",
            },
            status=403,
        )
    try:
        data = json.loads(request.body.decode() or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "bad_json"}, status=400)

    touched: set[str] = set()

    if "schedule_start" in data:
        v = data.get("schedule_start")
        parsed, err = _parse_schedule_api_date(v)
        if err:
            return JsonResponse({"ok": False, "error": err}, status=400)
        item.schedule_start = parsed
        touched.add("schedule_start")
        if not item.schedule_start:
            item.schedule_end = None
            touched.add("schedule_end")
        elif "duration_days" not in data and "schedule_end" not in data:
            item.schedule_end = item.schedule_start
            touched.add("schedule_end")

    if "schedule_end" in data:
        v = data.get("schedule_end")
        parsed, err = _parse_schedule_api_date(v)
        if err:
            return JsonResponse({"ok": False, "error": err}, status=400)
        item.schedule_end = parsed
        touched.add("schedule_end")

    if "duration_days" in data:
        try:
            ddays = int(data["duration_days"])
        except (TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "duration_days"}, status=400)
        if ddays < 1:
            ddays = 1
        if not item.schedule_start:
            return JsonResponse(
                {"ok": False, "error": "need_schedule_start"}, status=400
            )
        item.schedule_end = item.schedule_start + timedelta(days=ddays - 1)
        touched.add("schedule_end")

    if "schedule_status" in data:
        val = data["schedule_status"]
        if val not in dict(EstimateItem.SCHEDULE_STATUS_CHOICES):
            return JsonResponse({"ok": False, "error": "status"}, status=400)
        item.schedule_status = val
        touched.add("schedule_status")

    if "schedule_assignee_id" in data:
        aid = data.get("schedule_assignee_id")
        if aid in (None, "", "null"):
            item.schedule_assignee_id = None
        else:
            try:
                uid = int(aid)
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "assignee"}, status=400)
            u = (
                get_user_model()
                .objects.filter(pk=uid, company_users__company=project.company)
                .first()
            )
            if not u:
                return JsonResponse({"ok": False, "error": "assignee"}, status=400)
            item.schedule_assignee_id = uid
        touched.add("schedule_assignee")

    if "schedule_predecessor_id" in data:
        pid = data.get("schedule_predecessor_id")
        if pid in (None, "", "null"):
            item.schedule_predecessor_id = None
        else:
            try:
                pidi = int(pid)
            except (TypeError, ValueError):
                return JsonResponse({"ok": False, "error": "predecessor"}, status=400)
            pred = (
                EstimateItem.objects.filter(
                    pk=pidi,
                    section__project=project,
                    is_subsection_header=False,
                )
                .exclude(pk=item.pk)
                .first()
            )
            if not pred:
                return JsonResponse({"ok": False, "error": "predecessor"}, status=400)
            item.schedule_predecessor_id = pidi
        touched.add("schedule_predecessor")

    link_updates: list[EstimateItem] = []

    if touched:
        try:
            item.full_clean()
        except ValidationError as e:
            return JsonResponse(
                {"ok": False, "errors": e.message_dict or {"__all__": e.messages}},
                status=400,
            )
        item.save(update_fields=list(touched))

    if "schedule_successor_id" in data:
        sid_raw = data.get("schedule_successor_id")
        try:
            with transaction.atomic():
                if sid_raw in (None, "", "null"):
                    qs = EstimateItem.objects.filter(
                        section__project=project,
                        schedule_predecessor_id=item.pk,
                    ).select_related("section")
                    for dep in qs:
                        dep.schedule_predecessor_id = None
                        dep.full_clean()
                        dep.save(update_fields=["schedule_predecessor"])
                        link_updates.append(dep)
                else:
                    try:
                        sid = int(sid_raw)
                    except (TypeError, ValueError):
                        return JsonResponse(
                            {"ok": False, "error": "successor"}, status=400
                        )
                    succ = (
                        EstimateItem.objects.filter(
                            pk=sid,
                            section__project=project,
                            is_subsection_header=False,
                        )
                        .exclude(pk=item.pk)
                        .select_related("section")
                        .first()
                    )
                    if not succ:
                        return JsonResponse(
                            {"ok": False, "error": "successor"}, status=400
                        )
                    others = (
                        EstimateItem.objects.filter(
                            section__project=project,
                            schedule_predecessor_id=item.pk,
                        )
                        .exclude(pk=succ.pk)
                        .select_related("section")
                    )
                    for dep in others:
                        dep.schedule_predecessor_id = None
                        dep.full_clean()
                        dep.save(update_fields=["schedule_predecessor"])
                        link_updates.append(dep)
                    succ.schedule_predecessor_id = item.pk
                    succ.full_clean()
                    succ.save(update_fields=["schedule_predecessor"])
                    link_updates.append(succ)
        except ValidationError as e:
            return JsonResponse(
                {"ok": False, "errors": e.message_dict or {"__all__": e.messages}},
                status=400,
            )

    if not touched and "schedule_successor_id" not in data:
        return JsonResponse(_schedule_item_api_response_dict(item))

    return JsonResponse(
        _schedule_item_api_response_dict(
            item, link_updates=link_updates if link_updates else None
        )
    )


def _company_assignee_users(company: Company):
    user_ids = CompanyUser.objects.filter(
        company=company, is_active=True
    ).values_list("user_id", flat=True)
    return (
        User.objects.filter(Q(id__in=user_ids) | Q(id=company.owner_id))
        .distinct()
        .order_by("username")
    )


@login_required
def project_supply(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    _ensure_supply_resources(company)
    _ensure_finance_defaults(company)

    tab = request.GET.get("tab", "requests")
    if tab not in ("requests", "off_estimate", "approval", "procurement_approval", "orders"):
        tab = "requests"
    kpi = supply_services.compute_project_supply_kpis(project)

    can_create_supply = has_permission(request.user, company, "create_supply_request")
    can_approve_supply = has_permission(request.user, company, "approve_supply_request")
    can_approve_procurement = has_permission(
        request.user, company, "approve_procurement_payment"
    )
    can_procure_supply = has_permission(request.user, company, "procure_supply") or has_permission(
        request.user, company, "edit_supply"
    )
    can_receive_supply = has_permission(
        request.user, company, "receive_supply_warehouse"
    ) or has_permission(request.user, company, "edit_warehouse")
    can_edit_off_estimate = has_permission(
        request.user, company, "edit_off_estimate_supply"
    ) or can_create_supply
    can_view_off_estimate = (
        has_permission(request.user, company, "view_off_estimate_supply")
        or can_edit_off_estimate
    )
    if tab == "off_estimate" and not can_view_off_estimate:
        tab = "requests"
    if tab == "approval" and not can_approve_supply:
        tab = "requests"
    if tab == "procurement_approval" and not can_approve_procurement:
        tab = "requests"

    requests_qs = project.supply_requests.filter(
        estimate_item__isnull=False
    ).select_related(
        "resource",
        "estimate_item",
        "estimate_item__section",
        "created_by",
    ).prefetch_related("order_item", "order_item__order")

    st = request.GET.get("status", "").strip()
    if st in dict(SupplyRequest.STATUS_CHOICES):
        requests_qs = requests_qs.filter(status=st)
    df = request.GET.get("date_from", "").strip()
    dt_to = request.GET.get("date_to", "").strip()
    d_from = parse_date(df) if df else None
    d_to = parse_date(dt_to) if dt_to else None
    if d_from:
        requests_qs = requests_qs.filter(required_date__gte=d_from)
    if d_to:
        requests_qs = requests_qs.filter(required_date__lte=d_to)
    rtype = request.GET.get("resource_type", "").strip()
    if rtype in dict(Resource.TYPE_CHOICES):
        requests_qs = requests_qs.filter(resource__type=rtype)
    q = request.GET.get("q", "").strip()
    if q:
        requests_qs = requests_qs.filter(resource__name__icontains=q)

    requests_qs = requests_qs.order_by("-required_date", "-created_at")

    orders_qs = (
        SupplyOrder.objects.filter(project=project)
        .select_related("finance_operation", "off_estimate_request")
        .prefetch_related(
            "items__request__resource",
            "items__request__estimate_item",
            "items__off_estimate_item",
            "documents__uploaded_by",
        )
        .order_by("-created_at")
    )

    procurement_approval_orders = []
    if tab == "procurement_approval":
        from . import supply_payment_services as pay_svc

        procurement_approval_orders = pay_svc.list_pending_payment_approval(project)

    approval_data = {}
    project_warehouses = []
    if tab == "approval":
        approval_data = wf_services.list_pending_approval(project)
    if tab == "orders":
        project_warehouses = list(
            Warehouse.objects.filter(
                company=company, is_deleted=False
            ).filter(
                Q(project=project) | Q(project__isnull=True)
            ).order_by("project_id", "name")
        )

    req_initial = {}
    pre_ei = request.GET.get("estimate_item", "").strip()
    supply_selected_item_id = None
    supply_first_selectable_item_pk = None
    if pre_ei.isdigit():
        ei = (
            EstimateItem.objects.filter(
                pk=int(pre_ei),
                section__project=project,
                is_subsection_header=False,
            )
            .first()
        )
        if ei:
            req_initial = {
                "estimate_item": ei,
                "quantity": ei.quantity,
                "required_date": date.today(),
            }
            supply_selected_item_id = ei.pk

    supply_estimate_sections: list = []
    supply_first_selectable_item_pk = None
    if tab == "requests":
        from decimal import Decimal

        purchased_map = supply_services.purchased_qty_by_estimate_item(project)
        eligible_by_section: dict[int, list] = {}
        for item in supply_services.supply_eligible_estimate_items(project):
            eligible_by_section.setdefault(item.section_id, []).append(item)
        sections_qs = (
            EstimateSection.objects.filter(project=project)
            .prefetch_related(
                Prefetch(
                    "items",
                    queryset=EstimateItem.objects.order_by("order", "id"),
                )
            )
            .order_by("order", "id")
        )
        for sec in sections_qs:
            rows = []
            for item in eligible_by_section.get(sec.pk, []):
                rows.append(
                    {
                        "item": item,
                        "purchased_qty": purchased_map.get(item.pk, Decimal("0")),
                    }
                )
            if rows:
                supply_estimate_sections.append({"section": sec, "rows": rows})

        if not supply_selected_item_id:
            for block in supply_estimate_sections:
                for row in block["rows"]:
                    supply_first_selectable_item_pk = row["item"].pk
                    break
                if supply_first_selectable_item_pk:
                    break

    req_form = ProjectSupplyRequestForm(project=project, initial=req_initial)

    off_kpi = {}
    off_estimate_requests = []
    oes_form = ProjectOffEstimateSupplyRequestForm()
    assignee_users = []

    if tab == "off_estimate":
        off_kpi = oes_services.compute_off_estimate_kpis(project)
        off_estimate_requests = oes_services.filter_off_estimate_requests(
            project, request.GET
        )
        assignee_users = _company_assignee_users(company)
        export = request.GET.get("export", "").strip()
        if export == "xlsx":
            buf = oes_services.export_off_estimate_xlsx(
                project, off_estimate_requests
            )
            resp = HttpResponse(
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = (
                f'attachment; filename="off_estimate_{project.pk}.xlsx"'
            )
            return resp
        if export == "pdf":
            buf = oes_services.export_off_estimate_pdf(
                project, off_estimate_requests
            )
            resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
            resp["Content-Disposition"] = (
                f'attachment; filename="off_estimate_{project.pk}.pdf"'
            )
            return resp

    return render(
        request,
        "core/project/supply.html",
        {
            "project": project,
            "active_tab": "supply",
            "supply_tab": tab,
            "kpi": kpi,
            "supply_requests": requests_qs,
            "supply_orders": orders_qs,
            "req_form": req_form,
            "status_choices": SupplyRequest.STATUS_CHOICES,
            "resource_types": Resource.TYPE_CHOICES,
            "order_status_choices": SupplyOrder.STATUS_CHOICES,
            "procurement_status_choices": SupplyOrder.PROCUREMENT_STATUS_CHOICES,
            "payment_status_choices": SupplyOrder.PAYMENT_STATUS_CHOICES,
            "approval_estimate_requests": approval_data.get("estimate_requests", []),
            "approval_off_requests": approval_data.get("off_estimate_requests", []),
            "procurement_approval_orders": procurement_approval_orders,
            "project_warehouses": project_warehouses,
            "can_create_supply": can_create_supply,
            "can_approve_supply": can_approve_supply,
            "can_approve_procurement": can_approve_procurement,
            "can_procure_supply": can_procure_supply,
            "can_receive_supply": can_receive_supply,
            "filter_status": st,
            "filter_date_from": df,
            "filter_date_to": dt_to,
            "filter_resource_type": rtype,
            "filter_q": q,
            "preselect_estimate_item": pre_ei if req_initial else "",
            "supply_estimate_sections": supply_estimate_sections,
            "supply_selected_item_id": supply_selected_item_id,
            "supply_first_selectable_item_pk": supply_first_selectable_item_pk,
            "off_kpi": off_kpi,
            "off_estimate_requests": off_estimate_requests,
            "oes_form": oes_form,
            "oes_status_choices": OffEstimateSupplyRequest.STATUS_CHOICES,
            "oes_priority_choices": OffEstimateSupplyRequest.PRIORITY_CHOICES,
            "oes_unit_presets": OFF_ESTIMATE_UNIT_PRESETS,
            "assignee_users": assignee_users,
            "can_edit_off_estimate": can_edit_off_estimate,
            "can_create_supply": can_create_supply,
            "can_view_off_estimate": can_view_off_estimate,
        },
    )


@login_required
def timesheet_dashboard(request: HttpRequest) -> HttpResponse:
    """Табель компании (без привязки к проекту)."""
    company = get_current_company(request.user)
    if not company:
        return redirect("dashboard")
    if not has_permission(request.user, company, "view_timesheet"):
        return redirect("dashboard")
    from django.middleware.csrf import get_token

    return render(
        request,
        "core/timesheet.html",
        {"company": company, "csrf_token": get_token(request)},
    )


@login_required
def project_timesheet(request: HttpRequest, pk: int) -> HttpResponse:
    return redirect("timesheet_dashboard")


@login_required
@require_POST
@permission_required("create_supply_request")
def project_supply_request_create(
    request: HttpRequest, pk: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    form = ProjectSupplyRequestForm(request.POST, project=project)
    if not form.is_valid():
        for _f, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{_f}: {e}")
        return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=requests")

    item = form.cleaned_data["estimate_item"]
    resource = supply_services.get_or_create_resource_for_estimate_item(company, item)
    from decimal import Decimal

    price = item.sell_price or item.cost_price or Decimal("0")

    sr = SupplyRequest.objects.create(
        company=company,
        project=project,
        resource=resource,
        estimate_item=item,
        required_date=form.cleaned_data["required_date"],
        quantity=form.cleaned_data["quantity"],
        price_plan=price,
        supplier_name="",
        status=SupplyRequest.STATUS_APPROVAL,
        created_by=request.user,
    )
    wf_services.log_estimate_request_created(sr, request.user)
    messages.success(
        request,
        f"Заявка создана и отправлена на согласование: {sr.resource.name}.",
    )
    return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=requests")


@login_required
@require_POST
@permission_required("create_supply_request")
def project_off_estimate_request_create(
    request: HttpRequest, pk: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    form = ProjectOffEstimateSupplyRequestForm(request.POST)
    if not form.is_valid():
        for _f, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{_f}: {e}")
        return redirect(
            f"{reverse('project_supply', args=[project.pk])}?tab=off_estimate"
        )
    lines, line_errors = oes_services.parse_line_rows_from_post(request.POST)
    for e in line_errors:
        messages.error(request, e)
    if line_errors:
        return redirect(
            f"{reverse('project_supply', args=[project.pk])}?tab=off_estimate"
        )
    req = oes_services.create_off_estimate_request(
        company=company,
        project=project,
        user=request.user,
        note=(form.cleaned_data.get("note") or "").strip(),
        required_date=form.cleaned_data.get("required_date"),
        priority=form.cleaned_data["priority"],
        lines=lines,
    )
    wf_services.log_off_estimate_request_created(req, request.user)
    messages.success(
        request,
        f"Заявка {req.number} создана ({len(lines)} поз.) и отправлена на согласование.",
    )
    return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=off_estimate")


@login_required
@permission_required("view_off_estimate_supply")
def project_off_estimate_request_export(
    request: HttpRequest, pk: int, req_id: int
) -> HttpResponse:
    from urllib.parse import quote

    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    req = get_object_or_404(
        OffEstimateSupplyRequest,
        pk=req_id,
        project=project,
        company=project.company,
    )
    buf = export_single_off_estimate_request_xlsx(req)
    filename = export_filename(req)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return resp


@login_required
@require_POST
@permission_required("create_supply_request")
def project_off_estimate_request_delete(
    request: HttpRequest, pk: int, req_id: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    req = get_object_or_404(
        OffEstimateSupplyRequest, pk=req_id, project=project, company=company
    )
    number = req.number
    if req.status != OffEstimateSupplyRequest.STATUS_APPROVAL:
        messages.error(
            request,
            f"Заявку {number} можно удалить только на этапе согласования.",
        )
        return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=off_estimate")
    try:
        oes_services.delete_off_estimate_request(req)
    except ValueError as e:
        if str(e) == "warehouse_received":
            messages.error(
                request,
                f"Заявку {number} нельзя удалить: часть материалов уже принята на склад.",
            )
        else:
            messages.error(request, f"Не удалось удалить заявку {number}.")
    else:
        messages.success(request, f"Заявка {number} удалена.")
    return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=off_estimate")


@login_required
@require_POST
def project_supply_order_create(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    _ensure_finance_defaults(company)
    form = ProjectSupplyOrderCreateForm(request.POST, company=company, project=project)
    if not form.is_valid():
        for _f, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{_f}: {e}")
        return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=orders")

    supplier = form.cleaned_data["supplier"]
    ids = form.cleaned_data["request_ids"]
    if not ids:
        messages.error(request, "Выберите заявки.")
        return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=orders")

    order = SupplyOrder.objects.create(
        company=company,
        project=project,
        supplier=supplier,
        status=SupplyOrder.STATUS_NEW,
        payment_status=SupplyOrder.PAYMENT_DRAFT,
    )
    for rid in ids:
        req = SupplyRequest.objects.filter(
            pk=int(rid), project=project, company=company
        ).first()
        if not req:
            continue
        if SupplyOrderItem.objects.filter(request=req).exists():
            continue
        SupplyOrderItem.objects.create(
            order=order,
            request=req,
            quantity=req.quantity,
            price_fact=req.price_plan,
        )
        req.status = SupplyRequest.STATUS_IN_PROGRESS
        req.save(update_fields=["status"])

    order.recalc_total()
    order.save(update_fields=["total_amount"])

    messages.success(
        request,
        "Заказ создан. Отправьте его на оплату в «Снабжении», затем оплатите в «Финансы проекта».",
    )
    return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=orders")


@login_required
@require_POST
def project_supply_request_api(
    request: HttpRequest, pk: int, req_id: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    sr = get_object_or_404(SupplyRequest, pk=req_id, project=project)
    return JsonResponse(
        {
            "ok": False,
            "error": "readonly",
            "message": "Редактирование заявки после создания запрещено.",
        },
        status=403,
    )


@login_required
@require_POST
def project_supply_order_payment(
    request: HttpRequest, pk: int, order_id: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    order = get_object_or_404(SupplyOrder, pk=order_id, project=project)
    ps = request.POST.get("payment_status", "")
    if ps in dict(SupplyOrder.PAYMENT_STATUS_CHOICES):
        order.payment_status = ps
        order.save(update_fields=["payment_status"])
        messages.success(request, "Статус оплаты обновлён.")
    return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=orders")


@login_required
def project_finance_section(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    _ensure_finance_defaults(company)

    tab = request.GET.get("tab", "journal")
    kpi = finance_project_services.project_finance_kpis(project)

    income_form = ProjectFinanceIncomeForm(project=project)
    expense_form = ProjectFinanceExpenseForm(project=project)
    pay_so_form = ProjectPaySupplyOrderForm(initial={"pay_date": date.today()})
    pay_wa_form = ProjectPayWorkActForm(initial={"pay_date": date.today()})

    journal_qs = FinanceOperation.objects.filter(project=project).select_related(
        "account", "category", "supply_order", "work_act"
    )
    df = request.GET.get("date_from", "").strip()
    dto = request.GET.get("date_to", "").strip()
    contractor_f = request.GET.get("contractor", "").strip()
    op_type = request.GET.get("type", "").strip()
    d_from = parse_date(df) if df else None
    d_to = parse_date(dto) if dto else None
    if d_from:
        journal_qs = journal_qs.filter(date__gte=d_from)
    if d_to:
        journal_qs = journal_qs.filter(date__lte=d_to)
    if op_type in dict(FinanceOperation.TYPE_CHOICES):
        journal_qs = journal_qs.filter(type=op_type)
    if contractor_f:
        journal_qs = journal_qs.filter(contractor__icontains=contractor_f)
    journal_qs = journal_qs.order_by("-date", "-created_at")

    supply_for_payment = _supply_orders_for_payment(company=company, project=project)

    works_for_payment = (
        WorkAct.objects.filter(project=project)
        .filter(
            payment_status__in=(
                WorkAct.PAYMENT_AWAITING,
                WorkAct.PAYMENT_PARTIAL,
            )
        )
        .order_by("-act_date", "-created_at")
    )

    accounts = company.finance_accounts.all().order_by("name")
    categories_inc = company.finance_categories.filter(
        type=FinanceCategory.TYPE_INCOME
    ).order_by("name")
    categories_exp = company.finance_categories.filter(
        type=FinanceCategory.TYPE_EXPENSE
    ).order_by("name")

    return render(
        request,
        "core/project/finance.html",
        {
            "project": project,
            "active_tab": "finance",
            "finance_tab": tab,
            "kpi": kpi,
            "income_form": income_form,
            "expense_form": expense_form,
            "pay_so_form": pay_so_form,
            "pay_wa_form": pay_wa_form,
            "journal_operations": journal_qs,
            "supply_for_payment": supply_for_payment,
            "works_for_payment": works_for_payment,
            "accounts": accounts,
            "categories_inc": categories_inc,
            "categories_exp": categories_exp,
            "filter_date_from": df,
            "filter_date_to": dto,
            "filter_contractor": contractor_f,
            "filter_type": op_type,
            "journal_status_choices": FinanceOperation.JOURNAL_STATUS_CHOICES,
            "type_choices": FinanceOperation.TYPE_CHOICES,
            "payment_status_labels": dict(SupplyOrder.PAYMENT_STATUS_CHOICES),
            "work_payment_labels": dict(WorkAct.PAYMENT_STATUS_CHOICES),
            "today": date.today(),
        },
    )


@login_required
@require_POST
def project_finance_income_create(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    _ensure_finance_defaults(company)
    form = ProjectFinanceIncomeForm(request.POST, project=project)
    if form.is_valid():
        op = form.save(commit=False)
        op.company = company
        op.type = FinanceOperation.TYPE_INCOME
        op.created_by = request.user
        op.basis = FinanceOperation.BASIS_MANUAL
        op.journal_status = FinanceOperation.JOURNAL_PAID
        try:
            op.save()
            messages.success(request, "Доход добавлен в журнал.")
        except ValueError as exc:
            messages.error(request, str(exc))
    else:
        for fld, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{fld}: {e}")
    return redirect(f"{reverse('project_finance_section', args=[project.pk])}?tab=journal")


@login_required
@require_POST
def project_finance_expense_create(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    _ensure_finance_defaults(company)
    form = ProjectFinanceExpenseForm(request.POST, project=project)
    if form.is_valid():
        op = form.save(commit=False)
        op.company = company
        op.type = FinanceOperation.TYPE_EXPENSE
        op.created_by = request.user
        op.basis = FinanceOperation.BASIS_MANUAL
        op.journal_status = FinanceOperation.JOURNAL_PAID
        try:
            op.save()
            messages.success(request, "Расход добавлен в журнал.")
        except ValueError as exc:
            messages.error(request, str(exc))
    else:
        for fld, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{fld}: {e}")
    return redirect(f"{reverse('project_finance_section', args=[project.pk])}?tab=journal")


@login_required
def project_finance_operation_edit(
    request: HttpRequest, pk: int, op_id: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    op = get_object_or_404(
        FinanceOperation, pk=op_id, project=project, company=company
    )
    if request.method == "POST":
        form = ProjectFinanceOperationEditForm(request.POST, instance=op)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Операция обновлена.")
                return redirect(
                    f"{reverse('project_finance_section', args=[project.pk])}?tab=journal"
                )
            except ValueError as excpt:
                messages.error(request, str(excpt))
    else:
        form = ProjectFinanceOperationEditForm(instance=op)
    form.fields["account"].queryset = company.finance_accounts.all().order_by("name")
    if op.type == FinanceOperation.TYPE_INCOME:
        form.fields["category"].queryset = company.finance_categories.filter(
            type=FinanceCategory.TYPE_INCOME
        ).order_by("name")
    elif op.type == FinanceOperation.TYPE_EXPENSE:
        form.fields["category"].queryset = company.finance_categories.filter(
            type=FinanceCategory.TYPE_EXPENSE
        ).order_by("name")
    else:
        form.fields["category"].queryset = company.finance_categories.all().order_by(
            "type", "name"
        )
    return render(
        request,
        "core/project/finance_operation_edit.html",
        {
            "project": project,
            "active_tab": "finance",
            "form": form,
            "operation": op,
        },
    )


@login_required
@require_POST
def project_finance_operation_soft_delete(
    request: HttpRequest, pk: int, op_id: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    op = get_object_or_404(
        FinanceOperation, pk=op_id, project=project, company=project.company
    )
    try:
        op.soft_delete()
        messages.success(request, "Операция удалена из журнала (история сохранена).")
    except Exception as exc:
        messages.error(request, str(exc))
    return redirect(f"{reverse('project_finance_section', args=[project.pk])}?tab=journal")


@login_required
def project_finance_export_journal(request: HttpRequest, pk: int) -> HttpResponse:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    project, err = _get_project_or_403(request, pk)
    if err:
        return err

    journal_qs = FinanceOperation.objects.filter(project=project).select_related(
        "account", "category", "supply_order", "work_act"
    )
    df = request.GET.get("date_from", "").strip()
    dto = request.GET.get("date_to", "").strip()
    op_type = request.GET.get("type", "").strip()
    contractor_f = request.GET.get("contractor", "").strip()
    d_from = parse_date(df) if df else None
    d_to = parse_date(dto) if dto else None
    if d_from:
        journal_qs = journal_qs.filter(date__gte=d_from)
    if d_to:
        journal_qs = journal_qs.filter(date__lte=d_to)
    if op_type in dict(FinanceOperation.TYPE_CHOICES):
        journal_qs = journal_qs.filter(type=op_type)
    if contractor_f:
        journal_qs = journal_qs.filter(contractor__icontains=contractor_f)
    journal_qs = journal_qs.order_by("-date", "-created_at")

    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал"
    headers = [
        "Дата",
        "Тип",
        "Статья",
        "Контрагент",
        "Проект",
        "Сумма",
        "Статус",
        "Основание",
        "Комментарий",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    for row_idx, op in enumerate(journal_qs, 2):
        basis = finance_project_services.operation_basis_label(op)
        ws.cell(row=row_idx, column=1, value=op.date.strftime("%d.%m.%Y"))
        ws.cell(row=row_idx, column=2, value=op.get_type_display())
        ws.cell(
            row=row_idx,
            column=3,
            value=op.category.name if op.category else "",
        )
        ws.cell(row=row_idx, column=4, value=op.contractor)
        ws.cell(row=row_idx, column=5, value=project.name)
        ws.cell(row=row_idx, column=6, value=float(op.amount))
        ws.cell(row=row_idx, column=7, value=op.get_journal_status_display())
        ws.cell(row=row_idx, column=8, value=basis)
        ws.cell(row=row_idx, column=9, value=op.description)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="project_{project.pk}_finance_journal.xlsx"'
    )
    return response


@login_required
@require_POST
def project_finance_pay_supply_order(
    request: HttpRequest, pk: int, order_id: int
) -> HttpResponse:
    from decimal import Decimal

    from . import supply_payment_services as supply_pay

    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    _ensure_finance_defaults(company)
    order = get_object_or_404(SupplyOrder, pk=order_id, project=project)
    return_url = _finance_payment_orders_return_url(request, project_id=project.pk)
    if order.payment_status not in (
        SupplyOrder.PAYMENT_AWAITING,
        SupplyOrder.PAYMENT_PARTIAL,
    ):
        messages.error(request, "Этот заказ не в очереди на оплату.")
        return redirect(return_url)
    form = ProjectPaySupplyOrderForm(request.POST, request.FILES)
    if not form.is_valid():
        for fld, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{fld}: {e}")
        return redirect(return_url)
    amount = form.cleaned_data["amount"]
    pay_date = form.cleaned_data["pay_date"]
    payment_proof = form.cleaned_data["payment_proof"]
    remaining = order.remaining_amount
    if amount > remaining:
        messages.error(request, "Сумма превышает остаток к оплате.")
        return redirect(return_url)
    account = company.finance_accounts.order_by("id").first()
    category = (
        company.finance_categories.filter(
            type=FinanceCategory.TYPE_EXPENSE,
            name__icontains="постав",
        )
        .order_by("id")
        .first()
        or company.finance_categories.filter(type=FinanceCategory.TYPE_EXPENSE)
        .order_by("id")
        .first()
    )
    if not account or not category:
        messages.error(request, "Настройте счёт и статьи расхода в финансах компании.")
        return redirect(return_url)
    pay_target = order.payment_amount_display
    try:
        with transaction.atomic():
            supply_pay.upload_payment_proof(
                order,
                user=request.user,
                uploaded_file=payment_proof,
            )
            op = FinanceOperation(
                company=company,
                account=account,
                project=project,
                category=category,
                type=FinanceOperation.TYPE_EXPENSE,
                amount=amount,
                description=f"Оплата заказа поставщику ({order.supplier})",
                contractor=(order.supplier or "")[:255],
                date=pay_date,
                created_by=request.user,
                basis=FinanceOperation.BASIS_SUPPLY_ORDER,
                supply_order=order,
                journal_status=(
                    FinanceOperation.JOURNAL_PAID
                    if amount >= remaining
                    else FinanceOperation.JOURNAL_PARTIAL
                ),
            )
            op.save()
            paid = (order.paid_amount or Decimal("0")) + amount
            order.paid_amount = paid
            if paid >= pay_target:
                order.payment_status = SupplyOrder.PAYMENT_PAID
            else:
                order.payment_status = SupplyOrder.PAYMENT_PARTIAL
            order.save(update_fields=["paid_amount", "payment_status"])
            from . import expense_journal_services as ej

            ej.create_from_supply_payment(
                order=order,
                finance_operation=op,
                amount=amount,
                pay_date=pay_date,
                user=request.user,
            )
        messages.success(
            request,
            "Оплата зафиксирована: платёжка сохранена, запись добавлена в журнал.",
        )
    except ValueError as exc:
        err = str(exc)
        if err == "no_file":
            messages.error(request, "Прикрепите платёжное поручение.")
        elif err == "bad_extension":
            messages.error(
                request,
                "Допустимые форматы платёжки: pdf, doc, docx, xls, xlsx, jpg, png.",
            )
        else:
            messages.error(request, err)
    return redirect(return_url)


@login_required
@require_POST
def project_finance_upload_power_of_attorney(
    request: HttpRequest, pk: int, order_id: int
) -> HttpResponse:
    from . import supply_payment_services as supply_pay

    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    order = get_object_or_404(SupplyOrder, pk=order_id, project=project)
    return_url = _finance_payment_orders_return_url(request, project_id=project.pk)
    uploaded = request.FILES.get("power_of_attorney")
    try:
        supply_pay.upload_power_of_attorney(
            order,
            user=request.user,
            uploaded_file=uploaded,
        )
        messages.success(request, "Доверенность сохранена.")
    except ValueError as exc:
        err = str(exc)
        if err == "no_file":
            messages.error(request, "Выберите файл доверенности.")
        elif err == "bad_extension":
            messages.error(
                request,
                "Допустимые форматы: PDF, Word (.doc, .docx), Excel (.xls, .xlsx).",
            )
        elif err == "not_in_payment_queue":
            messages.error(request, "Заказ не в очереди на оплату.")
        else:
            messages.error(request, err)
    return redirect(return_url)


@login_required
@require_POST
def project_finance_pay_work_act(
    request: HttpRequest, pk: int, act_id: int
) -> HttpResponse:
    from decimal import Decimal

    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    _ensure_finance_defaults(company)
    act = get_object_or_404(WorkAct, pk=act_id, project=project)
    if act.payment_status not in (
        WorkAct.PAYMENT_AWAITING,
        WorkAct.PAYMENT_PARTIAL,
    ):
        messages.error(request, "Акт не в очереди на оплату.")
        return redirect(
            f"{reverse('project_finance_section', args=[project.pk])}?tab=work_payments"
        )
    form = ProjectPayWorkActForm(request.POST)
    if not form.is_valid():
        for fld, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{fld}: {e}")
        return redirect(
            f"{reverse('project_finance_section', args=[project.pk])}?tab=work_payments"
        )
    amount = form.cleaned_data["amount"]
    pay_date = form.cleaned_data["pay_date"]
    remaining = act.remaining_amount
    if amount > remaining:
        messages.error(request, "Сумма превышает остаток к оплате по акту.")
        return redirect(
            f"{reverse('project_finance_section', args=[project.pk])}?tab=work_payments"
        )
    account = company.finance_accounts.order_by("id").first()
    category = (
        company.finance_categories.filter(
            type=FinanceCategory.TYPE_EXPENSE,
            name__icontains="зарплат",
        )
        .order_by("id")
        .first()
        or company.finance_categories.filter(type=FinanceCategory.TYPE_EXPENSE)
        .order_by("id")
        .first()
    )
    if not account or not category:
        messages.error(request, "Настройте счёт и статьи расхода в финансах компании.")
        return redirect(
            f"{reverse('project_finance_section', args=[project.pk])}?tab=work_payments"
        )
    try:
        with transaction.atomic():
            desc = (act.work_type or "Работы")[:200]
            op = FinanceOperation(
                company=company,
                account=account,
                project=project,
                category=category,
                type=FinanceOperation.TYPE_EXPENSE,
                amount=amount,
                description=f"Оплата по акту: {desc}",
                contractor=(act.contractor or "")[:255],
                date=pay_date,
                created_by=request.user,
                basis=FinanceOperation.BASIS_WORK_ACT,
                work_act=act,
                journal_status=(
                    FinanceOperation.JOURNAL_PAID
                    if amount >= remaining
                    else FinanceOperation.JOURNAL_PARTIAL
                ),
            )
            op.save()
            paid = (act.paid_amount or Decimal("0")) + amount
            act.paid_amount = paid
            tot = act.amount or Decimal("0")
            if paid >= tot:
                act.payment_status = WorkAct.PAYMENT_PAID
            else:
                act.payment_status = WorkAct.PAYMENT_PARTIAL
            act.save(update_fields=["paid_amount", "payment_status"])
        messages.success(request, "Оплата проведена, запись в журнале создана.")
    except ValueError as exc:
        messages.error(request, str(exc))
    return redirect(
        f"{reverse('project_finance_section', args=[project.pk])}?tab=work_payments"
    )


@login_required
@require_POST
def project_supply_order_submit_payment(
    request: HttpRequest, pk: int, order_id: int
) -> HttpResponse:
    """Устаревший URL: перенаправление на согласование оплаты."""
    from django.http import HttpResponseForbidden

    from . import supply_payment_services as pay
    from .supply_workflow_views import _payment_error_message

    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    if not (
        has_permission(request.user, company, "procure_supply")
        or has_permission(request.user, company, "edit_supply")
    ):
        return HttpResponseForbidden()
    order = get_object_or_404(SupplyOrder, pk=order_id, project=project)
    try:
        pay.submit_order_for_payment_approval(order, user=request.user)
        messages.success(request, "Заказ отправлен на согласование оплаты.")
    except ValueError as exc:
        messages.error(request, _payment_error_message(exc))
    return redirect(f"{reverse('project_supply', args=[project.pk])}?tab=orders")


@login_required
@require_POST
def project_work_act_create(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    form = WorkActForm(request.POST)
    if form.is_valid():
        wa = form.save(commit=False)
        wa.company = company
        wa.project = project
        wa.created_by = request.user
        wa.payment_status = WorkAct.PAYMENT_DRAFT
        wa.save()
        messages.success(request, "Акт добавлен.")
    else:
        for fld, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{fld}: {e}")
    return redirect(reverse("project_documents", args=[project.pk]))


@login_required
@require_POST
def project_work_act_submit_payment(
    request: HttpRequest, pk: int, act_id: int
) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    act = get_object_or_404(WorkAct, pk=act_id, project=project)
    if act.payment_status != WorkAct.PAYMENT_DRAFT:
        messages.warning(request, "Акт уже в оплате или закрыт.")
        return redirect(reverse("project_documents", args=[project.pk]))
    act.payment_status = WorkAct.PAYMENT_AWAITING
    act.save(update_fields=["payment_status"])
    messages.success(
        request,
        "Акт отправлен в «Финансы проекта» — «Работы на оплату».",
    )
    return redirect(reverse("project_documents", args=[project.pk]))


@login_required
def project_construction(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    item_qs = EstimateItem.objects.order_by("order", "id")
    for iid in EstimateItem.objects.filter(section__project=project).values_list(
        "pk", flat=True
    ):
        construction_services.recalc_estimate_item_construction(iid)
    sections = (
        EstimateSection.objects.filter(project=project)
        .prefetch_related(Prefetch("items", queryset=item_qs))
        .order_by("order", "id")
    )
    kpi = construction_services.compute_construction_kpis(project)
    readonly = construction_services.construction_is_readonly(request.user, project)
    journal = (
        ConstructionWorkLog.objects.filter(estimate_item__section__project=project)
        .select_related("estimate_item", "estimate_item__section", "created_by")
        .prefetch_related("photos")
        .order_by("-work_date", "-created_at")[:200]
    )
    report_form = ConstructionWorkReportForm(
        initial={"work_date": date.today()},
    )
    return render(
        request,
        "core/project/construction.html",
        {
            "project": project,
            "active_tab": "construction",
            "sections": sections,
            "kpi": kpi,
            "construction_readonly": readonly,
            "construction_journal": journal,
            "construction_report_form": report_form,
        },
    )


@login_required
@require_POST
def project_construction_log(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    if construction_services.construction_is_readonly(request.user, project):
        messages.error(request, "Недостаточно прав для внесения отчётов.")
        return redirect("project_construction", pk=project.pk)
    item_id = request.POST.get("estimate_item_id", "").strip()
    if not item_id.isdigit():
        messages.error(request, "Не указана позиция.")
        return redirect("project_construction", pk=project.pk)
    item = get_object_or_404(
        EstimateItem.objects.select_related("section"),
        pk=int(item_id),
        section__project=project,
    )
    form = ConstructionWorkReportForm(request.POST)
    if not form.is_valid():
        for fld, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{fld}: {e}")
        return redirect("project_construction", pk=project.pk)
    files = [f for f in request.FILES.getlist("photos") if f]
    try:
        construction_services.create_construction_log_with_photos(
            item=item,
            user=request.user,
            work_date=form.cleaned_data["work_date"],
            volume=form.cleaned_data["volume"],
            comment=form.cleaned_data.get("comment") or "",
            files=files,
        )
        messages.success(request, "Отчёт сохранён, факт и статус обновлены.")
    except Exception as exc:
        messages.error(request, str(exc))
    return redirect("project_construction", pk=project.pk)


@login_required
def project_construction_journal_export(request: HttpRequest, pk: int) -> HttpResponse:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    rows = list(
        ConstructionWorkLog.objects.filter(estimate_item__section__project=project)
        .select_related("estimate_item", "estimate_item__section", "created_by")
        .prefetch_related("photos")
        .order_by("-work_date", "-created_at")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал"
    hdr = [
        "Дата",
        "Раздел",
        "Работа",
        "Что сделано (комментарий)",
        "Объём",
        "Кто",
        "Фото (шт.)",
    ]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    for r, log in enumerate(rows, 2):
        u = (
            log.created_by.get_username()
            if log.created_by_id
            else ""
        )
        ws.cell(row=r, column=1, value=log.work_date.strftime("%d.%m.%Y"))
        ws.cell(row=r, column=2, value=log.estimate_item.section.name)
        ws.cell(row=r, column=3, value=log.estimate_item.name or "—")
        ws.cell(row=r, column=4, value=log.comment)
        ws.cell(row=r, column=5, value=float(log.volume))
        ws.cell(row=r, column=6, value=u)
        ws.cell(row=r, column=7, value=len(log.photos.all()))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="project_{project.pk}_construction_journal.xlsx"'
    )
    return response


@login_required
def project_construction_journal_print(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    rows = (
        ConstructionWorkLog.objects.filter(estimate_item__section__project=project)
        .select_related("estimate_item", "estimate_item__section", "created_by")
        .prefetch_related("photos")
        .order_by("work_date", "created_at")
    )
    return render(
        request,
        "core/project/construction_journal_print.html",
        {
            "project": project,
            "journal_rows": list(rows),
        },
    )


@ensure_csrf_cookie
@login_required
def project_warehouses(request: HttpRequest, pk: int) -> HttpResponse:
    """Kanban инвентаря: колонки — склады проекта + «Списано», карточки — единицы инвентаря."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    company = project.company
    from django.db.models import Q
    written_off_wh = get_written_off_warehouse(company)
    warehouses = list(
        Warehouse.objects.filter(company=company, is_deleted=False).filter(
            Q(project=project) | Q(id=written_off_wh.id)
        ).order_by("name")
    )
    if written_off_wh not in warehouses:
        warehouses.append(written_off_wh)
    warehouses.sort(key=lambda w: (0 if w.project_id else 1, w.name))
    warehouse_filter = request.GET.get("warehouse", "")
    status_filter = request.GET.get("search_status", "")
    search_q = request.GET.get("q", "").strip()
    material_cf = material_names_casefold_for_company(company)
    items_qs = WarehouseInventoryItem.objects.filter(company=company).select_related(
        "warehouse"
    ).prefetch_related("logs").exclude(category="material")
    if warehouse_filter.isdigit():
        items_qs = items_qs.filter(warehouse_id=int(warehouse_filter))
    if status_filter:
        items_qs = items_qs.filter(status=status_filter)
    if search_q:
        items_qs = items_qs.filter(name__icontains=search_q)
    items_list = [
        i for i in items_qs
        if is_equipment_inventory_row(i, material_cf)
    ]
    items_by_warehouse = {}
    for w in warehouses:
        items_by_warehouse[w.id] = [i for i in items_list if i.warehouse_id == w.id]
    project_warehouses = [w for w in warehouses if w.id != written_off_wh.id]
    warehouse_columns = [(w, items_by_warehouse.get(w.id, [])) for w in project_warehouses]
    written_off_column = (written_off_wh, items_by_warehouse.get(written_off_wh.id, []))
    total_count = sum(1 for i in items_list if i.status != WarehouseInventoryItem.STATUS_WRITTEN_OFF)
    total_sum = sum((i.purchase_price or 0) for i in items_list if i.status != WarehouseInventoryItem.STATUS_WRITTEN_OFF)
    written_off_count = sum(1 for i in items_list if i.status == WarehouseInventoryItem.STATUS_WRITTEN_OFF)
    written_off_sum = sum((i.purchase_price or 0) for i in items_list if i.status == WarehouseInventoryItem.STATUS_WRITTEN_OFF)
    create_form = WarehouseInventoryCreateForm(company=company, project=project)
    can_edit_project_inventory = has_permission(request.user, company, "edit_warehouse")
    return render(request, "core/project/warehouses.html", {
        "project": project,
        "active_tab": "warehouses",
        "warehouses": warehouses,
        "warehouse_columns": warehouse_columns,
        "written_off_column": written_off_column,
        "warehouse_filter": warehouse_filter,
        "status_filter": status_filter,
        "search_q": search_q,
        "total_count": total_count,
        "total_sum": total_sum,
        "written_off_count": written_off_count,
        "written_off_sum": written_off_sum,
        "create_form": create_form,
        "can_edit_project_inventory": can_edit_project_inventory,
        "written_off_warehouse_id": written_off_wh.id,
    })


@login_required
def project_warehouse_create(request: HttpRequest, pk: int) -> HttpResponse:
    """Создать склад (модальное окно → POST)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    if request.method != "POST":
        return redirect("project_warehouses", pk=project.pk)
    name = request.POST.get("name", "").strip()
    if not name:
        messages.error(request, "Укажите название склада.")
        return redirect("project_warehouses", pk=project.pk)
    Warehouse.objects.create(company=project.company, project=project, name=name)
    messages.success(request, "Склад создан.")
    return redirect("project_warehouses", pk=project.pk)


@login_required
@require_POST
def project_warehouse_delete(request: HttpRequest, pk: int, warehouse_id: int) -> HttpResponse:
    """Мягкое удаление склада проекта (нельзя «Списано» и склад с остатками/инвентарём)."""
    from decimal import Decimal

    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    if not has_permission(request.user, project.company, "edit_warehouse"):
        messages.error(request, "Недостаточно прав для удаления склада.")
        return redirect("project_warehouses", pk=project.pk)
    wh = get_object_or_404(
        Warehouse,
        pk=warehouse_id,
        company=project.company,
        is_deleted=False,
    )
    written_off_wh = get_written_off_warehouse(project.company)
    if wh.id == written_off_wh.id:
        messages.error(request, "Склад «Списано» удалять нельзя.")
        return redirect("project_warehouses", pk=project.pk)
    if wh.project_id != project.pk:
        messages.error(request, "Можно удалять только склады, привязанные к этому проекту.")
        return redirect("project_warehouses", pk=project.pk)
    if WarehouseInventoryItem.objects.filter(warehouse=wh).exists():
        messages.error(
            request,
            "На складе есть инвентарь. Переместите или удалите позиции, затем повторите.",
        )
        return redirect("project_warehouses", pk=project.pk)
    if Stock.objects.filter(warehouse=wh).exclude(quantity=Decimal("0")).exists():
        messages.error(
            request,
            "На складе есть остатки материалов. Обнулите или переместите их.",
        )
        return redirect("project_warehouses", pk=project.pk)
    if FuelStock.objects.filter(warehouse=wh).exclude(quantity=Decimal("0")).exists():
        messages.error(
            request,
            "На складе учитывается ГСМ. Переместите или спишите топливо.",
        )
        return redirect("project_warehouses", pk=project.pk)
    wh.is_deleted = True
    wh.save(update_fields=["is_deleted"])
    messages.success(request, f"Склад «{wh.name}» удалён.")
    return redirect("project_warehouses", pk=project.pk)


@login_required
def project_inventory_create(request: HttpRequest, pk: int) -> HttpResponse:
    """Добавить инвентарь (GET — форма, POST — создание)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    default_warehouse_id = request.GET.get("warehouse") or request.POST.get("warehouse")
    default_wh = None
    if default_warehouse_id and default_warehouse_id.isdigit():
        default_wh = get_object_or_404(Warehouse, pk=int(default_warehouse_id), company=project.company)
    if request.method == "POST":
        form = WarehouseInventoryCreateForm(
            request.POST, request.FILES,
            company=project.company,
            project=project,
            default_warehouse=default_wh,
        )
        if form.is_valid():
            item = form.save(commit=False)
            item.company = project.company
            inv = (item.inventory_number or "").strip()
            if not inv:
                item.inventory_number = allocate_inventory_number(
                    project.company,
                    item.category or WarehouseInventoryItem.CATEGORY_OTHER,
                )
            item.save()
            log_inventory_action(item, InventoryLog.ACTION_CREATED, request.user, "Создан инвентарь")
            ensure_qr_image_file(item, request)
            messages.success(request, "Инвентарь добавлен.")
            return redirect("project_warehouses", pk=project.pk)
    else:
        form = WarehouseInventoryCreateForm(
            company=project.company,
            project=project,
            default_warehouse=default_wh,
        )
    return render(request, "core/inventory_create.html", {
        "project": project,
        "form": form,
        "active_tab": "warehouses",
    })


@login_required
def project_inventory_update(request: HttpRequest, pk: int, item_id: int) -> HttpResponse:
    """Редактировать инвентарь (GET — форма, POST — сохранение)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    item = get_object_or_404(WarehouseInventoryItem, pk=item_id, company=project.company)
    if request.method == "POST":
        form = WarehouseInventoryUpdateForm(request.POST, request.FILES, instance=item)
        if form.is_valid():
            old_status = item.status
            form.save()
            new_status = item.status
            if old_status != new_status:
                if new_status == WarehouseInventoryItem.STATUS_WRITTEN_OFF:
                    written_off_wh = get_written_off_warehouse(project.company)
                    if item.warehouse_id != written_off_wh.id:
                        transfer_inventory_item(item, written_off_wh, request.user)
                log_inventory_action(
                    item, InventoryLog.ACTION_STATUS_CHANGED, request.user,
                    description=f"Статус: {item.get_status_display()}",
                )
            else:
                log_inventory_action(item, InventoryLog.ACTION_UPDATED, request.user, "Изменён")
            messages.success(request, "Инвентарь обновлён.")
            return redirect("project_warehouses", pk=project.pk)
    else:
        form = WarehouseInventoryUpdateForm(instance=item)
    logs = item.logs.select_related("user").order_by("-created_at")[:50]
    can_edit_project_inventory = has_permission(request.user, project.company, "edit_warehouse")
    return render(request, "core/inventory_modal.html", {
        "project": project,
        "item": item,
        "form": form,
        "logs": logs,
        "active_tab": "warehouses",
        "can_edit_project_inventory": can_edit_project_inventory,
    })


@login_required
def project_inventory_transfer(request: HttpRequest, pk: int, item_id: int) -> HttpResponse:
    """Переместить инвентарь на другой склад (POST)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    item = get_object_or_404(WarehouseInventoryItem, pk=item_id, company=project.company)
    if request.method != "POST":
        form = InventoryTransferForm(item=item)
        return render(request, "core/inventory_transfer_modal.html", {
            "project": project,
            "item": item,
            "form": form,
        })
    form = InventoryTransferForm(request.POST, item=item)
    if not form.is_valid():
        messages.error(request, "Выберите склад назначения.")
        return redirect("project_warehouses", pk=project.pk)
    to_warehouse = form.cleaned_data["to_warehouse"]
    try:
        transfer_inventory_item(item, to_warehouse, request.user)
    except Exception as e:
        messages.error(request, str(e))
        return redirect("project_warehouses", pk=project.pk)
    messages.success(request, "Инвентарь перемещён.")
    return redirect("project_warehouses", pk=project.pk)


@login_required
@require_POST
def project_warehouse_inventory_delete(request: HttpRequest, pk: int, item_id: int) -> HttpResponse:
    """Удалить единицу инвентаря проекта (карточка на доске складов)."""
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    if not has_permission(request.user, project.company, "edit_warehouse"):
        messages.error(request, "Недостаточно прав для удаления.")
        return redirect("project_warehouses", pk=project.pk)
    item = get_object_or_404(WarehouseInventoryItem, pk=item_id, company=project.company)
    item.delete()
    messages.success(request, "Инвентарь удалён.")
    return redirect("project_warehouses", pk=project.pk)


@login_required
def project_documents(request: HttpRequest, pk: int) -> HttpResponse:
    project, err = _get_project_or_403(request, pk)
    if err:
        return err
    acts = (
        WorkAct.objects.filter(project=project)
        .select_related("created_by")
        .order_by("-act_date", "-created_at")
    )
    act_form = WorkActForm()
    can_edit_documents = has_permission(request.user, project.company, "edit_projects")
    return render(
        request,
        "core/project/documents.html",
        {
            "project": project,
            "active_tab": "documents",
            "work_acts": acts,
            "act_form": act_form,
            "work_payment_labels": dict(WorkAct.PAYMENT_STATUS_CHOICES),
            "can_edit_documents": can_edit_documents,
        },
    )


@login_required
def project_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """Страница проекта с вкладками Задачи / Финансы / Склад (legacy). Используется для обратной совместимости или как контент внутри разделов."""
    company = get_current_company(request.user)
    if not company:
        return redirect("dashboard")
    project = get_object_or_404(Project, pk=pk, company=company)
    active_tab = request.GET.get("tab", "tasks")

    task_form = TaskQuickForm()
    finance_form = FinanceForm()
    inventory_form = InventoryItemForm()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_task":
            task_form = TaskQuickForm(request.POST)
            if task_form.is_valid():
                task = task_form.save(commit=False)
                task.project = project
                task.created_by = request.user
                task.save()
                return redirect("project_construction", pk=project.pk)
        elif action == "create_finance":
            finance_form = FinanceForm(request.POST)
            if finance_form.is_valid():
                finance = finance_form.save(commit=False)
                finance.project = project
                finance.save()
                return redirect("project_finance_section", pk=project.pk)
        elif action == "create_inventory":
            inventory_form = InventoryItemForm(request.POST)
            if inventory_form.is_valid():
                item = inventory_form.save(commit=False)
                item.project = project
                item.save()
                return redirect("project_warehouses", pk=project.pk)

    context = {
        "project": project,
        "tasks": project.tasks.all().order_by("-created_at"),
        "finances": project.finances.all().order_by("-date"),
        "inventory_items": project.inventory_items.all().order_by("name"),
        "task_form": task_form,
        "finance_form": finance_form,
        "inventory_form": inventory_form,
        "active_tab": active_tab,
    }
    return render(request, "core/project_detail.html", context)


# ---------- Модуль «Задачи» (страница /tasks/) ----------


def _get_user_company_tasks(request: HttpRequest):
    """Задачи только в рамках компании текущего пользователя."""
    company = Company.objects.filter(owner=request.user).first()
    if not company:
        return Task.objects.none()
    return Task.objects.filter(project__company=company).select_related(
        "project", "assigned_to", "created_by"
    )


@login_required
def task_list(request: HttpRequest) -> HttpResponse:
    """Список задач с фильтрацией по проекту, статусу, ответственному и сортировкой."""
    tasks = _get_user_company_tasks(request)

    # Фильтры
    project_id = request.GET.get("project")
    if project_id:
        tasks = tasks.filter(project_id=project_id)
    status_filter = request.GET.get("status")
    if status_filter:
        tasks = tasks.filter(status=status_filter)
    assigned_id = request.GET.get("assigned")
    if assigned_id:
        tasks = tasks.filter(assigned_to_id=assigned_id)

    # Сортировка
    sort = request.GET.get("sort", "created_at")
    if sort == "due_date":
        tasks = tasks.order_by("due_date", "-created_at")
    else:
        tasks = tasks.order_by("-created_at")

    company = Company.objects.filter(owner=request.user).first()
    projects = company.projects.all().order_by("name") if company else []
    # Пользователи для фильтра «Ответственный»: владелец + все, кому назначены задачи в компании
    user_ids = set()
    if company:
        user_ids.add(company.owner_id)
        user_ids.update(
            Task.objects.filter(project__company=company).values_list("assigned_to_id", flat=True)
        )
    user_ids.discard(None)
    User = request.user.__class__
    assignee_choices = list(User.objects.filter(id__in=user_ids).order_by("username").values_list("id", "username"))

    try:
        filter_project_id = int(project_id) if project_id else None
    except (TypeError, ValueError):
        filter_project_id = None
    try:
        filter_assigned_id = int(assigned_id) if assigned_id else None
    except (TypeError, ValueError):
        filter_assigned_id = None

    context = {
        "tasks": tasks,
        "projects": projects,
        "assignee_choices": assignee_choices,
        "filter_project_id": filter_project_id,
        "filter_status": status_filter or "",
        "filter_assigned_id": filter_assigned_id,
        "sort": sort,
    }
    return render(request, "core/task_list.html", context)


@login_required
def task_create(request: HttpRequest) -> HttpResponse:
    """Создание задачи (выбор проекта из своей компании)."""
    company = Company.objects.filter(owner=request.user).first()
    if not company:
        messages.warning(request, "Сначала создайте компанию или проект.")
        return redirect("dashboard")
    if request.method == "POST":
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.created_by = request.user
            task.save()
            messages.success(request, "Задача создана.")
            return redirect("task_list")
    else:
        form = TaskForm()
        form.fields["project"].queryset = company.projects.all().order_by("name")
        _user_ids = {company.owner_id}
        _user_ids.update(
            Task.objects.filter(project__company=company).values_list("assigned_to_id", flat=True)
        )
        _user_ids.discard(None)
        form.fields["assigned_to"].queryset = request.user.__class__.objects.filter(
            id__in=_user_ids
        ).order_by("username")
    return render(request, "core/task_form.html", {"form": form, "title": "Новая задача"})


@login_required
def task_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """Редактирование задачи."""
    task = get_object_or_404(Task, pk=pk, project__company__owner=request.user)
    company = task.project.company
    if request.method == "POST":
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, "Задача обновлена.")
            return redirect("task_list")
    else:
        form = TaskForm(instance=task)
    form.fields["project"].queryset = company.projects.all().order_by("name")
    _user_ids = {company.owner_id}
    _user_ids.update(
        Task.objects.filter(project__company=company).values_list("assigned_to_id", flat=True)
    )
    _user_ids.discard(None)
    form.fields["assigned_to"].queryset = request.user.__class__.objects.filter(
        id__in=_user_ids
    ).order_by("username")
    return render(request, "core/task_form.html", {"form": form, "task": task, "title": "Редактирование задачи"})


@login_required
@require_POST
def task_status(request: HttpRequest, pk: int) -> HttpResponse:
    """Смена статуса задачи без перезагрузки страницы (POST → JSON)."""
    task = get_object_or_404(Task, pk=pk, project__company__owner=request.user)
    new_status = request.POST.get("status")
    allowed = dict(Task.STATUS_CHOICES).keys()
    if new_status not in allowed:
        return JsonResponse({"ok": False, "error": "invalid status"}, status=400)
    task.status = new_status
    task.save(update_fields=["status"])
    return JsonResponse({
        "ok": True,
        "status": task.status,
        "status_display": task.get_status_display(),
    })


@login_required
def company_settings(request: HttpRequest) -> HttpResponse:
    company = Company.objects.filter(owner=request.user).first()
    if request.method == "POST":
        form = CompanyForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            obj = form.save(commit=False)
            if company is None:
                obj.owner = request.user
                obj.save()
            else:
                form.save()
            messages.success(request, "Настройки компании обновлены.")
            return redirect("company_settings")
    else:
        form = CompanyForm(instance=company)

    company = Company.objects.filter(owner=request.user).first()
    return render(
        request,
        "core/company_settings.html",
        {"form": form, "company": company},
    )


# ---------- Настройки → Права доступа ----------


def _get_settings_company(request):
    """Компания для раздела настроек (текущая компания пользователя)."""
    return get_current_company(request.user)


@login_required
def settings_access(request: HttpRequest) -> HttpResponse:
    """Страница /settings/access/: вкладки Пользователи и Права и роли."""
    company = _get_settings_company(request)
    if not company:
        return redirect("dashboard")
    if not can_manage_access(request.user, company):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("<h1>403</h1><p>Недостаточно прав.</p>")

    from .company_roles import ensure_company_default_roles

    ensure_company_default_roles(company)

    active_tab = request.GET.get("tab", "users")
    if active_tab not in ("users", "roles"):
        active_tab = "users"

    company_users = (
        CompanyUser.objects.filter(company=company)
        .select_related("user", "role")
        .prefetch_related("project_accesses__project")
        .order_by("user__email")
    )
    # Кол-во проектов по каждому company_user
    user_rows = []
    for cu in company_users:
        project_count = cu.project_accesses.count()
        user_rows.append({
            "company_user": cu,
            "project_count": project_count,
        })

    roles = company.roles.all().order_by("is_system", "name")
    add_form = AddCompanyUserForm(company=company)

    return render(request, "core/settings_access.html", {
        "company": company,
        "active_tab": active_tab,
        "user_rows": user_rows,
        "roles": roles,
        "add_form": add_form,
    })


@login_required
@require_POST
def settings_access_add_user(request: HttpRequest) -> HttpResponse:
    """POST: добавление пользователя в компанию (из модального окна)."""
    company = _get_settings_company(request)
    if not company or not can_manage_access(request.user, company):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from .company_roles import ensure_company_default_roles

    ensure_company_default_roles(company)

    form = AddCompanyUserForm(request.POST, company=company)
    if not form.is_valid():
        messages.error(request, "Исправьте ошибки в форме.")
        return redirect("settings_access")

    from django.contrib.auth import get_user_model
    User = get_user_model()
    email = form.cleaned_data["email"].strip().lower()
    role = form.cleaned_data["role"]
    project_ids = form.cleaned_data.get("projects") or []
    role_in_project = form.cleaned_data.get("role_in_project") or ProjectAccess.ROLE_VIEWER
    auto_add = form.cleaned_data.get("auto_add_to_new_projects") or False

    # Только владелец может назначать роль «Владелец»
    if role.slug == CompanyRole.SLUG_OWNER and not (company.owner_id == request.user.id):
        messages.error(request, "Назначать роль «Владелец компании» может только владелец.")
        return redirect("settings_access")

    user = User.objects.filter(email__iexact=email).first()
    if not user:
        username = email[:150]
        n = 0
        while User.objects.filter(username=username).exists():
            n += 1
            suffix = f"_{n}"
            username = f"{email[: max(0, 150 - len(suffix))]}{suffix}"
        user = User(username=username, email=email, is_active=True)
        user.set_unusable_password()
        user.save()

    if CompanyUser.objects.filter(user=user, company=company).exists():
        messages.error(request, "Пользователь с таким email уже добавлен в компанию.")
        return redirect("settings_access")

    from .subscription_limits import can_add_company_user

    ok_u, err_u = can_add_company_user(
        company,
    )
    if not ok_u:
        messages.error(
            request,
            err_u or "Нельзя добавить пользователя.",
        )
        return redirect(
            "settings_access",
        )

    company_user = CompanyUser.objects.create(
        user=user,
        company=company,
        role=role,
        is_active=True,
        auto_add_to_new_projects=auto_add,
    )

    for pid in project_ids:
        try:
            project = company.projects.get(pk=int(pid))
            ProjectAccess.objects.get_or_create(
                company_user=company_user,
                project=project,
                defaults={"role_in_project": role_in_project},
            )
        except (ValueError, Project.DoesNotExist):
            pass

    messages.success(
        request,
        "Пользователь добавлен. Откройте «Изменить» и задайте пароль — "
        "сотрудник войдёт с email как логином.",
    )
    return redirect("settings_access")


@login_required
def settings_access_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """Редактирование пользователя компании (роль, статус, авто-добавление в проекты)."""
    company = _get_settings_company(request)
    if not company or not can_manage_access(request.user, company):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    company_user = get_object_or_404(CompanyUser, pk=pk, company=company)

    from .company_roles import ensure_company_default_roles

    ensure_company_default_roles(company)

    allow_set_password = company_user.user_id != request.user.id
    if request.method == "POST":
        form = EditCompanyUserForm(
            request.POST,
            instance=company_user,
            company=company,
            allow_set_password=allow_set_password,
        )
        if form.is_valid():
            new_role = form.cleaned_data["role"]
            if new_role and new_role.slug == CompanyRole.SLUG_OWNER and company.owner_id != request.user.id:
                messages.error(request, "Назначать роль «Владелец» может только владелец.")
                return redirect("settings_access")
            form.save()
            messages.success(request, "Изменения сохранены.")
            return redirect("settings_access")
    else:
        form = EditCompanyUserForm(
            instance=company_user,
            company=company,
            allow_set_password=allow_set_password,
        )
    return render(request, "core/settings_access_edit.html", {
        "form": form,
        "company_user": company_user,
        "company": company,
        "allow_set_password": allow_set_password,
    })


@login_required
@require_POST
def settings_access_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Удаление пользователя из компании. Нельзя удалить последнего владельца."""
    company = _get_settings_company(request)
    if not company or not can_manage_access(request.user, company):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    company_user = get_object_or_404(CompanyUser, pk=pk, company=company)
    if company_user.is_owner_role:
        owner_count = CompanyUser.objects.filter(
            company=company, role__slug=CompanyRole.SLUG_OWNER
        ).count()
        if owner_count <= 1:
            messages.error(request, "Нельзя удалить последнего владельца компании.")
            return redirect("settings_access")
    company_user.delete()
    messages.success(request, "Пользователь удалён из компании.")
    return redirect("settings_access")


@login_required
def delete_task(request: HttpRequest, pk: int) -> HttpResponse:
    task = get_object_or_404(Task, pk=pk, project__company__owner=request.user)
    task.delete()
    messages.success(request, "Задача удалена.")
    # Редирект на список задач; с project_detail можно вернуться по ссылке «Проекты»
    return redirect("task_list")


@login_required
def delete_finance(request: HttpRequest, pk: int) -> HttpResponse:
    finance = get_object_or_404(Finance, pk=pk, project__company__owner=request.user)
    project_id = finance.project_id
    finance.delete()
    return redirect("project_finance_section", pk=project_id)


@login_required
def delete_inventory_item(request: HttpRequest, pk: int) -> HttpResponse:
    item = get_object_or_404(
        InventoryItem, pk=pk, project__company__owner=request.user
    )
    project_id = item.project_id
    item.delete()
    return redirect("project_warehouses", pk=project_id)


# ---------- Модуль «Финансы»: журнал операций ----------


def _get_finance_company(request):
    """Компания текущего пользователя; только её финансы видны."""
    return get_current_company(request.user)


def _ensure_finance_defaults(company):
    """При первом заходе в раздел — создать один счёт и базовые статьи."""
    if not company:
        return
    if not company.finance_accounts.exists():
        Account.objects.create(company=company, name="Расчётный счёт", balance=0, currency="KZT")
    if not company.finance_categories.exists():
        for name, cat_type in [
            ("Оплата поставщикам", FinanceCategory.TYPE_EXPENSE),
            ("Зарплата", FinanceCategory.TYPE_EXPENSE),
            ("Материалы", FinanceCategory.TYPE_EXPENSE),
            ("Оплата от заказчика", FinanceCategory.TYPE_INCOME),
            ("Прочий доход", FinanceCategory.TYPE_INCOME),
        ]:
            FinanceCategory.objects.create(company=company, name=name, type=cat_type)


def _supply_orders_for_payment(*, company, project=None, project_id=None):
    """Заказы снабжения, согласованные к оплате (очередь бухгалтерии)."""
    from django.db.models import Case, IntegerField, Prefetch, When

    from .models import SupplyOrderDocument

    qs = SupplyOrder.objects.filter(company=company)
    if project is not None:
        qs = qs.filter(project=project)
    elif project_id:
        qs = qs.filter(project_id=project_id)
    return (
        qs.filter(payment_approved_at__isnull=False)
        .filter(
            payment_status__in=(
                SupplyOrder.PAYMENT_AWAITING,
                SupplyOrder.PAYMENT_PARTIAL,
                SupplyOrder.PAYMENT_PAID,
            )
        )
        .select_related("project", "off_estimate_request")
        .prefetch_related(
            Prefetch(
                "documents",
                queryset=SupplyOrderDocument.objects.select_related(
                    "uploaded_by"
                ).order_by("doc_type", "-version", "-id"),
            ),
            "items__request__resource",
            "items__off_estimate_item",
        )
        .annotate(
            _pay_sort=Case(
                When(payment_status=SupplyOrder.PAYMENT_AWAITING, then=0),
                When(payment_status=SupplyOrder.PAYMENT_PARTIAL, then=1),
                When(payment_status=SupplyOrder.PAYMENT_PAID, then=2),
                default=3,
                output_field=IntegerField(),
            )
        )
        .order_by("_pay_sort", "-payment_approved_at", "-created_at")
    )


def _finance_payment_orders_return_url(request, *, project_id=None):
    """Куда вернуться после оплаты заказа снабжения."""
    from django.utils.http import url_has_allowed_host_and_scheme

    next_url = (request.GET.get("next") or request.POST.get("next") or "").strip()
    if next_url and url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}
    ):
        return next_url
    if project_id:
        return (
            f"{reverse('project_finance_section', args=[project_id])}?tab=payment_orders"
        )
    return f"{reverse('finance_dashboard')}?tab=payment_orders"


@login_required
def finance_dashboard(request: HttpRequest) -> HttpResponse:
    """Страница Финансы: карточки счетов, журнал расходов, заказы на оплату."""
    company = _get_finance_company(request)
    if not company:
        messages.warning(request, "Создайте компанию или проект.")
        return redirect("dashboard")
    _ensure_finance_defaults(company)

    accounts = company.finance_accounts.all().order_by("name")
    active_tab = request.GET.get("tab", "expense_journal")
    if active_tab == "journal":
        active_tab = "expense_journal"

    projects = company.projects.all().order_by("name")

    payment_project_param = request.GET.get("payment_project", "").strip()
    filter_payment_project_id = (
        int(payment_project_param)
        if payment_project_param.isdigit()
        else None
    )
    supply_for_payment = _supply_orders_for_payment(
        company=company,
        project_id=filter_payment_project_id,
    )
    pay_return_query = "/finance/?tab=payment_orders"
    if filter_payment_project_id:
        pay_return_query += f"&payment_project={filter_payment_project_id}"

    from .access_utils import has_permission

    context = {
        "company": company,
        "accounts": accounts,
        "projects": projects,
        "active_tab": active_tab,
        "filter_payment_project_id": filter_payment_project_id,
        "supply_for_payment": supply_for_payment,
        "pay_return_query": pay_return_query,
        "today": date.today(),
        "can_edit_expense_journal": has_permission(
            request.user, company, "edit_finance"
        ),
    }
    return render(request, "core/finance_dashboard.html", context)


@login_required
def finance_income(request: HttpRequest) -> HttpResponse:
    """➕ Доход: создание операции типа income."""
    company = _get_finance_company(request)
    if not company:
        return redirect("finance_dashboard")
    _ensure_finance_defaults(company)

    if request.method == "POST":
        form = FinanceIncomeForm(request.POST)
        if form.is_valid():
            op = form.save(commit=False)
            op.company = company
            op.type = FinanceOperation.TYPE_INCOME
            op.created_by = request.user
            op.save()
            messages.success(request, "Доход записан.")
            return redirect("finance_dashboard")
    else:
        form = FinanceIncomeForm()
    form.fields["account"].queryset = company.finance_accounts.all().order_by("name")
    form.fields["category"].queryset = company.finance_categories.filter(type=FinanceCategory.TYPE_INCOME).order_by("name")
    form.fields["project"].queryset = company.projects.all().order_by("name")
    return render(request, "core/finance_operation_form.html", {
        "form": form,
        "title": "➕ Доход",
        "operation_type": "income",
    })


@login_required
def finance_expense(request: HttpRequest) -> HttpResponse:
    """➖ Расход: создание операции типа expense (проверка баланса в модели)."""
    company = _get_finance_company(request)
    if not company:
        return redirect("finance_dashboard")
    _ensure_finance_defaults(company)

    if request.method == "POST":
        form = FinanceExpenseForm(request.POST)
        if form.is_valid():
            op = form.save(commit=False)
            op.company = company
            op.type = FinanceOperation.TYPE_EXPENSE
            op.created_by = request.user
            try:
                op.save()
                messages.success(request, "Расход записан.")
                return redirect("finance_dashboard")
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = FinanceExpenseForm()
    form.fields["account"].queryset = company.finance_accounts.all().order_by("name")
    form.fields["category"].queryset = company.finance_categories.filter(type=FinanceCategory.TYPE_EXPENSE).order_by("name")
    form.fields["project"].queryset = company.projects.all().order_by("name")
    return render(request, "core/finance_operation_form.html", {
        "form": form,
        "title": "➖ Расход",
        "operation_type": "expense",
    })


@login_required
def finance_transfer(request: HttpRequest) -> HttpResponse:
    """🔁 Перевод: создание операции типа transfer между счетами."""
    company = _get_finance_company(request)
    if not company:
        return redirect("finance_dashboard")
    _ensure_finance_defaults(company)

    if request.method == "POST":
        form = FinanceTransferForm(request.POST)
        if form.is_valid():
            op = form.save(commit=False)
            op.company = company
            op.type = FinanceOperation.TYPE_TRANSFER
            op.created_by = request.user
            try:
                op.save()
                messages.success(request, "Перевод выполнен.")
                return redirect("finance_dashboard")
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = FinanceTransferForm()
    form.fields["account"].queryset = company.finance_accounts.all().order_by("name")
    form.fields["account_to"].queryset = company.finance_accounts.all().order_by("name")
    return render(request, "core/finance_operation_form.html", {
        "form": form,
        "title": "🔁 Перевод",
        "operation_type": "transfer",
    })


@login_required
@require_POST
def finance_operation_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """Мягкое удаление операции (баланс откатывается)."""
    op = get_object_or_404(FinanceOperation, pk=pk, company__owner=request.user)
    try:
        op.soft_delete()
        messages.success(request, "Операция удалена из журнала (история сохранена).")
    except Exception as exc:
        messages.error(request, str(exc))
    return redirect("finance_dashboard")


# ---------- Модуль «Снабжение»: заявки и заказы ----------


def _get_supply_company(request):
    """Компания текущего пользователя для раздела Снабжение."""
    return get_current_company(request.user)


def _ensure_supply_resources(company):
    """Минимум один ресурс для выбора в заявках."""
    if company and not company.supply_resources.exists():
        Resource.objects.create(
            company=company,
            name="Прочие материалы",
            type=Resource.TYPE_MATERIAL,
            unit="шт.",
        )


@login_required
def supply_dashboard(request: HttpRequest) -> HttpResponse:
    """Главная страница Снабжения: вкладки Заявки / Заказы, таблица заявок с фильтрами."""
    company = _get_supply_company(request)
    if not company:
        messages.warning(request, "Создайте компанию или проект.")
        return redirect("dashboard")
    _ensure_supply_resources(company)

    active_tab = request.GET.get("tab", "requests")
    # Фильтры заявок
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    project_id = request.GET.get("project")
    resource_type = request.GET.get("resource_type")
    without_order = request.GET.get("without_order") == "1"

    requests_qs = company.supply_requests.all().select_related(
        "project", "resource", "created_by"
    ).prefetch_related("order_item", "order_item__order").order_by("-required_date", "-created_at")

    if date_from:
        requests_qs = requests_qs.filter(required_date__gte=date_from)
    if date_to:
        requests_qs = requests_qs.filter(required_date__lte=date_to)
    if project_id:
        requests_qs = requests_qs.filter(project_id=project_id)
    if resource_type:
        requests_qs = requests_qs.filter(resource__type=resource_type)
    if without_order:
        requests_qs = requests_qs.filter(order_item__isnull=True)

    orders = company.supply_orders.all().prefetch_related("items", "items__request", "items__request__resource").order_by("-created_at")
    projects = company.projects.all().order_by("name")

    filter_project_id = int(project_id) if project_id and str(project_id).isdigit() else None
    from datetime import date
    today = date.today()

    context = {
        "company": company,
        "today": today,
        "requests_list": requests_qs,
        "orders": orders,
        "projects": projects,
        "active_tab": active_tab,
        "filter_date_from": date_from or "",
        "filter_date_to": date_to or "",
        "filter_project_id": filter_project_id,
        "filter_resource_type": resource_type or "",
        "filter_without_order": without_order,
    }
    return render(request, "core/supply_dashboard.html", context)


@login_required
def supply_request_create(request: HttpRequest) -> HttpResponse:
    """➕ Заявка: создание заявки на снабжение."""
    company = _get_supply_company(request)
    if not company:
        return redirect("supply_dashboard")
    _ensure_supply_resources(company)

    if request.method == "POST":
        form = SupplyRequestForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.company = company
            obj.created_by = request.user
            obj.save()
            messages.success(request, "Заявка создана.")
            return redirect("supply_dashboard")
    else:
        form = SupplyRequestForm()
    form.fields["project"].queryset = company.projects.all().order_by("name")
    form.fields["resource"].queryset = company.supply_resources.all().order_by("type", "name")
    return render(request, "core/supply_request_form.html", {"form": form})


@login_required
def supply_order_create(request: HttpRequest) -> HttpResponse:
    """Создание заказа из выбранных заявок."""
    company = _get_supply_company(request)
    if not company:
        return redirect("supply_dashboard")

    if request.method == "POST":
        form = SupplyOrderCreateForm(request.POST, company=company)
        if form.is_valid():
            supplier = form.cleaned_data["supplier"]
            request_ids = form.cleaned_data["request_ids"]
            if not request_ids:
                messages.error(request, "Выберите хотя бы одну заявку.")
                return redirect("supply_order_create")
            order = SupplyOrder.objects.create(company=company, supplier=supplier, status=SupplyOrder.STATUS_NEW)
            for rid in request_ids:
                req = SupplyRequest.objects.get(pk=rid, company=company)
                SupplyOrderItem.objects.create(
                    order=order,
                    request=req,
                    quantity=req.quantity,
                    price_fact=req.price_plan,
                )
                req.status = SupplyRequest.STATUS_IN_PROGRESS
                req.save(update_fields=["status"])
            messages.success(request, "Заказ создан.")
            return redirect("supply_order_detail", pk=order.pk)
    else:
        form = SupplyOrderCreateForm(company=company)
    return render(request, "core/supply_order_create.html", {"form": form, "company": company})


@login_required
def supply_order_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """Просмотр заказа и смена статуса."""
    order = get_object_or_404(
        SupplyOrder.objects.prefetch_related("items__request__resource", "items__request__project"),
        pk=pk,
        company__owner=request.user,
    )
    if request.method == "POST":
        new_status = request.POST.get("status")
        if new_status in dict(SupplyOrder.STATUS_CHOICES):
            order.status = new_status
            order.save(update_fields=["status"])
            messages.success(request, "Статус заказа обновлён.")
            return redirect("supply_order_detail", pk=order.pk)
    return render(request, "core/supply_order_detail.html", {
        "order": order,
        "status_choices": SupplyOrder.STATUS_CHOICES,
    })


# ---------- Модуль «Склады» ----------


def _get_warehouse_company(request):
    """Компания текущего пользователя для раздела Склады."""
    return get_current_company(request.user)


@login_required
def warehouses_dashboard(request: HttpRequest) -> HttpResponse:
    """Дашборд складов: общий остаток, таблица складов, вкладки Операции / Заказы / Остатки."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")

    warehouses_qs = company.warehouses.all().prefetch_related("stock_items__resource")
    total_balance = sum(
        si.total_sum for w in warehouses_qs for si in w.stock_items.all()
    )

    # Остатки по складам: склад -> (позиций, сумма)
    warehouse_stats = []
    for w in warehouses_qs:
        items = list(w.stock_items.all())
        positions = len([i for i in items if i.quantity > 0])
        total = sum(i.total_sum for i in items)
        warehouse_stats.append({"warehouse": w, "positions": positions, "total": total})

    operations = (
        WarehouseOperation.objects.filter(company=company)
        .select_related("warehouse", "from_warehouse", "to_warehouse", "resource", "order", "created_by")
        .order_by("-created_at")[:200]
    )
    balances = StockItem.objects.filter(warehouse__company=company, quantity__gt=0).select_related(
        "warehouse", "resource"
    ).order_by("warehouse__name", "resource__name")

    # Заказы снабжения (для вкладки «Заказы»)
    supply_orders = SupplyOrder.objects.filter(company=company).select_related().order_by("-created_at")[:100]

    tab = request.GET.get("tab", "operations")
    if tab not in ("operations", "orders", "balances"):
        tab = "operations"
    return render(request, "core/warehouses_dashboard.html", {
        "company": company,
        "warehouses": warehouses_qs,
        "warehouse_stats": warehouse_stats,
        "total_balance": total_balance,
        "operations": operations,
        "balances": balances,
        "supply_orders": supply_orders,
        "active_tab": tab,
    })


def _limit_warehouse_choices(form, company):
    """Ограничить выбор складов/ресурсов/заказов компанией."""
    form.fields["warehouse"].queryset = company.warehouses.all().order_by("name")
    form.fields["resource"].queryset = company.supply_resources.all().order_by("type", "name")
    if "order" in form.fields:
        form.fields["order"].queryset = SupplyOrder.objects.filter(company=company).order_by("-created_at")
        form.fields["order"].required = False


@login_required
def warehouses_incoming(request: HttpRequest) -> HttpResponse:
    """Поступление на склад."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")

    if request.method == "POST":
        form = WarehouseIncomingForm(request.POST)
        _limit_warehouse_choices(form, company)
        if form.is_valid():
            op = form.save(commit=False)
            op.company = company
            op.operation_type = WarehouseOperation.TYPE_INCOMING
            op.created_by = request.user
            op.save()
            messages.success(request, "Поступление проведено.")
            return redirect("warehouses_dashboard")
    else:
        form = WarehouseIncomingForm()
        _limit_warehouse_choices(form, company)
    return render(request, "core/warehouse_operation_form.html", {
        "form": form,
        "company": company,
        "operation_label": "Поступление",
        "form_url": reverse("warehouses_incoming"),
    })


@login_required
def warehouses_outgoing(request: HttpRequest) -> HttpResponse:
    """Списание со склада."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")

    if request.method == "POST":
        form = WarehouseOutgoingForm(request.POST)
        _limit_warehouse_choices(form, company)
        if form.is_valid():
            warehouse = form.cleaned_data["warehouse"]
            resource = form.cleaned_data["resource"]
            quantity = form.cleaned_data["quantity"]
            stock = StockItem.objects.filter(warehouse=warehouse, resource=resource).first()
            if not stock or stock.quantity < quantity:
                messages.error(request, "Недостаточно остатка на складе.")
                form = WarehouseOutgoingForm(request.POST)
                _limit_warehouse_choices(form, company)
                return render(request, "core/warehouse_operation_form.html", {
                    "form": form,
                    "company": company,
                    "operation_label": "Списание",
                    "form_url": reverse("warehouses_outgoing"),
                })
            op = form.save(commit=False)
            op.company = company
            op.operation_type = WarehouseOperation.TYPE_OUTGOING
            op.price = stock.price_avg
            op.created_by = request.user
            op.save()
            messages.success(request, "Списание проведено.")
            return redirect("warehouses_dashboard")
    else:
        form = WarehouseOutgoingForm()
        _limit_warehouse_choices(form, company)
    return render(request, "core/warehouse_operation_form.html", {
        "form": form,
        "company": company,
        "operation_label": "Списание",
        "form_url": reverse("warehouses_outgoing"),
    })


@login_required
def warehouses_transfer(request: HttpRequest) -> HttpResponse:
    """Перемещение между складами."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")

    if request.method == "POST":
        form = WarehouseTransferForm(request.POST)
        form.fields["from_warehouse"].queryset = company.warehouses.all().order_by("name")
        form.fields["to_warehouse"].queryset = company.warehouses.all().order_by("name")
        form.fields["resource"].queryset = company.supply_resources.all().order_by("type", "name")
        if form.is_valid():
            from_wh = form.cleaned_data["from_warehouse"]
            to_wh = form.cleaned_data["to_warehouse"]
            if from_wh == to_wh:
                messages.error(request, "Выберите разные склады.")
                form = WarehouseTransferForm(request.POST)
                form.fields["from_warehouse"].queryset = company.warehouses.all().order_by("name")
                form.fields["to_warehouse"].queryset = company.warehouses.all().order_by("name")
                form.fields["resource"].queryset = company.supply_resources.all().order_by("type", "name")
                return render(request, "core/warehouse_operation_form.html", {
                    "form": form,
                    "company": company,
                    "operation_label": "Перемещение",
                    "form_url": reverse("warehouses_transfer"),
                })
            stock = StockItem.objects.filter(warehouse=from_wh, resource=form.cleaned_data["resource"]).first()
            quantity = form.cleaned_data["quantity"]
            if not stock or stock.quantity < quantity:
                messages.error(request, "Недостаточно остатка на складе-источнике.")
                form = WarehouseTransferForm(request.POST)
                form.fields["from_warehouse"].queryset = company.warehouses.all().order_by("name")
                form.fields["to_warehouse"].queryset = company.warehouses.all().order_by("name")
                form.fields["resource"].queryset = company.supply_resources.all().order_by("type", "name")
                return render(request, "core/warehouse_operation_form.html", {
                    "form": form,
                    "company": company,
                    "operation_label": "Перемещение",
                    "form_url": reverse("warehouses_transfer"),
                })
            price = form.cleaned_data.get("price") or (stock.price_avg if stock else 0)
            op = form.save(commit=False)
            op.company = company
            op.operation_type = WarehouseOperation.TYPE_TRANSFER
            op.warehouse = None
            op.price = price
            op.created_by = request.user
            op.save()
            messages.success(request, "Перемещение проведено.")
            return redirect("warehouses_dashboard")
    else:
        form = WarehouseTransferForm()
        form.fields["from_warehouse"].queryset = company.warehouses.all().order_by("name")
        form.fields["to_warehouse"].queryset = company.warehouses.all().order_by("name")
        form.fields["resource"].queryset = company.supply_resources.all().order_by("type", "name")
    return render(request, "core/warehouse_operation_form.html", {
        "form": form,
        "company": company,
        "operation_label": "Перемещение",
        "form_url": reverse("warehouses_transfer"),
    })


# ---------- Склады (Material / Stock / StockMovement): список, карточка, операции ----------


@login_required
@ensure_csrf_cookie
@permission_required("view_warehouse")
def warehouse_inventory_erp(request: HttpRequest) -> HttpResponse:
    """ERP-интерфейс учёта инвентаря (React + API)."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    return render(request, "core/warehouse_inventory_erp.html", {"company": company})


@login_required
def warehouse_list(request: HttpRequest) -> HttpResponse:
    """Список складов: название, локация, кол-во материалов, общая стоимость. Кнопки: Создать, Переместить, Отчёт."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    warehouses = (
        Warehouse.objects.filter(company=company, is_deleted=False)
        .prefetch_related("stocks__material")
        .order_by("name")
    )
    from decimal import Decimal
    stats = []
    for w in warehouses:
        stocks = [s for s in w.stocks.all() if s.quantity > 0]
        total_value = sum((s.quantity * s.price_avg for s in stocks), Decimal("0"))
        stats.append({"warehouse": w, "materials_count": len(stocks), "total_value": total_value})
    return render(request, "core/warehouse_list.html", {
        "company": company,
        "warehouses": warehouses,
        "stats": stats,
    })


@login_required
def warehouse_create(request: HttpRequest) -> HttpResponse:
    """Создать склад."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    if request.method == "POST":
        form = WarehouseCreateForm(request.POST)
        if form.is_valid():
            w = form.save(commit=False)
            w.company = company
            w.save()
            messages.success(request, "Склад создан.")
            return redirect("warehouse_detail", pk=w.pk)
    else:
        form = WarehouseCreateForm()
    return render(request, "core/warehouse_form.html", {"form": form, "company": company})


@login_required
def warehouse_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """Карточка склада: вкладки Остатки, Движение, Инвентаризация."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    warehouse = get_object_or_404(
        Warehouse.objects.select_related("project"),
        pk=pk,
        company=company,
        is_deleted=False,
    )
    tab = request.GET.get("tab", "stocks")
    if tab not in ("stocks", "movements", "inventory"):
        tab = "stocks"
    stocks = (
        Stock.objects.filter(warehouse=warehouse, quantity__gt=0)
        .select_related("material")
        .order_by("material__category", "material__name")
    )
    resource_stocks = (
        StockItem.objects.filter(warehouse=warehouse, quantity__gt=0)
        .select_related("resource")
        .order_by("resource__type", "resource__name")
    )
    from django.db.models import Q

    movement_base = StockMovement.objects.filter(
        Q(warehouse_from=warehouse) | Q(warehouse_to=warehouse)
    )
    has_material_movements = movement_base.exists()
    movements = movement_base.select_related(
        "material", "warehouse_from", "warehouse_to", "project"
    ).order_by("-date", "-created_at")[:500]
    has_positive_material_stock = stocks.exists()
    show_movements_without_balance_hint = has_material_movements and not has_positive_material_stock
    return render(request, "core/warehouse_detail.html", {
        "warehouse": warehouse,
        "company": company,
        "active_tab": tab,
        "stocks": stocks,
        "resource_stocks": resource_stocks,
        "movements": movements,
        "show_movements_without_balance_hint": show_movements_without_balance_hint,
    })


@login_required
def stock_incoming(request: HttpRequest) -> HttpResponse:
    """Добавить поступление на склад."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    if request.method == "POST":
        form = StockIncomingForm(request.POST, company=company)
        if form.is_valid():
            from datetime import date
            cd = form.cleaned_data
            apply_incoming(
                material=cd["material"],
                warehouse=cd["warehouse"],
                quantity=cd["quantity"],
                price=cd["price"],
                date=cd["date"],
                comment=cd.get("comment", ""),
                user=request.user,
            )
            messages.success(request, "Поступление проведено.")
            return redirect("warehouse_detail", pk=cd["warehouse"].pk)
    else:
        form = StockIncomingForm(company=company)
    return render(request, "core/stock_operation_form.html", {
        "form": form,
        "company": company,
        "title": "Добавить поступление",
        "form_url": reverse("stock_incoming"),
    })


@login_required
def stock_writeoff(request: HttpRequest) -> HttpResponse:
    """Добавить списание со склада."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    if request.method == "POST":
        form = StockWriteoffForm(request.POST, company=company)
        if form.is_valid():
            cd = form.cleaned_data
            try:
                apply_writeoff(
                    material=cd["material"],
                    warehouse=cd["warehouse"],
                    quantity=cd["quantity"],
                    date=cd["date"],
                    comment=cd.get("comment", ""),
                    project=cd.get("project"),
                    user=request.user,
                )
            except ValueError as e:
                messages.error(request, str(e))
                return render(request, "core/stock_operation_form.html", {
                    "form": form,
                    "company": company,
                    "title": "Добавить списание",
                    "form_url": reverse("stock_writeoff"),
                })
            messages.success(request, "Списание проведено.")
            return redirect("warehouse_detail", pk=cd["warehouse"].pk)
    else:
        form = StockWriteoffForm(company=company)
    return render(request, "core/stock_operation_form.html", {
        "form": form,
        "company": company,
        "title": "Добавить списание",
        "form_url": reverse("stock_writeoff"),
    })


@login_required
def stock_transfer(request: HttpRequest) -> HttpResponse:
    """Перемещение между складами."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    if request.method == "POST":
        form = StockTransferForm(request.POST, company=company)
        if form.is_valid():
            cd = form.cleaned_data
            try:
                apply_transfer(
                    material=cd["material"],
                    warehouse_from=cd["warehouse_from"],
                    warehouse_to=cd["warehouse_to"],
                    quantity=cd["quantity"],
                    date=cd["date"],
                    comment=cd.get("comment", ""),
                    user=request.user,
                )
            except ValueError as e:
                messages.error(request, str(e))
                return render(request, "core/stock_operation_form.html", {
                    "form": form,
                    "company": company,
                    "title": "Перемещение между складами",
                    "form_url": reverse("stock_transfer"),
                })
            messages.success(request, "Перемещение проведено.")
            return redirect("warehouse_list")
    else:
        form = StockTransferForm(company=company)
    return render(request, "core/stock_operation_form.html", {
        "form": form,
        "company": company,
        "title": "Перемещение между складами",
        "form_url": reverse("stock_transfer"),
    })


@login_required
def material_create(request: HttpRequest) -> HttpResponse:
    """Добавить материал в справочник компании."""
    company = _get_warehouse_company(request)
    if not company:
        return redirect("dashboard")
    if request.method == "POST":
        form = MaterialCreateForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            m.company = company
            m.save()
            messages.success(request, "Материал добавлен.")
            return redirect("warehouse_list")
    else:
        form = MaterialCreateForm()
    return render(request, "core/material_form.html", {"form": form, "company": company})


def _equipment_audit(equipment: Equipment, user, message: str) -> None:
    EquipmentAuditLog.objects.create(
        equipment=equipment,
        user=user,
        message=message[:2000],
    )


@login_required
def company_equipment_list(request: HttpRequest) -> HttpResponse:
    company = _get_warehouse_company(request)
    if not company or not has_permission(request.user, company, "view_warehouse"):
        messages.error(request, "Нет доступа к справочнику техники.")
        return redirect("dashboard")
    qs = Equipment.objects.filter(company=company).select_related(
        "project", "fuel_type", "base_warehouse", "responsible"
    )
    st = request.GET.get("status", "").strip()
    if st and st in dict(Equipment.STATUS_CHOICES):
        qs = qs.filter(status=st)
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(license_plate__icontains=q)
            | Q(inventory_number__icontains=q)
            | Q(make_model__icontains=q)
        )
    return render(
        request,
        "core/equipment_list.html",
        {
            "company": company,
            "equipment_list": qs.order_by("name")[:500],
            "status_filter": st,
            "search_q": q,
            "status_choices": Equipment.STATUS_CHOICES,
        },
    )


@login_required
def company_equipment_create(request: HttpRequest) -> HttpResponse:
    company = _get_warehouse_company(request)
    if not company or not has_permission(request.user, company, "edit_warehouse"):
        messages.error(request, "Нет права редактировать технику.")
        return redirect("dashboard")
    if request.method == "POST":
        form = EquipmentForm(request.POST, request.FILES, company=company)
        if form.is_valid():
            from .fuel_services import ensure_default_fuel_types

            ensure_default_fuel_types(company)
            obj = form.save(commit=False)
            obj.company = company
            obj.save()
            _equipment_audit(obj, request.user, "Создана запись в справочнике техники")
            messages.success(request, "Техника добавлена.")
            return redirect("company_equipment_detail", pk=obj.pk)
    else:
        form = EquipmentForm(company=company)
    return render(
        request,
        "core/equipment_form.html",
        {"form": form, "company": company, "is_edit": False},
    )


@login_required
def company_equipment_edit(request: HttpRequest, pk: int) -> HttpResponse:
    company = _get_warehouse_company(request)
    if not company or not has_permission(request.user, company, "edit_warehouse"):
        messages.error(request, "Нет права редактировать технику.")
        return redirect("dashboard")
    obj = get_object_or_404(Equipment, pk=pk, company=company)
    if request.method == "POST":
        form = EquipmentForm(
            request.POST, request.FILES, instance=obj, company=company
        )
        if form.is_valid():
            form.save()
            _equipment_audit(obj, request.user, "Изменена карточка техники")
            messages.success(request, "Сохранено.")
            return redirect("company_equipment_detail", pk=obj.pk)
    else:
        form = EquipmentForm(instance=obj, company=company)
    return render(
        request,
        "core/equipment_form.html",
        {"form": form, "company": company, "is_edit": True, "equipment": obj},
    )


@login_required
def company_equipment_detail(request: HttpRequest, pk: int) -> HttpResponse:
    company = _get_warehouse_company(request)
    if not company or not has_permission(request.user, company, "view_warehouse"):
        messages.error(request, "Нет доступа.")
        return redirect("dashboard")
    obj = get_object_or_404(Equipment, pk=pk, company=company)
    fuel_logs = list(
        obj.fuel_logs.select_related("transaction", "transaction__fuel_type")
        .order_by("-created_at")[:100]
    )
    return render(
        request,
        "core/equipment_detail.html",
        {
            "company": company,
            "equipment": obj,
            "audit_logs": obj.audit_logs.select_related("user").all()[:100],
            "documents": obj.documents.all()[:100],
            "fuel_logs": fuel_logs,
        },
    )


@login_required
@require_POST
def company_equipment_document_add(request: HttpRequest, pk: int) -> HttpResponse:
    company = _get_warehouse_company(request)
    if not company or not has_permission(request.user, company, "edit_warehouse"):
        messages.error(request, "Нет права.")
        return redirect("dashboard")
    obj = get_object_or_404(Equipment, pk=pk, company=company)
    title = (request.POST.get("title") or "").strip()
    f = request.FILES.get("file")
    if f:
        EquipmentDocument.objects.create(equipment=obj, title=title, file=f)
        _equipment_audit(obj, request.user, f"Загружен документ: {title or f.name}")
        messages.success(request, "Файл добавлен.")
    else:
        messages.error(request, "Выберите файл.")
    return redirect("company_equipment_detail", pk=obj.pk)


@login_required
def company_equipment_qr_card(request: HttpRequest, token) -> HttpResponse:
    """Карточка по QR (токен), для быстрого просмотра на объекте."""
    obj = get_object_or_404(Equipment, qr_token=token)
    company = _get_warehouse_company(request)
    if not company or obj.company_id != company.id:
        messages.error(request, "Техника другой компании.")
        return redirect("dashboard")
    if not has_permission(request.user, company, "view_warehouse"):
        return redirect("dashboard")
    return render(request, "core/equipment_qr_card.html", {"equipment": obj, "company": company})


# ---------- Модуль «Отчёты» ----------


def _get_reports_company(request: HttpRequest) -> Company | None:
    return get_current_company(request.user)


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        y, m, d = value.split("-")
        return date(int(y), int(m), int(d))
    except Exception:
        return None


def _get_period_range(request: HttpRequest) -> tuple[date, date, dict]:
    """
    Период по GET:
    - date_from, date_to (приоритетно)
    - period=month|quarter|year + year + month/quarter
    """
    today = date.today()
    period = request.GET.get("period", "year")
    year = int(request.GET.get("year") or today.year)
    month = int(request.GET.get("month") or today.month)
    quarter = int(request.GET.get("quarter") or ((today.month - 1) // 3 + 1))

    df = _parse_date(request.GET.get("date_from"))
    dt = _parse_date(request.GET.get("date_to"))
    if df and dt:
        meta = {"period": "custom", "year": year, "month": month, "quarter": quarter, "date_from": df, "date_to": dt}
        return df, dt, meta

    if period == "month":
        df = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        dt = date(year, month, last_day)
    elif period == "quarter":
        start_month = (quarter - 1) * 3 + 1
        df = date(year, start_month, 1)
        end_month = start_month + 2
        last_day = calendar.monthrange(year, end_month)[1]
        dt = date(year, end_month, last_day)
    else:  # year
        period = "year"
        df = date(year, 1, 1)
        dt = date(year, 12, 31)

    meta = {"period": period, "year": year, "month": month, "quarter": quarter, "date_from": df, "date_to": dt}
    return df, dt, meta


def _qs_base(period_meta: dict) -> str:
    """Query string без basis/export, чтобы строить ссылки переключателей."""
    df: date = period_meta["date_from"]
    dt: date = period_meta["date_to"]
    return (
        f"period={period_meta.get('period','year')}"
        f"&year={period_meta.get('year','')}"
        f"&quarter={period_meta.get('quarter','')}"
        f"&month={period_meta.get('month','')}"
        f"&date_from={df.isoformat()}&date_to={dt.isoformat()}"
    )


@login_required
def reports_index(request: HttpRequest) -> HttpResponse:
    """Главная /reports/ с вкладками."""
    company = _get_reports_company(request)
    if not company:
        return redirect("dashboard")

    tab = request.GET.get("tab", "management")
    if tab not in ("management", "project", "settlements"):
        tab = "management"

    projects = company.projects.all().order_by("name")
    return render(request, "core/reports_index.html", {
        "company": company,
        "active_tab": tab,
        "projects": projects,
    })


def _pnl_xlsx_response(pnl: dict, filename: str = "pnl.xlsx") -> HttpResponse:
    from io import BytesIO
    from decimal import Decimal

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"

    months: list[date] = pnl["months"]
    lines = pnl["lines"]

    header = ["Статья"] + [m.strftime("%b %Y") for m in months] + ["Итого"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    green = PatternFill("solid", fgColor="E8F5E9")
    red = PatternFill("solid", fgColor="FFEBEE")

    for ln in lines:
        row = [ln.label] + [float(ln.values.get(m, Decimal("0"))) for m in months] + [float(ln.total)]
        ws.append(row)
        # подсветка строки по итоговому значению (прибыль/убыток)
        fill = green if ln.total >= 0 else red
        for cell in ws[ws.max_row]:
            cell.fill = fill

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
def reports_pnl(request: HttpRequest) -> HttpResponse:
    """Отчёт P&L."""
    company = _get_reports_company(request)
    if not company:
        return redirect("dashboard")

    date_from, date_to, period_meta = _get_period_range(request)
    basis = request.GET.get("basis", "cash")
    if basis not in ("cash", "accrual"):
        basis = "cash"

    pnl = report_services.build_pnl(company=company, date_from=date_from, date_to=date_to, basis=basis)
    months = pnl["months"]
    rows = [
        {
            "key": ln.key,
            "label": ln.label,
            "values": [ln.values.get(m) for m in months],
            "total": ln.total,
        }
        for ln in pnl["lines"]
    ]

    if request.GET.get("export") == "xlsx":
        return _pnl_xlsx_response(pnl, filename=f"pnl_{date_from}_{date_to}.xlsx")

    return render(request, "core/reports_pnl.html", {
        "company": company,
        "months": months,
        "rows": rows,
        "date_from": date_from,
        "date_to": date_to,
        "period_meta": period_meta,
        "basis": basis,
        "qs_base": _qs_base(period_meta),
    })


@login_required
def reports_cashflow(request: HttpRequest) -> HttpResponse:
    """Отчёт Cash Flow."""
    company = _get_reports_company(request)
    if not company:
        return redirect("dashboard")

    date_from, date_to, period_meta = _get_period_range(request)
    basis = request.GET.get("basis", "cash")
    if basis not in ("cash", "accrual"):
        basis = "cash"

    cashflow = report_services.build_cashflow(company=company, date_from=date_from, date_to=date_to, basis=basis)
    months = cashflow["months"]
    rows = [
        {
            "key": ln.key,
            "label": ln.label,
            "values": [ln.values.get(m) for m in months],
            "total": ln.total,
        }
        for ln in cashflow["sections"]
    ]
    return render(request, "core/reports_cashflow.html", {
        "company": company,
        "months": months,
        "rows": rows,
        "date_from": date_from,
        "date_to": date_to,
        "period_meta": period_meta,
        "basis": basis,
        "qs_base": _qs_base(period_meta),
    })


@login_required
def reports_project(request: HttpRequest, id: int) -> HttpResponse:
    """Проектный отчёт (план/факт/маржинальность)."""
    company = _get_reports_company(request)
    if not company:
        return redirect("dashboard")

    project = get_object_or_404(Project, pk=id, company=company)
    date_from, date_to, period_meta = _get_period_range(request)
    data = report_services.build_project_report(company=company, project=project, date_from=date_from, date_to=date_to)

    return render(request, "core/reports_project.html", {
        "company": company,
        "project": project,
        "data": data,
        "date_from": date_from,
        "date_to": date_to,
        "period_meta": period_meta,
        "qs_base": _qs_base(period_meta),
    })

