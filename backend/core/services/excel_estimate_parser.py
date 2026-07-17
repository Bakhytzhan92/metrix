"""
Парсер импорта сметы из Excel.

Поддерживаемые форматы:
1. Таблица работ: № позиции | Наименование работ | Ед. изм. | Кол-во (по заголовкам).
2. BOQ ABC: колонки 1, 3, 4, 5 (разделы и позиции по правилам локальной сметы).
3. Экспорт Metrix: Раздел | Наименование | Тип | Ед | Кол-во | Цена.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, BinaryIO, Iterator

from openpyxl import load_workbook

# --- текст и фильтры ---

_INVISIBLE = re.compile(r"[\u200b-\u200f\u202a-\u202e\u2060\ufeff]")
_MULTI_SPACE = re.compile(r"\s+")
_CYRILLIC = re.compile(r"[\u0400-\u04FF\u0500-\u052F]")
_LATIN = re.compile(r"[A-Za-z]")
_DIGITS_ONLY = re.compile(r"^[\d\s.,]+$")
_SPECIALS_ONLY = re.compile(r"^[\s\W_]+$", re.UNICODE)

_ENGLISH_SKIP = frozenset(
    s.upper()
    for s in (
        "SECTION",
        "EARTH WORKS",
        "DESCRIPTION",
        "QUANTITY",
        "UNIT",
        "TOTAL",
        "SUMMARY",
        "SUBTOTAL",
        "AMOUNT",
        "BILL OF QUANTITIES",
        "NO",
        "ITEM",
        "CODE",
        "RATE",
        "PRICE",
        "COST",
        "REMARKS",
        "NOTES",
    )
)

_TOTALS_SKIP = frozenset(
    s.upper()
    for s in (
        "ИТОГО",
        "ВСЕГО",
        "ВСЕГО ПО РАЗДЕЛУ",
        "ВСЕГО ПО СМЕТЕ",
        "НАКЛАДНЫЕ РАСХОДЫ",
        "СМЕТНАЯ ПРИБЫЛЬ",
        "НДС",
        "ИТОГО С НДС",
    )
)

_PLACEHOLDER_UNITS = frozenset({"", "/", "*", "—", "-", "–", "/ "})

DEFAULT_IMPORT_SECTION = "Смета"

_PLACEHOLDER_SECTION_PREFIXES = ("импорт из",)


def _is_placeholder_section_title(name: str) -> bool:
    low = (name or "").strip().lower()
    if not low:
        return True
    if low in ("импорт из excel", "импорт из"):
        return True
    return any(low.startswith(p) for p in _PLACEHOLDER_SECTION_PREFIXES)


def _collect_section_title_xlsx(ws: Any, header_row: int, *, max_col: int = 12) -> str:
    """Самая длинная строка над шапкой таблицы — название раздела."""
    candidates: list[str] = []
    for row_idx in range(1, header_row):
        parts: list[str] = []
        for col in range(1, max_col + 1):
            val = cell_name_text(ws.cell(row_idx, col).value)
            if val and not _header_cell_kind(val) and not is_junk_name(val):
                parts.append(val)
        if parts:
            line = " ".join(parts).strip()
            if line and not _is_placeholder_section_title(line):
                candidates.append(line)
    if candidates:
        return max(candidates, key=len)[:255]
    return ""

_HEADER_POS = re.compile(r"№|п/п|позици", re.IGNORECASE)
_HEADER_NAME = re.compile(r"наименован", re.IGNORECASE)
_HEADER_UNIT = re.compile(r"ед\.?\s*изм", re.IGNORECASE)
_HEADER_QTY = re.compile(r"кол-?во|количество", re.IGNORECASE)

# ARGB из Excel → стиль заголовка раздела в UI
_FILL_TO_ACCENT: dict[str, str] = {}
for _rgb, _accent in (
    ("FFCC0000", "red"),
    ("FFC00000", "red"),
    ("FFFF0000", "red"),
    ("FF800000", "bordeaux"),
    ("FF7030A0", "bordeaux"),
    ("FF993366", "bordeaux"),
    ("FFFFC000", "gold"),
    ("FFFF9900", "gold"),
    ("FFD99694", "gold"),
    ("FFE26B0A", "gold"),
):
    _FILL_TO_ACCENT[_rgb.upper()] = _accent


@dataclass
class ParsedExcelRow:
    kind: str  # "section" | "position"
    name: str
    list_no: str = ""  # № из 1-го столбца Excel
    unit: str = ""
    quantity: Decimal | None = None
    header_accent: str = ""
    source_row: int = 0


@dataclass
class ExcelParseResult:
    rows: list[ParsedExcelRow] = field(default_factory=list)
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


def clean_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        s = str(int(value))
    else:
        s = str(value)
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = s.replace("\t", " ")
    s = _INVISIBLE.sub("", s)
    s = _MULTI_SPACE.sub(" ", s).strip()
    return s


# Разделители двуязычных ячеек BOQ: «русский / English», «м3 грунта / m3 of soil»
_BILINGUAL_SPLIT = re.compile(r"\s*/\s*|\s*\\\s*")
# Хвост «/ MONOLITHIC …» или «/ Repair of …» без кириллицы в той же части
_TRAILING_LATIN_TAIL = re.compile(
    r"\s*/\s*(?![\u0400-\u04FF\u0500-\u052F])[\s\S]*$",
    re.IGNORECASE,
)
# Английские слова в хвосте (Canal, DISMANTLING, Dismantling works …)
_LATIN_WORD = re.compile(r"[A-Za-z]{3,}")
# Типичные английские дубли в BOQ/сметах (в т.ч. «Сanal» с кириллической С)
_TRAILING_EN_WORDS = frozenset(
    w.lower()
    for w in (
        "Canal",
        "Channel",
        "Dismantling",
        "EARTHWORKS",
        "WORKS",
        "Repair",
        "OTHER",
        "LINING",
        "TRAYS",
        "CONCRETE",
        "STRUCTURE",
    )
)
# Визуально похожие кириллица ↔ латиница (С↔C, а↔a …)
_HOMOGLYPH_MAP = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "С": "C",
        "Е": "E",
        "Н": "H",
        "К": "K",
        "М": "M",
        "О": "O",
        "Р": "P",
        "Т": "T",
        "Х": "X",
        "а": "a",
        "в": "b",
        "с": "c",
        "е": "e",
        "н": "h",
        "к": "k",
        "м": "m",
        "о": "o",
        "р": "p",
        "т": "t",
        "х": "x",
    }
)


def _latin_fold(word: str) -> str:
    """Омографы кириллицы в латиницу для сравнения (Сanal → Canal)."""
    return word.translate(_HOMOGLYPH_MAP)


def _should_strip_trailing_word(word: str) -> bool:
    w = word.strip(".,;:!?")
    if not w:
        return False
    if re.fullmatch(r"[\(\)\d.\-+/№]+", w):
        return False
    folded = _latin_fold(w).lower()
    if folded in _TRAILING_EN_WORDS:
        return True
    if has_cyrillic(w):
        return False
    if re.fullmatch(r"[A-Za-z][A-Za-z.\-]*", w) and len(w) >= 3:
        return True
    return False


def _strip_trailing_english_words(text: str) -> str:
    """Убирает последние слова-англицизмы: Canal, Сanal, Works …"""
    words = text.split()
    while words and _should_strip_trailing_word(words[-1]):
        words.pop()
    return " ".join(words).strip()


def _is_english_tail(fragment: str) -> bool:
    """Фрагмент без кириллицы, похожий на английский дубляж наименования."""
    if not fragment or has_cyrillic(fragment):
        return False
    return bool(_LATIN_WORD.search(fragment))


def _strip_latin_suffix_after_cyrillic(text: str) -> str:
    """
    Убирает хвост без кириллицы после русского текста:
    «… (ГР-29) Canal 4-K-1-1. DISMANTLING WORKS» → «… (ГР-29)».
    """
    words = text.split()
    if len(words) < 2:
        return text
    for i in range(1, len(words)):
        head = " ".join(words[:i]).strip()
        tail = " ".join(words[i:]).strip()
        if has_cyrillic(head) and _is_english_tail(tail):
            return head
    return text


def keep_cyrillic_text(text: str) -> str:
    """
    Оставляет только русский/казахский фрагмент ячейки.
    Английский дубляж после «/», через «/» или отдельным хвостом отбрасывается.
    """
    s = clean_cell_text(text)
    if not s:
        return ""
    if not has_cyrillic(s):
        return ""

    parts = [p.strip() for p in _BILINGUAL_SPLIT.split(s) if p.strip()]
    if len(parts) > 1:
        cyr_parts = [p for p in parts if has_cyrillic(p)]
        if not cyr_parts:
            return ""
        if len(cyr_parts) == 1:
            s = cyr_parts[0]
        else:
            keys = [_unit_token_key(p) for p in cyr_parts]
            if len(set(keys)) == 1 and keys[0] and all(len(p) <= 8 for p in cyr_parts):
                s = next(p for p in cyr_parts if has_cyrillic(p))
            else:
                s = " ".join(cyr_parts)

    s = _TRAILING_LATIN_TAIL.sub("", s).strip()
    for _ in range(6):
        prev = s
        s = _strip_latin_suffix_after_cyrillic(s)
        s = re.sub(
            r"\s*\(\s*[A-Za-z][^)]*\)\s*$",
            "",
            s,
        ).strip()
        if s == prev:
            break
    s = _strip_trailing_english_words(s)
    # Хвост из латинских слов без кириллицы (Dismantling works …)
    s = re.sub(
        r"\s+(?:(?:[A-Za-z]{2,}[A-Za-z0-9.\-/]*)\s*)+$",
        "",
        s,
    ).strip()
    s = _strip_trailing_english_words(s)
    s = re.sub(
        r"\s+(?:[СCсc][AaАа][NnНн][AaАа][LlЛл]\.?)\s*$",
        "",
        s,
    ).strip()
    s = _MULTI_SPACE.sub(" ", s).strip()
    return s if has_cyrillic(s) else ""


def normalize_estimate_name(text: str) -> str:
    """Наименование раздела/позиции без английского дубляжа (импорт и отображение)."""
    cleaned = keep_cyrillic_text(text)
    return cleaned if cleaned else (text or "").strip()


# Соответствие латиница ↔ кириллица для коротких единиц (t/т, m3/м3 …)
_UNIT_TOKEN_MAP = {
    "t": "т",
    "m3": "м3",
    "m2": "м2",
    "kg": "кг",
    "km": "км",
    "m": "м",
    "l": "л",
    "pcs": "шт",
    "pc": "шт",
    "шт": "шт",
    "компл": "компл",
    "комплект": "комплект",
}


def _unit_token_key(token: str) -> str:
    """Ключ для сравнения дублей «т/т», «m3/м3»."""
    t = _latin_fold(token.strip().lower().rstrip(".,;:"))
    t = t.replace("³", "3").replace(" ", "")
    if t in _UNIT_TOKEN_MAP:
        return _UNIT_TOKEN_MAP[t]
    if t == "т":
        return "т"
    return t


def _collapse_duplicate_unit_tokens(text: str) -> str:
    """«т т», «т/т», «m3 м3» → одна единица."""
    s = _MULTI_SPACE.sub(" ", (text or "").strip())
    if not s:
        return s
    if re.fullmatch(r"(?i)(т|t)\s*/\s*(т|t)\.?", s):
        return "т"
    words = s.split()
    if len(words) == 2:
        k0, k1 = _unit_token_key(words[0]), _unit_token_key(words[1])
        if k0 and k0 == k1 and len(k0) <= 6:
            return words[0] if has_cyrillic(words[0]) else words[1]
    return s


def normalize_excel_unit(text: Any) -> str:
    """Ед. изм. без английского дубляжа; т/т → т."""
    if text is None or text is False:
        return ""
    if isinstance(text, (int, float)):
        return clean_cell_text(text)
    cleaned = keep_cyrillic_text(text)
    if not cleaned:
        return clean_cell_text(text)
    return _collapse_duplicate_unit_tokens(cleaned)


def _normalize_rgb(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, str):
        s = raw.strip().upper()
        if len(s) == 8 and s.startswith("FF"):
            return s
        if len(s) == 6:
            return "FF" + s
        return None
    try:
        s = str(raw).strip().upper()
        if hasattr(raw, "rgb"):
            s = str(raw.rgb).strip().upper()
    except Exception:
        return None
    if s in ("00000000", "FFFFFFFF", "FF000000", "NONE", "AUTO"):
        return None
    if len(s) == 8:
        return s
    if len(s) == 6:
        return "FF" + s
    return None


def cell_fill_accent(cell: Any) -> str:
    """Цвет заливки ячейки (кол. 3) → accent для заголовка раздела."""
    if cell is None:
        return ""
    fill = getattr(cell, "fill", None)
    if not fill or getattr(fill, "fill_type", None) in (None, "none"):
        return ""
    candidates: list[str] = []
    for attr in ("fgColor", "start_color", "bgColor"):
        part = getattr(fill, attr, None)
        if part is None:
            continue
        rgb = _normalize_rgb(getattr(part, "rgb", None) or part)
        if rgb:
            candidates.append(rgb)
    for rgb in candidates:
        accent = _FILL_TO_ACCENT.get(rgb.upper())
        if accent:
            return accent
    for rgb in candidates:
        r = int(rgb[2:4], 16)
        g = int(rgb[4:6], 16)
        b = int(rgb[6:8], 16)
        if r >= 160 and g < 100 and b < 100:
            return "red" if r >= 200 else "bordeaux"
        if r >= 200 and g >= 140 and b < 80:
            return "gold"
    return ""


def _is_placeholder_unit(unit: str) -> bool:
    u = unit.strip()
    if u in _PLACEHOLDER_UNITS:
        return True
    if u.replace(" ", "") in ("/", "*"):
        return True
    return False


def has_cyrillic(text: str) -> bool:
    return bool(_CYRILLIC.search(text))


def is_english_only_line(text: str) -> bool:
    t = text.strip()
    if not t:
        return True
    if not _LATIN.search(t):
        return False
    if has_cyrillic(t):
        return False
    upper = re.sub(r"[^A-Za-z\s]", "", t).strip().upper()
    if upper in _ENGLISH_SKIP:
        return True
    words = [w for w in upper.split() if w]
    if words and all(w in _ENGLISH_SKIP for w in words):
        return True
    return not has_cyrillic(t) and bool(_LATIN.search(t))


def is_totals_line(name: str) -> bool:
    key = re.sub(r"\s+", " ", name.strip().upper())
    if key in _TOTALS_SKIP:
        return True
    for prefix in _TOTALS_SKIP:
        if key.startswith(prefix + " ") or key.startswith(prefix + ":"):
            return True
    return False


def is_junk_name(name: str) -> bool:
    if not name:
        return True
    if _DIGITS_ONLY.match(name):
        return True
    if _SPECIALS_ONLY.match(name):
        return True
    if is_english_only_line(name):
        return True
    if is_totals_line(name):
        return True
    return False


def parse_excel_list_no(value: Any) -> str:
    """№ п/п из 1-го столбца (целое или короткая метка)."""
    if value is None or value is False:
        return ""
    if isinstance(value, (int, float)):
        try:
            if value == int(value):
                return str(int(value))
        except (ValueError, OverflowError):
            return ""
        return clean_cell_text(value)[:16]
    s = clean_cell_text(value)
    if not s:
        return ""
    upper = s.upper().replace("№", "").strip()
    if upper in ("NO", "N", "П/П", "ПП"):
        return ""
    if re.fullmatch(r"\d+", s):
        return s
    return s[:16]


def parse_quantity(value: Any) -> Decimal | None:
    if value is None or value is False:
        return None
    if isinstance(value, (int, float)):
        try:
            d = Decimal(str(value))
            return d if d > 0 else None
        except (InvalidOperation, ValueError):
            return None
    s = clean_cell_text(value)
    if not s:
        return None
    s = re.sub(r"^[−–—-]+\s*", "", s)
    s = s.replace(" ", "").replace(",", ".")
    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    return d if d > 0 else None


def cell_name_text(value: Any) -> str:
    """Наименование: кириллица без дубляжа, иначе очищенный исходный текст."""
    cyr = keep_cyrillic_text(value)
    if cyr:
        return cyr
    return clean_cell_text(value)


def _header_cell_kind(text: str) -> str | None:
    raw = clean_cell_text(text)
    if not raw or len(raw) > 48:
        return None
    low = raw.lower()
    if _HEADER_NAME.search(low):
        return "name"
    if _HEADER_UNIT.search(low):
        return "unit"
    if _HEADER_QTY.search(low):
        return "qty"
    if _HEADER_POS.search(low) and "наименован" not in low:
        return "pos"
    return None


@dataclass
class _WorkTableColumns:
    header_row: int = 0
    pos_col: int = 1
    name_col: int = 2
    unit_col: int = 3
    qty_col: int = 4
    section_title: str = DEFAULT_IMPORT_SECTION


def classify_row(
    name: str,
    unit: str,
    qty: Decimal | None,
    *,
    fill_accent: str = "",
) -> str | None:
    """section | position | None (пропуск)."""
    if is_junk_name(name):
        return None
    if not has_cyrillic(name):
        return None
    unit_ok = _is_placeholder_unit(unit)
    if unit_ok and qty is None:
        return "section"
    if not unit_ok and qty is not None:
        return "position"
    if fill_accent and unit_ok and qty is None:
        return "section"
    return None


def _seekable_copy(file: BinaryIO) -> BinaryIO:
    if hasattr(file, "seek"):
        try:
            file.seek(0)
            return file
        except (OSError, AttributeError, TypeError):
            pass
    data = file.read()
    return BytesIO(data)


def _detect_work_table_xlsx(ws: Any, *, max_scan: int = 25, max_col: int = 12) -> _WorkTableColumns | None:
    for row_idx in range(1, min(max_scan, ws.max_row or 0) + 1):
        mapping: dict[str, int] = {}
        for col in range(1, max_col + 1):
            kind = _header_cell_kind(ws.cell(row_idx, col).value)
            if kind:
                mapping[kind] = col
        if "name" not in mapping:
            continue
        if "qty" not in mapping and "unit" not in mapping:
            continue
        cols = _WorkTableColumns(header_row=row_idx)
        cols.name_col = mapping["name"]
        cols.pos_col = mapping.get("pos", max(1, cols.name_col - 1))
        cols.unit_col = mapping.get("unit", cols.name_col + 1)
        cols.qty_col = mapping.get("qty", cols.unit_col + 1)
        title = _collect_section_title_xlsx(ws, row_idx)
        if title:
            cols.section_title = title
        return cols
    return None


def _parse_simple_work_table_xlsx(file: BinaryIO) -> ExcelParseResult | None:
    wb = load_workbook(_seekable_copy(file), data_only=True, read_only=False)
    try:
        ws = wb.active
        if not ws:
            return None
        cols = _detect_work_table_xlsx(ws)
        if not cols:
            return None
        result = ExcelParseResult()
        result.rows.append(
            ParsedExcelRow(
                kind="section",
                name=cols.section_title,
                source_row=max(1, cols.header_row - 1),
            )
        )
        for row_idx in range(cols.header_row + 1, (ws.max_row or 0) + 1):
            name = cell_name_text(ws.cell(row_idx, cols.name_col).value)
            if not name or is_junk_name(name) or _header_cell_kind(name):
                continue
            if is_totals_line(name):
                result.skipped += 1
                continue
            unit = normalize_excel_unit(ws.cell(row_idx, cols.unit_col).value) or "шт"
            qty = parse_quantity(ws.cell(row_idx, cols.qty_col).value)
            if qty is None:
                result.skipped += 1
                continue
            pos_no = parse_excel_list_no(ws.cell(row_idx, cols.pos_col).value)
            result.rows.append(
                ParsedExcelRow(
                    kind="position",
                    name=name,
                    list_no=pos_no,
                    unit=unit,
                    quantity=qty,
                    source_row=row_idx,
                )
            )
        return result if len(result.rows) > 1 else None
    finally:
        wb.close()


def _parse_simple_work_table_xls(file: BinaryIO) -> ExcelParseResult | None:
    import xlrd

    book = xlrd.open_workbook(file_contents=_seekable_copy(file).read())
    sheet = book.sheet_by_index(0)
    cols = _detect_work_table_xls(sheet)
    if not cols:
        return None
    result = ExcelParseResult()
    result.rows.append(
        ParsedExcelRow(
            kind="section",
            name=cols.section_title,
            source_row=max(1, cols.header_row - 1),
        )
    )
    for row_idx in range(cols.header_row, sheet.nrows):
        name = cell_name_text(
            sheet.cell_value(row_idx, cols.name_col - 1)
            if cols.name_col - 1 < sheet.ncols
            else ""
        )
        if not name or is_junk_name(name) or _header_cell_kind(name):
            continue
        if is_totals_line(name):
            result.skipped += 1
            continue
        unit = (
            normalize_excel_unit(
                sheet.cell_value(row_idx, cols.unit_col - 1)
                if cols.unit_col - 1 < sheet.ncols
                else ""
            )
            or "шт"
        )
        qty = parse_quantity(
            sheet.cell_value(row_idx, cols.qty_col - 1)
            if cols.qty_col - 1 < sheet.ncols
            else ""
        )
        if qty is None:
            result.skipped += 1
            continue
        pos_no = parse_excel_list_no(
            sheet.cell_value(row_idx, cols.pos_col - 1)
            if cols.pos_col - 1 < sheet.ncols
            else ""
        )
        result.rows.append(
            ParsedExcelRow(
                kind="position",
                name=name,
                list_no=pos_no,
                unit=unit,
                quantity=qty,
                source_row=row_idx + 1,
            )
        )
    return result if len(result.rows) > 1 else None


def _collect_section_title_xls(sheet: Any, header_row: int, *, max_col: int = 12) -> str:
    candidates: list[str] = []
    for row_idx in range(header_row - 1):
        parts: list[str] = []
        for col in range(min(max_col, sheet.ncols)):
            val = cell_name_text(sheet.cell_value(row_idx, col))
            if val and not _header_cell_kind(val) and not is_junk_name(val):
                parts.append(val)
        if parts:
            line = " ".join(parts).strip()
            if line and not _is_placeholder_section_title(line):
                candidates.append(line)
    if candidates:
        return max(candidates, key=len)[:255]
    return ""


def _detect_work_table_xls(sheet: Any, *, max_scan: int = 25, max_col: int = 12) -> _WorkTableColumns | None:
    for row_idx in range(min(max_scan, sheet.nrows)):
        mapping: dict[str, int] = {}
        for col in range(min(max_col, sheet.ncols)):
            kind = _header_cell_kind(sheet.cell_value(row_idx, col))
            if kind:
                mapping[kind] = col + 1
        if "name" not in mapping:
            continue
        if "qty" not in mapping and "unit" not in mapping:
            continue
        cols = _WorkTableColumns(header_row=row_idx + 1)
        cols.name_col = mapping["name"]
        cols.pos_col = mapping.get("pos", max(1, cols.name_col - 1))
        cols.unit_col = mapping.get("unit", cols.name_col + 1)
        cols.qty_col = mapping.get("qty", cols.unit_col + 1)
        title = _collect_section_title_xls(sheet, row_idx + 1)
        if title:
            cols.section_title = title
        return cols
    return None


def _parse_gectaro_export_format(file: BinaryIO) -> ExcelParseResult:
    """Старый формат экспорта Metrix: Раздел | Наименование | Тип | Ед | Кол-во | Цена."""
    result = ExcelParseResult()
    wb = load_workbook(_seekable_copy(file), data_only=True, read_only=True)
    try:
        ws = wb.active
        if not ws:
            result.errors.append("Нет листа в файле")
            return result
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return result

    header = [clean_cell_text(c).lower() for c in rows[0]]
    if not header or "раздел" not in header[0]:
        return result
    if len(header) < 2 or "наименование" not in header[1]:
        return result

    current_section = ""
    last_section_added = ""
    for row_idx, row in enumerate(rows[1:], start=2):
        parts = [clean_cell_text(c) for c in (list(row)[:6] or [])]
        while len(parts) < 6:
            parts.append("")
        section_name = cell_name_text(parts[0])
        name = cell_name_text(parts[1])
        unit = normalize_excel_unit(parts[3]) or "шт"
        qty = parse_quantity(parts[4])
        if section_name:
            current_section = section_name
        if not name:
            result.skipped += 1
            continue
        if not current_section:
            result.skipped += 1
            continue
        if qty is None:
            result.skipped += 1
            continue
        if current_section != last_section_added:
            result.rows.append(
                ParsedExcelRow(
                    kind="section",
                    name=current_section,
                    source_row=row_idx,
                )
            )
            last_section_added = current_section
        result.rows.append(
            ParsedExcelRow(
                kind="position",
                name=name,
                unit=unit,
                quantity=qty,
                source_row=row_idx,
            )
        )
    return result


def _consume_boq_rows(
    row_iter: Iterator[tuple[int, str, str, str, str, Decimal | None, str]],
) -> ExcelParseResult:
    result = ExcelParseResult()
    header_like = 0
    for row_idx, list_no, name, unit, _c5_raw, qty, accent in row_iter:
        if not name and not unit and qty is None:
            continue
        name = name or cell_name_text(_c5_raw)
        if not name:
            result.skipped += 1
            continue
        if row_idx <= 6 and is_english_only_line(name):
            header_like += 1
            result.skipped += 1
            continue

        kind = classify_row(name, unit, qty, fill_accent=accent)
        if kind is None:
            result.skipped += 1
            continue

        if kind == "section":
            result.rows.append(
                ParsedExcelRow(
                    kind="section",
                    name=name,
                    list_no=list_no,
                    header_accent=accent or "",
                    source_row=row_idx,
                )
            )
        else:
            result.rows.append(
                ParsedExcelRow(
                    kind="position",
                    name=name,
                    list_no=list_no,
                    unit=unit or "шт",
                    quantity=qty,
                    source_row=row_idx,
                )
            )

    if not result.rows and header_like:
        result.errors.append("Не найдено строк для импорта — проверьте формат файла.")
    return result


def _iter_xlsx_rows(
    file: BinaryIO,
) -> Iterator[tuple[int, str, str, str, str, Decimal | None, str]]:
    wb = load_workbook(_seekable_copy(file), data_only=True, read_only=False)
    try:
        ws = wb.active
        if not ws:
            return
        for row_idx in range(1, (ws.max_row or 0) + 1):
            c1 = parse_excel_list_no(ws.cell(row_idx, 1).value)
            c2 = cell_name_text(ws.cell(row_idx, 2).value)
            c3 = cell_name_text(ws.cell(row_idx, 3).value)
            name = c3 or c2
            c4 = normalize_excel_unit(ws.cell(row_idx, 4).value)
            c5_raw = ws.cell(row_idx, 5).value
            c6_raw = ws.cell(row_idx, 6).value
            qty = parse_quantity(c5_raw)
            if qty is None:
                qty = parse_quantity(c6_raw)
            unit = c4 or normalize_excel_unit(ws.cell(row_idx, 3).value)
            accent = cell_fill_accent(ws.cell(row_idx, 3)) or cell_fill_accent(
                ws.cell(row_idx, 2)
            )
            yield row_idx, c1, name, unit, c5_raw if c5_raw is not None else "", qty, accent
    finally:
        wb.close()


def _iter_xls_rows(
    file: BinaryIO,
) -> Iterator[tuple[int, str, str, str, str, Decimal | None, str]]:
    import xlrd

    book = xlrd.open_workbook(file_contents=_seekable_copy(file).read())
    sheet = book.sheet_by_index(0)
    for row_idx in range(sheet.nrows):
        r = row_idx + 1
        c1 = parse_excel_list_no(
            sheet.cell_value(row_idx, 0) if sheet.ncols > 0 else ""
        )
        c3 = cell_name_text(sheet.cell_value(row_idx, 2) if sheet.ncols > 2 else "")
        c2 = cell_name_text(sheet.cell_value(row_idx, 1) if sheet.ncols > 1 else "")
        name = c3 or c2
        c4 = normalize_excel_unit(
            sheet.cell_value(row_idx, 3) if sheet.ncols > 3 else ""
        )
        c5_raw = sheet.cell_value(row_idx, 4) if sheet.ncols > 4 else ""
        c6_raw = sheet.cell_value(row_idx, 5) if sheet.ncols > 5 else ""
        qty = parse_quantity(c5_raw)
        if qty is None:
            qty = parse_quantity(c6_raw)
        unit = c4
        accent = ""
        try:
            xf = sheet.cell_xf_index(row_idx, 2) if sheet.ncols > 2 else 0
            bk = book.xf_list[xf]
            bg = getattr(getattr(bk, "background", None), "background_colour_index", None)
            if bg is not None and bg != 9:
                # xlrd palette — грубая эвристика для красных/жёлтых
                if bg in (10, 11, 12):
                    accent = "red"
                elif bg in (13, 14, 51, 52):
                    accent = "gold"
        except Exception:
            pass
        yield r, c1, name, unit, c5_raw, qty, accent


def parse_excel_estimate(file) -> ExcelParseResult:
    """
    Читает .xlsx / .xls и возвращает упорядоченные строки разделов и позиций.
    """
    name = (getattr(file, "name", "") or "").lower()
    if not name.endswith(".xlsx") and not (
        name.endswith(".xls") and not name.endswith(".xlsx")
    ):
        result = ExcelParseResult()
        result.errors.append("Поддерживаются только файлы .xlsx и .xls")
        return result

    if hasattr(file, "seek"):
        try:
            file.seek(0)
        except (OSError, AttributeError, TypeError):
            pass

    if name.endswith(".xlsx"):
        simple = _parse_simple_work_table_xlsx(file)
    else:
        simple = _parse_simple_work_table_xls(file)
    if simple and simple.rows:
        return simple

    if hasattr(file, "seek"):
        try:
            file.seek(0)
        except (OSError, AttributeError, TypeError):
            pass
    if name.endswith(".xlsx"):
        legacy = _parse_gectaro_export_format(file)
        if legacy.rows:
            return legacy

    if hasattr(file, "seek"):
        try:
            file.seek(0)
        except (OSError, AttributeError, TypeError):
            pass

    if name.endswith(".xlsx"):
        return _consume_boq_rows(_iter_xlsx_rows(file))
    return _consume_boq_rows(_iter_xls_rows(file))
