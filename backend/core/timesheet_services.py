"""Табель работников: бизнес-логика, Excel, аналитика."""

from __future__ import annotations

import calendar
from datetime import date, datetime
from io import BytesIO
from typing import Any

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import (
    Company,
    Employee,
    Timesheet,
    TimesheetEntry,
    TimesheetEntryLog,
    TimesheetMember,
)

TIMESHEET_PLACE_SITE = Timesheet.PLACE_SITE
TIMESHEET_PLACE_OFFICE = Timesheet.PLACE_OFFICE
VALID_TIMESHEET_PLACES = {TIMESHEET_PLACE_SITE, TIMESHEET_PLACE_OFFICE}


def normalize_timesheet_place(raw: str | None) -> str:
    p = (raw or "").strip().lower()
    if p in (TIMESHEET_PLACE_OFFICE, "офис", "office"):
        return TIMESHEET_PLACE_OFFICE
    return TIMESHEET_PLACE_SITE

User = get_user_model()

STATUS_META: list[dict[str, str]] = [
    {"code": TimesheetEntry.STATUS_PRESENT, "short": "Я", "label": "Явка", "color": "#22c55e"},
    {"code": TimesheetEntry.STATUS_OFF, "short": "В", "label": "Выходной", "color": "#94a3b8"},
    {"code": TimesheetEntry.STATUS_VACATION, "short": "О", "label": "Отпуск", "color": "#6366f1"},
    {"code": TimesheetEntry.STATUS_ABSENT, "short": "Н", "label": "Неявка", "color": "#ef4444"},
    {"code": TimesheetEntry.STATUS_HALF, "short": "П", "label": "Полдня", "color": "#f59e0b"},
]

STATUS_SHORT_BY_CODE = {m["code"]: m["short"] for m in STATUS_META}
STATUS_LABEL_BY_CODE = {m["code"]: m["label"] for m in STATUS_META}
VALID_STATUS_CODES = set(STATUS_SHORT_BY_CODE)


def month_bounds(year: int, month: int) -> tuple[date, date, int]:
    days = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, days), days


def get_or_create_timesheet(
    company: Company,
    year: int,
    month: int,
    place: str = TIMESHEET_PLACE_SITE,
) -> Timesheet:
    place = normalize_timesheet_place(place)
    ts, _ = Timesheet.objects.get_or_create(
        company=company,
        year=year,
        month=month,
        place=place,
    )
    return ts


def company_members_qs(
    company: Company,
    *,
    place: str = TIMESHEET_PLACE_SITE,
    brigade: str = "",
):
    place = normalize_timesheet_place(place)
    qs = (
        TimesheetMember.objects.filter(
            company=company,
            workplace=place,
            is_active=True,
            employee__status=Employee.STATUS_ACTIVE,
        )
        .select_related("employee")
        .order_by("employee__full_name", "employee_id")
    )
    if brigade:
        qs = qs.filter(employee__brigade__iexact=brigade.strip())
    return qs


def serialize_member(tm: TimesheetMember) -> dict[str, Any]:
    e = tm.employee
    return {
        "id": e.pk,
        "project_employee_id": tm.pk,
        "full_name": e.full_name,
        "position": e.position or "",
        "status": e.status,
        "status_display": e.get_status_display(),
        "workplace": tm.workplace,
    }


def entries_map_for_month(
    company: Company,
    year: int,
    month: int,
    place: str = TIMESHEET_PLACE_SITE,
) -> dict[str, str]:
    place = normalize_timesheet_place(place)
    start, end, _ = month_bounds(year, month)
    out: dict[str, str] = {}
    qs = TimesheetEntry.objects.filter(
        company=company,
        date__gte=start,
        date__lte=end,
        timesheet__place=place,
    ).values_list("employee_id", "date", "status")
    for eid, d, st in qs:
        out[f"{eid}:{d.isoformat()}"] = st or ""
    return out


def compute_analytics(
    company: Company,
    year: int,
    month: int,
    *,
    place: str = TIMESHEET_PLACE_SITE,
    today: date | None = None,
) -> dict[str, Any]:
    place = normalize_timesheet_place(place)
    today = today or date.today()
    members = list(company_members_qs(company, place=place))
    total_workers = len(members)
    emp_ids = [tm.employee_id for tm in members]

    on_site_today = 0
    absent_today = 0
    if emp_ids and today.year == year and today.month == month:
        today_entries = TimesheetEntry.objects.filter(
            company=company,
            employee_id__in=emp_ids,
            date=today,
            timesheet__place=place,
        )
        for ent in today_entries:
            if ent.status in (TimesheetEntry.STATUS_PRESENT, TimesheetEntry.STATUS_HALF):
                on_site_today += 1
            elif ent.status == TimesheetEntry.STATUS_ABSENT:
                absent_today += 1

    start, end, days_in_month = month_bounds(year, month)
    filled = TimesheetEntry.objects.filter(
        company=company,
        date__gte=start,
        date__lte=end,
        employee_id__in=emp_ids,
        timesheet__place=place,
        status__in=[TimesheetEntry.STATUS_PRESENT, TimesheetEntry.STATUS_HALF],
    ).count()
    possible = max(1, total_workers * days_in_month)
    attendance_pct = round(100.0 * filled / possible, 1)

    return {
        "total_workers": total_workers,
        "on_site_today": on_site_today,
        "absent_today": absent_today,
        "attendance_pct": attendance_pct,
    }


def log_entry_change(
    *,
    company: Company,
    employee: Employee,
    entry: TimesheetEntry | None,
    day: date,
    old_status: str,
    new_status: str,
    user: User | None,
) -> None:
    TimesheetEntryLog.objects.create(
        company=company,
        employee=employee,
        entry=entry,
        date=day,
        old_status=old_status or "",
        new_status=new_status or "",
        edited_by=user,
    )


@transaction.atomic
def upsert_cell(
    *,
    company: Company,
    employee_id: int,
    day: date,
    status: str,
    comment: str = "",
    user: User | None = None,
    place: str = TIMESHEET_PLACE_SITE,
) -> TimesheetEntry | None:
    place = normalize_timesheet_place(place)
    if day > date.today():
        raise ValueError("future date")
    if status and status not in VALID_STATUS_CODES:
        raise ValueError("invalid status")
    employee = Employee.objects.filter(
        pk=employee_id,
        company_id=company.pk,
    ).first()
    if not employee:
        raise ValueError("employee not found")
    if not TimesheetMember.objects.filter(
        company=company,
        employee=employee,
        workplace=place,
        is_active=True,
    ).exists():
        raise ValueError("employee not in timesheet")

    ts = get_or_create_timesheet(company, day.year, day.month, place)
    ent = TimesheetEntry.objects.filter(
        timesheet=ts, employee=employee, date=day
    ).first()
    old_status = ent.status if ent else ""

    if not status:
        if ent:
            log_entry_change(
                company=company,
                employee=employee,
                entry=ent,
                day=day,
                old_status=old_status,
                new_status="",
                user=user,
            )
            ent.delete()
        return None

    if ent:
        ent.status = status
        ent.comment = comment or ent.comment
        ent.edited_by = user
        ent.save(update_fields=["status", "comment", "edited_by", "edited_at"])
    else:
        ent = TimesheetEntry.objects.create(
            timesheet=ts,
            company=company,
            employee=employee,
            date=day,
            status=status,
            comment=comment or "",
            edited_by=user,
        )

    if old_status != status:
        log_entry_change(
            company=company,
            employee=employee,
            entry=ent,
            day=day,
            old_status=old_status,
            new_status=status,
            user=user,
        )
    return ent


@transaction.atomic
def bulk_upsert_cells(
    *,
    company: Company,
    updates: list[dict[str, Any]],
    user: User | None = None,
    place: str = TIMESHEET_PLACE_SITE,
) -> int:
    place = normalize_timesheet_place(place)
    count = 0
    for u in updates:
        eid = u.get("employee_id")
        ds = u.get("date")
        st = u.get("status", "")
        if not eid or not ds:
            continue
        try:
            day = datetime.strptime(str(ds)[:10], "%Y-%m-%d").date()
        except ValueError:
            continue
        upsert_cell(
            company=company,
            employee_id=int(eid),
            day=day,
            status=st or "",
            comment=str(u.get("comment") or ""),
            user=user,
            place=place,
        )
        count += 1
    return count


def export_timesheet_xlsx(
    company: Company,
    year: int,
    month: int,
    place: str = TIMESHEET_PLACE_SITE,
) -> BytesIO:
    place = normalize_timesheet_place(place)
    place_label = "Офис" if place == TIMESHEET_PLACE_OFFICE else "Объект"
    start, end, days = month_bounds(year, month)
    employees = [tm.employee for tm in company_members_qs(company, place=place)]
    entries = entries_map_for_month(company, year, month, place)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Табель {month:02d}.{year}"

    title = (
        f"ТАБЕЛЬ учёта рабочего времени ({place_label}) — "
        f"{company.name} — {month:02d}.{year}"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days + 2)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)

    headers = ["№", "ФИО", "Должность"] + [str(d) for d in range(1, days + 1)]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fill_map = {
        TimesheetEntry.STATUS_PRESENT: PatternFill("solid", fgColor="DCFCE7"),
        TimesheetEntry.STATUS_OFF: PatternFill("solid", fgColor="F1F5F9"),
        TimesheetEntry.STATUS_VACATION: PatternFill("solid", fgColor="E0E7FF"),
        TimesheetEntry.STATUS_ABSENT: PatternFill("solid", fgColor="FEE2E2"),
        TimesheetEntry.STATUS_HALF: PatternFill("solid", fgColor="FEF3C7"),
    }

    for i, emp in enumerate(employees, 1):
        row = 3 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=emp.full_name)
        ws.cell(row=row, column=3, value=emp.position or "")
        for d in range(1, days + 1):
            day = date(year, month, d)
            key = f"{emp.pk}:{day.isoformat()}"
            st = entries.get(key, "")
            col = 3 + d
            cell = ws.cell(
                row=row,
                column=col,
                value=STATUS_SHORT_BY_CODE.get(st, ""),
            )
            cell.alignment = Alignment(horizontal="center")
            if st in fill_map:
                cell.fill = fill_map[st]

    legend_row = 3 + len(employees) + 2
    ws.cell(row=legend_row, column=1, value="Условные обозначения:")
    for j, m in enumerate(STATUS_META):
        ws.cell(
            row=legend_row,
            column=2 + j,
            value=f"{m['short']} — {m['label']}",
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@transaction.atomic
def import_employees_from_xlsx(
    company: Company,
    file_obj,
    *,
    place: str = TIMESHEET_PLACE_SITE,
) -> dict[str, int]:
    place = normalize_timesheet_place(place)
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    created = 0
    linked = 0
    updated = 0

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row:
            continue
        full_name = str(row[0] or "").strip()
        if not full_name or full_name.lower() in ("фio", "фio", "фио", "fio"):
            continue
        position = str(row[1] or "").strip() if len(row) > 1 else ""

        emp = Employee.objects.filter(
            company=company, full_name__iexact=full_name
        ).first()
        if emp:
            changed = False
            if position and emp.position != position:
                emp.position = position
                changed = True
            if emp.status != Employee.STATUS_ACTIVE:
                emp.status = Employee.STATUS_ACTIVE
                changed = True
            if changed:
                emp.save()
                updated += 1
        else:
            emp = Employee.objects.create(
                company=company,
                full_name=full_name,
                position=position,
                status=Employee.STATUS_ACTIVE,
            )
            created += 1

        _, was_created = TimesheetMember.objects.get_or_create(
            company=company,
            employee=emp,
            workplace=place,
            defaults={"is_active": True},
        )
        if was_created:
            linked += 1
        elif not TimesheetMember.objects.filter(
            company=company,
            employee=emp,
            workplace=place,
            is_active=True,
        ).exists():
            TimesheetMember.objects.filter(
                company=company,
                employee=emp,
                workplace=place,
            ).update(is_active=True)
            linked += 1

    return {"created": created, "linked": linked, "updated": updated}


def _normalize_employee_status(status: str) -> str:
    s = (status or "").strip()
    if s in (Employee.STATUS_ACTIVE, Employee.STATUS_INACTIVE):
        return s
    return Employee.STATUS_ACTIVE


@transaction.atomic
def create_timesheet_member(
    company: Company,
    *,
    full_name: str,
    position: str = "",
    phone: str = "",
    brigade: str = "",
    status: str = Employee.STATUS_ACTIVE,
    place: str = TIMESHEET_PLACE_SITE,
) -> TimesheetMember:
    place = normalize_timesheet_place(place)
    full_name = (full_name or "").strip()
    if not full_name:
        raise ValueError("full_name")

    emp = Employee.objects.filter(
        company=company, full_name__iexact=full_name
    ).first()
    if emp:
        emp.position = (position or "").strip()
        emp.phone = (phone or "").strip()
        emp.brigade = (brigade or "").strip()
        emp.status = _normalize_employee_status(status)
        emp.save()
    else:
        emp = Employee.objects.create(
            company=company,
            full_name=full_name,
            position=(position or "").strip(),
            phone=(phone or "").strip(),
            brigade=(brigade or "").strip(),
            status=_normalize_employee_status(status),
        )

    tm, _ = TimesheetMember.objects.get_or_create(
        company=company,
        employee=emp,
        workplace=place,
        defaults={"is_active": True},
    )
    if not tm.is_active:
        tm.is_active = True
        tm.save(update_fields=["is_active"])
    return tm


@transaction.atomic
def update_timesheet_member(
    company: Company,
    employee_id: int,
    *,
    full_name: str | None = None,
    position: str | None = None,
    phone: str | None = None,
    brigade: str | None = None,
    status: str | None = None,
    place: str = TIMESHEET_PLACE_SITE,
) -> TimesheetMember:
    place = normalize_timesheet_place(place)
    tm = (
        TimesheetMember.objects.filter(
            company=company,
            employee_id=employee_id,
            workplace=place,
            is_active=True,
        )
        .select_related("employee")
        .first()
    )
    if not tm:
        raise ValueError("employee")

    emp = tm.employee
    if full_name is not None:
        name = full_name.strip()
        if not name:
            raise ValueError("full_name")
        dup = (
            Employee.objects.filter(company=company, full_name__iexact=name)
            .exclude(pk=emp.pk)
            .exists()
        )
        if dup:
            raise ValueError("duplicate_name")
        emp.full_name = name
    if position is not None:
        emp.position = position.strip()
    if phone is not None:
        emp.phone = phone.strip()
    if brigade is not None:
        emp.brigade = brigade.strip()
    if status is not None:
        emp.status = _normalize_employee_status(status)
    emp.save()
    return tm


@transaction.atomic
def remove_timesheet_member(
    company: Company,
    employee_id: int,
    *,
    place: str = TIMESHEET_PLACE_SITE,
) -> bool:
    place = normalize_timesheet_place(place)
    updated = TimesheetMember.objects.filter(
        company=company,
        employee_id=employee_id,
        workplace=place,
        is_active=True,
    ).update(is_active=False)
    if not updated:
        raise ValueError("employee")
    return True


def brigades_for_company(company: Company) -> list[str]:
    qs = (
        Employee.objects.filter(
            timesheet_memberships__company=company,
            timesheet_memberships__is_active=True,
            status=Employee.STATUS_ACTIVE,
        )
        .exclude(brigade="")
        .values_list("brigade", flat=True)
        .distinct()
        .order_by("brigade")
    )
    return list(qs)
